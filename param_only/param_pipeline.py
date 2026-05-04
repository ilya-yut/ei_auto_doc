#!/usr/bin/env python3
"""
Parameters-only EI doc pipeline.
Uses same input/ (Structure, Available fields, Code; optional _Parameters.docx). No Metadata.
Prepare: discover input/, write manifest and prompts for sections 03, 04, 05, 06, 07 (07 = structure only, no ABAP) to param_only/run/.
Verify: check 03/04/05/06/07 response files, 04 rules, 06 rules (use case param counts), and no duplicate `###`+`##` same title (params-only: keep `###` only).
Assemble: build one .md from 03+04+05+06+07 (dedupe headings per section), write param_only/output/Explanation_<basename>_params.md and .docx.
Run from repo root: python param_only/param_pipeline.py prepare [--skip-verify] [--yes] | verify | assemble
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

PARAM_ONLY_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = PARAM_ONLY_DIR.parent
INPUT_DIR = PROJECT_ROOT / "input"
OUTPUT_DIR = PARAM_ONLY_DIR / "output"
PROMPTS_DIR = PROJECT_ROOT / "prompts"
RUN_DIR = PARAM_ONLY_DIR / "run"
PIPELINE_DIR = PROJECT_ROOT / "scripts" / "pipeline"

SECTION_SPEC_PARAM = [
    ("03", "PROMPT_Parameters_Reference_Table_section.md", ["params"]),
    ("04", "PROMPT_Parameter_Configuration_Guidelines_section.md", ["structure", "params", "code", "params_docx"]),
    ("05", "PROMPT_Parameter_Relationships_section.md", ["structure", "params", "code", "params_docx"]),
    ("06", "PROMPT_Default_Values_and_Practical_Examples_section.md", ["structure", "params", "code"]),
    ("07", "07_prompt_str_only.md", ["structure"]),  # template in param_only; structure only, no ABAP
]

PREFIXES = {
    "params": ("Available fields_", "Available fields _"),
    "structure": ("Structure_", "Structure _"),
    "code": ("Code_", "Code _"),
}

# Regex: CALL FUNCTION '...' or "..." (ABAP). Capture the function name.
_CALL_FUNCTION_RE = re.compile(
    r"CALL\s+FUNCTION\s+['\"]([^'\"]+)['\"]",
    re.IGNORECASE,
)
# Regex: FUNCTION name (first line of ABAP FM, e.g. "  FUNCTION /SKN/F_SW_01_01_SM21.")
_FUNCTION_DECL_RE = re.compile(
    r"^\s*FUNCTION\s+(\S+)",
    re.IGNORECASE,
)


def _build_fm_to_code_path_map() -> dict[str, Path]:
    """Scan input/ and input/old for Code_*.txt; map FUNCTION name (normalized) -> file path."""
    fm_to_path: dict[str, Path] = {}
    for folder in (INPUT_DIR, INPUT_DIR / "old"):
        if not folder.exists():
            continue
        for path in list(folder.glob("Code_*.txt")) + list(folder.glob("Code _*.txt")):
            try:
                head = path.read_text(encoding="utf-8", errors="replace").splitlines()[:5]
            except OSError:
                continue
            for line in head:
                m = _FUNCTION_DECL_RE.match(line)
                if m:
                    name = m.group(1).strip().rstrip(".").upper()
                    if name and name not in fm_to_path:
                        fm_to_path[name] = path.resolve()
                    break
    return fm_to_path


def _find_called_function_code(main_code_path: Path) -> Path | None:
    """
    If the main code calls another function module (CALL FUNCTION '...') and that
    function's source exists in input/ or input/old, return its path; else None.
    """
    try:
        text = main_code_path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return None
    map_ = _build_fm_to_code_path_map()
    for m in _CALL_FUNCTION_RE.finditer(text):
        called = m.group(1).strip().upper()
        if not called:
            continue
        if called in map_:
            resolved = map_[called].resolve()
            main_resolved = main_code_path.resolve()
            if resolved != main_resolved:
                return map_[called]
    return None


def _stem_from_path(path: Path, prefixes: tuple[str, ...]) -> str:
    name = path.stem
    for prefix in prefixes:
        if name.startswith(prefix):
            return name[len(prefix) :].strip()
    return name


def _normalize_stem(s: str) -> str:
    return re.sub(r"[\s_]+", "_", s.strip().lower()).strip("_")


def _stems_similar(s1: str, s2: str) -> bool:
    return _normalize_stem(s1) == _normalize_stem(s2)


def _discover_inputs_param_only(assume_yes: bool = False) -> dict[str, Path] | None:
    """Find Code, Structure, Available fields (no Metadata). Optional _Parameters.docx."""
    code = list(dict.fromkeys(
        list(INPUT_DIR.glob("Code_*.txt")) + list(INPUT_DIR.glob("Code _*.txt"))
    ))
    struct = list(dict.fromkeys(
        list(INPUT_DIR.glob("Structure_*.xlsx")) + list(INPUT_DIR.glob("Structure _*.xlsx"))
    ))
    avail = list(dict.fromkeys(
        list(INPUT_DIR.glob("Available fields_*.xlsx")) + list(INPUT_DIR.glob("Available fields _*.xlsx"))
    ))
    params_docx = list(INPUT_DIR.glob("_Parameters.docx"))
    if not code or not struct or not avail:
        return None

    for candidates, key, type_name in [
        (avail, "params", "Available fields"),
        (struct, "structure", "Structure"),
        (code, "code", "Code"),
    ]:
        if len(candidates) > 1:
            stems = [_stem_from_path(p, PREFIXES[key]) for p in candidates]
            similar = len(set(stems)) == 1
            print(f"Multiple {type_name} files found (names {'match' if similar else 'differ'}):")
            for i, p in enumerate(candidates, 1):
                print(f"  {i}. {p.name}")
            print(f"  → Use first: {candidates[0].name}")
            if not assume_yes:
                try:
                    reply = input("  Proceed with first? [y/N]: ").strip().lower()
                except EOFError:
                    reply = "n"
                if reply not in ("y", "yes"):
                    return None
            if key == "params":
                avail = [candidates[0]]
            elif key == "structure":
                struct = [candidates[0]]
            else:
                code = [candidates[0]]

    return {
        "code": code[0].resolve(),
        "structure": struct[0].resolve(),
        "params": avail[0].resolve(),
        "params_docx": params_docx[0].resolve() if params_docx else None,
    }


def _verify_stem_match_param_only(paths: dict[str, Path], assume_yes: bool = False) -> list[str]:
    """Structure, params, code must share same stem."""
    errors = []
    stems = {}
    for key in ("params", "structure", "code"):
        p = paths.get(key)
        if p:
            stems[key] = _stem_from_path(p, PREFIXES[key])
    common = stems.get("structure") or stems.get("code")
    if not common:
        return ["Could not determine common stem from structure/code."]
    ok = (
        _stems_similar(stems.get("structure", ""), stems.get("params", ""))
        and _stems_similar(stems.get("structure", ""), stems.get("code", ""))
    )
    if not ok:
        if _stems_similar(stems.get("structure", ""), stems.get("params", "")) and _stems_similar(stems.get("structure", ""), stems.get("code", "")):
            ok = True
        else:
            errors.append(
                f"Structure, Available fields, and Code stems must match. "
                f"Found: Structure_{stems.get('structure', '?')} , Available fields_{stems.get('params', '?')} , Code_{stems.get('code', '?')}"
            )
    return errors


def _extract_structure_names_from_code(code_path: Path) -> list[str]:
    text = code_path.read_text(encoding="utf-8", errors="replace")
    pattern = re.compile(r"T_DATA\s+STRUCTURE\s+([/A-Za-z0-9_]+)", re.IGNORECASE)
    names = []
    for m in pattern.finditer(text):
        name = m.group(1).strip()
        if "/" in name:
            names.append(name)
    return list(dict.fromkeys(names))


def _structure_names_in_xlsx(struct_path: Path) -> set[str]:
    import openpyxl
    wb = openpyxl.load_workbook(struct_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=500, values_only=True))
    wb.close()
    if not rows:
        return set()
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx = None
    for i, h in enumerate(headers):
        if h and "structure" in h.lower() and "name" in h.lower():
            col_idx = i
            break
    if col_idx is None:
        col_idx = 0
    names = set()
    for row in rows[1:]:
        if row and len(row) > col_idx and row[col_idx] is not None:
            names.add(str(row[col_idx]).strip())
    return names


def verify_input(skip_verify: bool, assume_yes: bool) -> list[str]:
    if skip_verify:
        return []
    errors = []
    paths = _discover_inputs_param_only(assume_yes=assume_yes)
    if paths is None:
        return ["Could not discover input files. Need Code, Structure, Available fields in input/."]
    errors.extend(_verify_stem_match_param_only(paths, assume_yes=assume_yes))
    if errors:
        return errors
    try:
        import openpyxl
        wb = openpyxl.load_workbook(paths["structure"], read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
        wb.close()
        if not rows:
            errors.append("Structure file must have a sheet with Structure Name and at least one row.")
        else:
            headers = [str(c).strip() if c is not None else "" for c in rows[0]]
            if not any("structure" in h.lower() and "name" in h.lower() for h in headers) and not headers:
                errors.append("Structure file must have a sheet with Structure Name and at least one row.")
    except Exception as e:
        errors.append(f"Structure file unreadable: {e}")
    try:
        text = paths["code"].read_text(encoding="utf-8", errors="replace")
        if not text.strip():
            errors.append("Code file is empty or unreadable.")
    except Exception as e:
        errors.append(f"Code file is empty or unreadable: {e}")
    code_names = _extract_structure_names_from_code(paths["code"])
    struct_names = _structure_names_in_xlsx(paths["structure"])
    struct_names_upper = {s.upper() for s in struct_names}
    for name in code_names:
        if name.upper() not in struct_names_upper:
            errors.append(
                f"Structure file does not match code: T_DATA structure '{name}' is not listed in the Structure file."
            )
    return errors


_04_FORBIDDEN_PHRASES = [
    "output only",
    "output field",
    "not a filter",
    "used for output display only",
]
_04_FORBIDDEN_INTERNAL_NAMES = [
    "R_DATUM",
    "R_UDATE",
    "SY_DATLO",
    "SY_TIMLO",
    "DATE_FROM",
    "SY-DATUM",
]
_04_PARAMS_REQUIRING_OPTIONS = [
    "DATE_REF_FLD",
    "TIME_REF_FLD",
    "DURATION_UNIT",
    "TIME_DIFF_UNIT",
    "STATE_COLOR",
    "STATE_ICON",
    "WP_TYPE",
    "SLGMODE",
    "RQARCHSTAT",
    "TS",
    "STATUS",
]
_04_GENERIC_PHRASE_MAX_OCCURRENCES = [
    ("set to focus on", 5),
    ("set to narrow by", 5),
    ("when relevant to the data flow", 3),
    ("when needed", 5),
    ("when applicable", 3),
]


def _load_pipeline_core():
    """Reuse 03 expansion and 04 grouping checks from scripts/pipeline/pipeline.py."""
    import importlib.util

    path = PIPELINE_DIR / "pipeline.py"
    spec = importlib.util.spec_from_file_location("_ei_pipeline_core", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


_PIPELINE_CORE = None

# Params-only bundle: 03–07 are all `###` subsections (no 01/02). If both `### Title` and `## Title`
# appear, keep `###` and drop `##` (same rule as full pipeline sections 03–06).
_PARAM_ONLY_PREFER_H2_SECTIONS = frozenset()


def _param_duplicate_heading_issues(md: str, section_num: str) -> list[str]:
    """Detect back-to-back `### Title` and `## Title` with the same title."""
    prefer_h2 = section_num in _PARAM_ONLY_PREFER_H2_SECTIONS
    lines = md.splitlines()
    issues: list[str] = []
    i = 0
    while i < len(lines):
        s = lines[i].strip()
        if s.startswith("### ") and not s.startswith("####"):
            title_h3 = s[4:].strip()
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j < len(lines):
                s2 = lines[j].strip()
                if s2.startswith("## ") and not s2.startswith("###"):
                    title_h2 = s2[3:].strip()
                    if title_h3 == title_h2:
                        fix = (
                            "remove the `###` line (keep `##` only)"
                            if prefer_h2
                            else "remove the `##` line (keep `###` only)"
                        )
                        issues.append(
                            f"duplicate heading '{title_h3}' (~line {i + 1}: `###` then `##`); {fix}"
                        )
        i += 1
    return issues


def _param_dedupe_section_headings(md: str, section_num: str) -> str:
    """Resolve `### Title` + `## Title` duplicate; params-only keeps `###`."""
    prefer_h2 = section_num in _PARAM_ONLY_PREFER_H2_SECTIONS
    lines = md.splitlines(keepends=True)
    out: list[str] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        s = line.strip()
        if s.startswith("### ") and not s.startswith("####"):
            title_h3 = s[4:].strip()
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j < len(lines):
                s2 = lines[j].strip()
                if s2.startswith("## ") and not s2.startswith("###"):
                    title_h2 = s2[3:].strip()
                    if title_h3 == title_h2:
                        if prefer_h2:
                            i = j
                            continue
                        out.append(line)
                        i = j + 1
                        continue
        out.append(line)
        i += 1
    return "".join(out)


def _core():
    global _PIPELINE_CORE
    if _PIPELINE_CORE is None:
        _PIPELINE_CORE = _load_pipeline_core()
    return _PIPELINE_CORE


def verify_responses() -> list[str]:
    """Check 03/04/05 exist and 04 meets Parameter Configuration Guidelines rules."""
    errors = []
    for num, _, _ in SECTION_SPEC_PARAM:
        r = RUN_DIR / f"{num}_response.md"
        if not r.exists():
            errors.append(f"Missing response file: {r.name}")
    if errors:
        return errors

    for num, _, _ in SECTION_SPEC_PARAM:
        r = RUN_DIR / f"{num}_response.md"
        text = r.read_text(encoding="utf-8")
        for msg in _param_duplicate_heading_issues(text, num):
            errors.append(f"{r.name}: {msg}")

    r04 = RUN_DIR / "04_response.md"
    r03 = RUN_DIR / "03_response.md"
    if r04.exists():
        text04 = r04.read_text(encoding="utf-8")
        lines04 = text04.splitlines()
        for phrase in _04_FORBIDDEN_PHRASES:
            for i, line in enumerate(lines04, 1):
                if phrase.lower() in line.lower():
                    errors.append(f"04_response.md line {i}: forbidden phrase '{phrase}'")
                    break
        for name in _04_FORBIDDEN_INTERNAL_NAMES:
            for i, line in enumerate(lines04, 1):
                if name in line:
                    errors.append(
                        f"04_response.md line {i}: forbidden internal name '{name}' "
                        "(date/time params: business meaning only; no R_DATUM, SY_DATLO, etc.)"
                    )
                    break
        text04_lower = text04.lower()
        for phrase, max_occ in _04_GENERIC_PHRASE_MAX_OCCURRENCES:
            count = text04_lower.count(phrase.lower())
            if count > max_occ:
                errors.append(
                    f"04_response.md: generic phrase '{phrase}' appears {count} times (max {max_occ})."
                )
        if r03.exists():
            text03 = r03.read_text(encoding="utf-8")
            pc = _core()
            param_names_ordered, _row_count = pc._param_names_ordered_from_03_table(text03)
            expected_count = len(param_names_ordered)
            params_in_03 = set(param_names_ordered)
            for first, last in pc._serial_series_from_03_param_names(param_names_ordered):
                if not pc._serial_group_heading_present(text04, first, last):
                    gh = f"**{first} - {last}**"
                    gs = f"**{first}/{last}**"
                    errors.append(
                        f"04_response.md: serial-number series {first}..{last} must be grouped "
                        f'(e.g. "{gh}", "{gs}", or en-dash between names)'
                    )
            for stem1, stem2, lo, hi, suf in pc._PARALLEL_TAB_SLOT_GROUPS:
                a, b = pc._parallel_slot_param_names(stem1, stem2, lo, hi, suf)
                if not all(n in params_in_03 for n in a + b):
                    continue
                f1, l1 = a[0], a[-1]
                f2, l2 = b[0], b[-1]
                if not pc._parallel_tab12_heading_ok(text04, f1, l1, f2, l2):
                    ex = f"**{f1} - {l1} / {f2} - {l2}**"
                    ab = f"**{stem1}{lo}-{hi}{suf} / {stem2}{lo}-{hi}{suf}**"
                    errors.append(
                        f"04_response.md: parallel table parameters {f1}..{l1} and {f2}..{l2} must be grouped "
                        f'(e.g. "{ex}" or abbreviated "{ab}")'
                    )
            for param in _04_PARAMS_REQUIRING_OPTIONS:
                has_suffixed = any(re.match(rf"^{re.escape(param)}\d+$", p) for p in params_in_03)
                if has_suffixed:
                    if not re.search(rf"\*\*{re.escape(param)}\d+\s+Options:\*\*", text04):
                        errors.append(
                            f"04_response.md: parameters like {param}1 must have an Options subsection "
                            f'(e.g. "**{param}1 Options:**" per suffixed parameter)'
                        )
                elif param in params_in_03:
                    options_marker = f"**{param} Options:**"
                    if options_marker not in text04:
                        errors.append(
                            f"04_response.md: {param} must have an Options subsection "
                            f"(add '{options_marker}' with possible values)"
                        )
            match = re.search(r"ALL\s+(\d+)\s+parameters", text04, re.IGNORECASE)
            if match and expected_count > 0:
                n_in_04 = int(match.group(1))
                if n_in_04 != expected_count:
                    errors.append(
                        f"04_response.md IMPORTANT line says {n_in_04} parameters but 03 expands to {expected_count} parameters"
                    )
            elif expected_count > 0 and not re.search(r"IMPORTANT.*\d+.*parameters", text04, re.IGNORECASE):
                errors.append("04_response.md missing IMPORTANT line with parameter count (e.g. ALL N parameters)")

    # 06_response.md: Practical Configuration Examples – each use case >= 2 params, at least one use case 3–5 params (when EI has >= 3 params)
    r06 = RUN_DIR / "06_response.md"
    r03 = RUN_DIR / "03_response.md"
    if r06.exists():
        text06 = r06.read_text(encoding="utf-8")
        param_counts = _06_practical_example_param_counts(text06)
        for i, n in enumerate(param_counts):
            if n < 2:
                errors.append(
                    f"06_response.md: Use Case {i + 1} has {n} parameter(s) in its code block; "
                    "each use case must have at least 2 parameters."
                )
        # Require at least one use case with 3–5 params only when the EI has 3+ parameters in the reference table
        total_params = 0
        if r03.exists():
            names, _ = _core()._param_names_ordered_from_03_table(r03.read_text(encoding="utf-8"))
            total_params = len(names)
        if total_params >= 3 and param_counts and not any(3 <= n <= 5 for n in param_counts) and not any(n >= 6 for n in param_counts):
            if not any(n >= 3 for n in param_counts):
                errors.append(
                    "06_response.md: No use case has 3–5 (or more) parameters. "
                    "At least one practical configuration example must include 3–5 parameters in its code block."
                )

    return errors


def _06_practical_example_param_counts(text: str) -> list[int]:
    """Parse 06_response.md Practical Configuration Examples; return list of parameter counts per use case code block."""
    counts = []
    in_practical_section = False
    in_block = False
    for line in text.splitlines():
        if re.match(r"^###\s+Practical", line, re.IGNORECASE):
            in_practical_section = True
            continue
        if not in_practical_section:
            continue
        if line.strip().startswith("```"):
            if in_block:
                in_block = False
            else:
                in_block = True
                counts.append(0)
            continue
        if in_block:
            s = line.strip()
            if s and "=" in s and not s.startswith("#") and re.match(r"^[A-Za-z0-9_]+\s*=", s):
                counts[-1] += 1
    return counts


def _write_manifest_at(basename: str, title: str) -> None:
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    manifest = RUN_DIR / "manifest.txt"
    manifest.write_text(f"output_basename={basename}\ntitle={title}\n", encoding="utf-8")


def prepare(skip_verify: bool = False, assume_yes: bool = False) -> None:
    """Discover input (no Metadata), write manifest and 03/04/05/06/07 prompts to param_only/run/."""
    errs = verify_input(skip_verify=skip_verify, assume_yes=assume_yes)
    if errs:
        print("Verification failed:")
        for e in errs:
            print("  -", e)
        print("Fix the issues above or re-run with --skip-verify to ignore.")
        sys.exit(1)

    paths = _discover_inputs_param_only(assume_yes=assume_yes)
    if paths is None:
        print("Could not discover input files. Need Code, Structure, Available fields in input/.", file=sys.stderr)
        sys.exit(1)

    RUN_DIR.mkdir(parents=True, exist_ok=True)
    for num, _, _ in SECTION_SPEC_PARAM:
        r = RUN_DIR / f"{num}_response.md"
        if r.exists():
            r.unlink()

    stem = _stem_from_path(paths["structure"], PREFIXES["structure"])
    basename = stem
    title = f"Parameters: {basename}"
    _write_manifest_at(basename, title)

    called_code_path = _find_called_function_code(paths["code"])
    if called_code_path is not None:
        paths["called_code"] = called_code_path

    def _repl(s):
        return str(s).replace("\\", "\\\\")

    def replace_placeholders(text: str) -> str:
        text = re.sub(r"\[Provide the structure file path[^\]]*\]", _repl(paths["structure"]), text)
        text = re.sub(r"\[Provide the output structure / fields file path[^\]]*\]", _repl(paths["structure"]), text)
        text = re.sub(r"\[Provide the Parameters sheet path[^\]]*\]", _repl(paths["params"]), text)
        text = re.sub(r"\[Provide the file path or paste the parameters table content here\]", _repl(paths["params"]), text)
        text = re.sub(r"\[Provide the code file path[^\]]*\]", _repl(paths["code"]), text)
        text = re.sub(r"\[Provide the ABAP source\]", _repl(paths["code"]), text)
        if paths.get("called_code"):
            text = re.sub(
                r"\[Additional code \(called function\)[^\]]*\]",
                _repl(paths["called_code"]),
                text,
            )
        else:
            text = re.sub(
                r"\[Additional code \(called function\)[^\]]*\]",
                "Not provided.",
                text,
            )
        if paths.get("params_docx"):
            text = re.sub(
                r"Selected parameters file[^\n]*\n[^\n]*",
                "Selected parameters file: " + _repl(paths["params_docx"]) + "\n",
                text,
                count=1,
            )
        else:
            text = re.sub(
                r"Selected parameters file[^\n]*\n[^\n]*",
                "Selected parameters file: (optional – omit if not in input)\n",
                text,
                count=1,
            )
        return text

    for num, template_name, _ in SECTION_SPEC_PARAM:
        if num == "07":
            template_path = PARAM_ONLY_DIR / "07_prompt_str_only.md"
        else:
            template_path = PROMPTS_DIR / template_name
        if not template_path.exists():
            print(f"Missing prompt template: {template_path}", file=sys.stderr)
            sys.exit(1)
        text = template_path.read_text(encoding="utf-8")
        text = replace_placeholders(text)
        (RUN_DIR / f"{num}_prompt.txt").write_text(text, encoding="utf-8")

    print("Prepare done. Output basename:", basename)
    print("In Cursor, send the instruction from param_only/CURSOR_INSTRUCTION_PARAMS.txt")
    print("When 03/04/05/06/07 response files are in param_only/run/, run: python param_only/param_pipeline.py verify  (optional)")
    print("Then run: python param_only/param_pipeline.py assemble")


def assemble() -> None:
    """Read manifest and 03/04/05/06/07, build one .md, write param_only/output/Explanation_<basename>_params.md and .docx."""
    if not (RUN_DIR / "manifest.txt").exists():
        print("Run prepare first. No param_only/run/manifest.txt found.", file=sys.stderr)
        sys.exit(1)
    manifest = (RUN_DIR / "manifest.txt").read_text(encoding="utf-8")
    basename = ""
    title = ""
    for line in manifest.splitlines():
        if line.startswith("output_basename="):
            basename = line.split("=", 1)[1].strip()
        elif line.startswith("title="):
            title = line.split("=", 1)[1].strip()
    if not basename:
        print("manifest.txt missing output_basename=", file=sys.stderr)
        sys.exit(1)

    verr = verify_responses()
    if verr:
        print("Response verification failed:")
        for e in verr:
            print("  -", e)
        print("Fix the issues above (e.g. edit 04_response.md) then run assemble again.")
        sys.exit(1)

    parts = []
    for num, _, _ in SECTION_SPEC_PARAM:
        r = RUN_DIR / f"{num}_response.md"
        chunk = r.read_text(encoding="utf-8")
        chunk = _param_dedupe_section_headings(chunk, num)
        parts.append(chunk)

    full_md = f"# {title}\n\n" + "\n\n".join(parts)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_md = OUTPUT_DIR / f"Explanation_{basename}_params.md"
    out_md.write_text(full_md, encoding="utf-8")
    print("Wrote", out_md)

    out_docx = OUTPUT_DIR / f"Explanation_{basename}_params.docx"
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "md_to_docx",
            PIPELINE_DIR / "md_to_docx.py",
        )
        md2docx = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(md2docx)
        md2docx.convert_md_to_docx(out_md, out_docx)
        print("Wrote", out_docx)
    except ImportError as e:
        print("MD->DOCX converter unavailable (install markdown, beautifulsoup4, python-docx). Skipping .docx.", e)
    except Exception as e:
        print("MD->DOCX conversion failed. Markdown is ready.", e)


def main() -> None:
    parser = argparse.ArgumentParser(description="Parameters-only pipeline: prepare | verify | assemble")
    parser.add_argument("mode", choices=["prepare", "verify", "assemble"])
    parser.add_argument("--skip-verify", action="store_true", help="skip input verification when running prepare")
    parser.add_argument("--yes", "-y", action="store_true", help="assume yes for proceed prompts")
    args = parser.parse_args()
    if args.mode == "prepare":
        prepare(skip_verify=args.skip_verify, assume_yes=args.yes)
    elif args.mode == "verify":
        verr = verify_responses()
        if verr:
            print("Response verification failed:")
            for e in verr:
                print("  -", e)
            sys.exit(1)
        print("Response verification passed.")
    else:
        assemble()


if __name__ == "__main__":
    main()
