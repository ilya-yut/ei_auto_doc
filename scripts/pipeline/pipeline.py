#!/usr/bin/env python3
"""
EI Doc Pipeline (Option B, no API).
Verify: one of each input file (Metadata, Available fields, Structure, Code); stems Metadata/Structure/Code match; Available fields may match that or output name from metadata (B8/B9). If names only differ slightly (e.g. space), prompt user to proceed. Optional --yes to assume yes.
Prepare: clear run/, discover input/, read Metadata, write 7 prompts with paths injected.
Assemble: read 7 responses, build one .md, convert to .docx.
Verify: check response files (e.g. 03: Parameters table rows A–Z by Field; 01/02: no parameter or output-field names from 03/07 or global blocklist—business language only; 04: no "output only", parameter count match; unused parameters (not on output sheet and not used in ABAP) must have **`Not in use`** after the main description; used parameters must not; parameters in `run/unused_params.txt` must not appear in 05/06; when `input/params_dictionary.xlsx` lists a parameter, 04 main explanation must match that dictionary text exactly (`**Not in use**` line excluded from compare); optional `input/checked params.txt` toggles yellow `<mark>` wrapping for dictionary-only parameters (unchecked only; verify fails if checked dictionary params use `<mark>`; BACKDAYS/FORWDAYS exempt from duplicate mark when dictionary text equals the mandatory window sentence; FORWDAYS anchor lines mirror BACKDAYS per §3c); no duplicate `###`+`##` same title per section; 02: no **Training** subsection under Suggested Resolution; 06: blank line before **Purpose:**; no ranges on BACKDAYS/FORWDAYS/DURATION/DURATION_UNIT/DURATION_D/DURATION_H/DURATION_M in use cases; DURATION_UNIT=F requires Purpose "exactly N full days ago"). Run before or after generating responses; assemble runs verify automatically. `prepare --generate-037` sorts the Parameters sheet rows A–Z when building 03/04/07, marks unused parameters, writes `run/unused_params.txt`, and prefers dictionary text for 04 where available. MD→DOCX renders parameter **Options** lists and list items under Parameter Relationships / Default Values as plain paragraphs (no Word list bullets). Document title: `Exception Indicator: Name ( ID)`.
Run from repo root: python scripts/pipeline/pipeline.py prepare [--skip-verify] [--yes] [--generate-037] | verify | assemble
  --generate-037: after prepare, build 03_response.md, 04_response.md, and 07_response.md from input xlsx/code (no separate generator script).
"""
from __future__ import annotations

import argparse
import hashlib
import html
import re
import sys
from pathlib import Path

PIPELINE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = PIPELINE_DIR.parent.parent
INPUT_DIR = PROJECT_ROOT / "input"
OUTPUT_DIR = PROJECT_ROOT / "output"
PROMPTS_DIR = PROJECT_ROOT / "prompts"
RUN_DIR = PIPELINE_DIR / "run"
# Optional: SAP-oriented parameter explanations (Field name -> text). Used when generating 04 and when verifying 04.
PARAMS_DICTIONARY_PATH = INPUT_DIR / "params_dictionary.xlsx"
# Optional: one parameter name per line (A–Z field names). When present and non-empty, dictionary text for params
# listed in the dictionary but NOT in this file is wrapped in <mark>…</mark> (yellow in Word via md_to_docx).
CHECKED_PARAMS_PATH = INPUT_DIR / "checked params.txt"
UNUSED_PARAMS_RUN_FILE = RUN_DIR / "unused_params.txt"

sys.path.insert(0, str(PIPELINE_DIR))
from unused_params import (  # noqa: E402
    NOT_IN_USE_LINE,
    analyze_unused_params,
    format_unused_params_prompt_block,
    read_unused_params_file,
    write_unused_params_file,
)

UNUSED_PARAMS_PROMPT_PLACEHOLDER = "[UNUSED_PARAMS_PIPELINE_BLOCK]"

SECTION_SPEC = [
    ("01", "PROMPT_General_Overview_section.md", ["structure", "params", "code"]),
    ("02", "PROMPT_Problem_Description_and_Suggested_Resolution_section.md", ["structure", "params", "code"]),
    ("03", "PROMPT_Parameters_Reference_Table_section.md", ["params"]),
    ("04", "PROMPT_Parameter_Configuration_Guidelines_section.md", ["structure", "params", "code", "params_docx"]),
    ("05", "PROMPT_Parameter_Relationships_section.md", ["structure", "params", "code", "params_docx"]),
    ("06", "PROMPT_Default_Values_and_Practical_Examples_section.md", ["structure", "params", "code"]),
    ("07", "PROMPT_EI_Function_Structure_and_ABAP_Code_section.md", ["structure", "code"]),
]

# Option B: accept prefix with or without space before underscore (e.g. Metadata_ vs Metadata _)
PREFIXES = {
    "metadata": ("Metadata_", "Metadata _"),
    "params": ("Available fields_", "Available fields _"),
    "structure": ("Structure_", "Structure _"),
    "code": ("Code_", "Code _"),
}


def _stem_from_path(path: Path, prefixes: tuple[str, ...]) -> str:
    """Return the stem (name after prefix, without extension). Tries each prefix (e.g. Metadata_ then Metadata _)."""
    name = path.stem
    for prefix in prefixes:
        if name.startswith(prefix):
            return name[len(prefix) :].strip()
    return name


def _normalize_stem(s: str) -> str:
    """Normalize stem for similarity: spaces/underscores interchangeable, lower, stripped."""
    return re.sub(r"[\s_]+", "_", s.strip().lower()).strip("_")


def _stems_similar(s1: str, s2: str) -> bool:
    """True if stems are effectively the same (e.g. only space vs underscore difference)."""
    return _normalize_stem(s1) == _normalize_stem(s2)


def _discover_inputs(assume_yes: bool = False) -> dict[str, Path] | None:
    """Find one of each: Code, Structure, Available fields, Metadata. If multiple similar (e.g. space in prefix), ask user to proceed. Optional _Parameters.docx."""
    code = list(dict.fromkeys(
        list(INPUT_DIR.glob("Code_*.txt")) + list(INPUT_DIR.glob("Code _*.txt"))
    ))
    struct = list(dict.fromkeys(
        list(INPUT_DIR.glob("Structure_*.xlsx")) + list(INPUT_DIR.glob("Structure _*.xlsx"))
    ))
    avail = list(dict.fromkeys(
        list(INPUT_DIR.glob("Available fields_*.xlsx")) + list(INPUT_DIR.glob("Available fields _*.xlsx"))
    ))
    meta = list(dict.fromkeys(
        list(INPUT_DIR.glob("Metadata_*.xlsx")) + list(INPUT_DIR.glob("Metadata _*.xlsx"))
    ))
    params_docx = list(INPUT_DIR.glob("_Parameters.docx"))
    if not code or not struct or not avail or not meta:
        return None

    # If multiple files for any type, show and ask to proceed with first (similar names = e.g. space in prefix only)
    for candidates, key, type_name in [
        (meta, "metadata", "Metadata"),
        (avail, "params", "Available fields"),
        (struct, "structure", "Structure"),
        (code, "code", "Code"),
    ]:
        if len(candidates) > 1:
            stems = [_stem_from_path(p, PREFIXES[key]) for p in candidates]
            similar = len(set(stems)) == 1
            print(f"Multiple {type_name} files found (names {'match' if similar else 'differ'}):")
            for i, p in enumerate(candidates, 1):
                print(f"  {i}. {p.name}")
            print(f"  -> Use first: {candidates[0].name}")
            if not assume_yes:
                try:
                    reply = input("  Proceed with first? [y/N]: ").strip().lower()
                except EOFError:
                    reply = "n"
                if reply not in ("y", "yes"):
                    return None
            # use first
            if key == "metadata":
                meta = [candidates[0]]
            elif key == "params":
                avail = [candidates[0]]
            elif key == "structure":
                struct = [candidates[0]]
            else:
                code = [candidates[0]]

    return {
        "code": code[0].resolve(),
        "structure": struct[0].resolve(),
        "params": avail[0].resolve(),
        "metadata": meta[0].resolve(),
        "params_docx": params_docx[0].resolve() if params_docx else None,
    }


def _metadata_sheet(wb):
    """Use sheet with General metadata: 'General', 'Metadata general', or first sheet."""
    for name in ("General", "Metadata general"):
        if name in wb.sheetnames:
            return wb[name]
    return wb.worksheets[0] if wb.worksheets else wb.active


def _read_metadata_title(metadata_path: Path) -> tuple[str, str]:
    """Read Exception indicator ID (B8) and name (B9) from Metadata. Prefer sheet 'General' if present."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(metadata_path, read_only=True)
        ws = _metadata_sheet(wb)
        rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
        wb.close()
        if len(rows) >= 9 and len(rows[8]) > 1 and len(rows[7]) > 1:
            return (str(rows[7][1] or "").strip(), str(rows[8][1] or "").strip())
    except Exception:
        pass
    return ("", "")


def _verify_stem_match(
    paths: dict[str, Path],
    metadata_id: str = "",
    metadata_name: str = "",
    assume_yes: bool = False,
) -> list[str]:
    """Metadata, Structure, Code must share same stem. Available fields may match that stem or output name (from metadata contents). If names only differ slightly (e.g. space), ask user to proceed."""
    errors = []
    stems = {}
    for key in ("metadata", "params", "structure", "code"):
        p = paths.get(key)
        if p:
            stems[key] = _stem_from_path(p, PREFIXES[key])

    common_stem = stems.get("metadata") or stems.get("structure") or stems.get("code")
    meta_struct_code_ok = (
        stems.get("metadata") == stems.get("structure") == stems.get("code")
    )

    # Available fields stem can match common_stem OR output name from metadata (B8/B9): name, ID, or "name_ID" / "name - ID"
    params_stem = stems.get("params", "")
    combined_underscore = "_".join(filter(None, [metadata_name.strip(), metadata_id.strip()]))
    combined_dash = " - ".join(filter(None, [metadata_name.strip(), metadata_id.strip()]))
    params_ok = (
        params_stem == common_stem
        or params_stem == metadata_name
        or params_stem == metadata_id
        or params_stem == combined_underscore
        or params_stem == combined_dash
        or _stems_similar(params_stem, common_stem or "")
        or _stems_similar(params_stem, metadata_name)
        or _stems_similar(params_stem, metadata_id)
        or _stems_similar(params_stem, combined_underscore)
        or _stems_similar(params_stem, combined_dash)
    )

    if not meta_struct_code_ok:
        # Metadata / Structure / Code don't match; check if very similar
        if common_stem and _stems_similar(stems.get("metadata", ""), stems.get("structure", "")) and _stems_similar(stems.get("structure", ""), stems.get("code", "")):
            print("File names differ slightly (e.g. space vs underscore):")
            print(f"  Metadata: {stems.get('metadata', '?')} | Structure: {stems.get('structure', '?')} | Code: {stems.get('code', '?')}")
            if not assume_yes:
                try:
                    reply = input("  Proceed anyway? [y/N]: ").strip().lower()
                except EOFError:
                    reply = "n"
                if reply in ("y", "yes"):
                    params_ok = True if _stems_similar(params_stem, common_stem or "") or params_stem == common_stem else params_ok
                else:
                    errors.append(
                        "Metadata, Structure, and Code file names (after prefix) must match. "
                        f"Found: {stems.get('metadata', '?')} , {stems.get('structure', '?')} , {stems.get('code', '?')} ."
                    )
            else:
                params_ok = True
        else:
            errors.append(
                "Metadata, Structure, and Code file names (after prefix) must match. "
                f"Found: Metadata_{stems.get('metadata', '?')} , Structure_{stems.get('structure', '?')} , Code_{stems.get('code', '?')} . "
                "Available fields may match that name or the output name from metadata (B8/B9)."
            )
    elif not params_ok:
        # Params stem differs; check if very similar
        if _stems_similar(params_stem, common_stem or "") or _stems_similar(params_stem, metadata_name) or _stems_similar(params_stem, metadata_id):
            print("Available fields file name differs slightly from Metadata/Structure/Code (or from output name in metadata).")
            print(f"  Available fields: {params_stem} | others: {common_stem} | metadata name/ID: {metadata_name!r} / {metadata_id!r}")
            if not assume_yes:
                try:
                    reply = input("  Proceed anyway? [y/N]: ").strip().lower()
                except EOFError:
                    reply = "n"
                if reply in ("y", "yes"):
                    params_ok = True
            else:
                params_ok = True
        if not params_ok:
            errors.append(
                "Available fields file name (after prefix) should match Metadata/Structure/Code or the output name from metadata (B8/B9). "
                f"Found: Available fields_{params_stem} ; others: {common_stem} ; metadata name/ID: {metadata_name!r} / {metadata_id!r} ."
            )

    return errors


def _extract_structure_names_from_code(code_path: Path) -> list[str]:
    """Extract only the T_DATA structure name from ABAP (the main output structure). Other STRUCTURE refs (e.g. INCLUDE) are ignored."""
    text = code_path.read_text(encoding="utf-8", errors="replace")
    # T_DATA STRUCTURE /SKN/... or T_DATA STRUCTURE  /SKN/...
    pattern = re.compile(r"T_DATA\s+STRUCTURE\s+([/A-Za-z0-9_]+)", re.IGNORECASE)
    names = []
    for m in pattern.finditer(text):
        name = m.group(1).strip()
        if "/" in name:
            names.append(name)
    return list(dict.fromkeys(names))


def _structure_names_in_xlsx(struct_path: Path) -> set[str]:
    """Read 'Structure Name' column (unique values) from first sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(struct_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=500, values_only=True))
    wb.close()
    if not rows:
        return set()
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_idx = None
    for i, h in enumerate(headers):
        if h and "structure" in h.lower() and "name" in h.lower():
            col_idx = i
            break
    if col_idx is None:
        col_idx = 0
    names = set()
    for row in rows[1:]:
        if row and len(row) > col_idx and row[col_idx] is not None:
            names.add(str(row[col_idx]).strip())
    return names


def verify(skip_verify: bool = False, assume_yes: bool = False) -> list[str]:
    """Run all verification checks. Return list of error strings (empty if OK)."""
    if skip_verify:
        return []
    errors = []

    # Presence and cardinality (may prompt if multiple similar files)
    paths = _discover_inputs(assume_yes=assume_yes)
    if paths is None:
        code = list(INPUT_DIR.glob("Code_*.txt")) + list(INPUT_DIR.glob("Code _*.txt"))
        struct = list(INPUT_DIR.glob("Structure_*.xlsx")) + list(INPUT_DIR.glob("Structure _*.xlsx"))
        avail = list(INPUT_DIR.glob("Available fields_*.xlsx")) + list(INPUT_DIR.glob("Available fields _*.xlsx"))
        meta = list(INPUT_DIR.glob("Metadata_*.xlsx")) + list(INPUT_DIR.glob("Metadata _*.xlsx"))
        if not meta:
            errors.append("Missing: exactly one Metadata_*.xlsx in input/")
        elif len(meta) > 1:
            errors.append("Multiple Metadata_*.xlsx found; use one function set per run.")
        if not avail:
            errors.append("Missing: exactly one Available fields_*.xlsx in input/")
        elif len(avail) > 1:
            errors.append("Multiple Available fields_*.xlsx found; use one function set per run.")
        if not struct:
            errors.append("Missing: exactly one Structure_*.xlsx in input/")
        elif len(struct) > 1:
            errors.append("Multiple Structure_*.xlsx found; use one function set per run.")
        if not code:
            errors.append("Missing: exactly one Code_*.txt in input/")
        elif len(code) > 1:
            errors.append("Multiple Code_*.txt found; use one function set per run.")
        return errors

    # Stem match: Metadata/Structure/Code must match; Available fields may match that or output name from metadata
    metadata_id, metadata_name = _read_metadata_title(paths["metadata"])
    errors.extend(
        _verify_stem_match(paths, metadata_id, metadata_name, assume_yes=assume_yes)
    )

    # Per-file checks
    try:
        import openpyxl
        wb = openpyxl.load_workbook(paths["metadata"], read_only=True)
        ws = _metadata_sheet(wb)
        rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
        wb.close()
        if len(rows) < 9:
            errors.append("Metadata file must have Exception indicator ID (row 8) and Exception indicator name (row 9) in the General section.")
        else:
            _ = str(rows[7][1] or "").strip()
            _ = str(rows[8][1] or "").strip()
    except Exception as e:
        errors.append(f"Metadata file unreadable or missing required cells: {e}")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(paths["params"], read_only=True)
        if "Parameters" not in wb.sheetnames:
            errors.append("Available fields file must contain a 'Parameters' sheet with at least one parameter row.")
        else:
            ws = wb["Parameters"]
            rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
            wb.close()
            if len(rows) < 2:
                errors.append("Available fields file must contain a 'Parameters' sheet with at least one parameter row.")
    except Exception as e:
        errors.append(f"Available fields file unreadable: {e}")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(paths["structure"], read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
        wb.close()
        if not rows:
            errors.append("Structure file must have a sheet with Structure Name and at least one row.")
        else:
            headers = [str(c).strip() if c is not None else "" for c in rows[0]]
            if not any("structure" in h.lower() and "name" in h.lower() for h in headers) and not headers:
                errors.append("Structure file must have a sheet with Structure Name and at least one row.")
    except Exception as e:
        errors.append(f"Structure file unreadable: {e}")

    try:
        text = paths["code"].read_text(encoding="utf-8", errors="replace")
        if not text.strip():
            errors.append("Code file is empty or unreadable.")
    except Exception as e:
        errors.append(f"Code file is empty or unreadable: {e}")

    # T_DATA structure from code must be present in Structure xlsx (case-insensitive)
    code_names = _extract_structure_names_from_code(paths["code"])
    struct_names = _structure_names_in_xlsx(paths["structure"])
    struct_names_upper = {s.upper() for s in struct_names}
    for name in code_names:
        if name.upper() not in struct_names_upper:
            errors.append(
                f"Structure file does not match code: T_DATA structure '{name}' is not listed in the Structure file."
            )

    return errors


# Forbidden phrases in Parameter Configuration Guidelines (04); rule: do not classify as "output only" / "not a filter"
_04_FORBIDDEN_PHRASES = [
    "output only",
    "output field",
    "not a filter",
    "used for output display only",
]
# Date/time rule: 04 must not name internal variables (prompt: business meaning only)
_04_FORBIDDEN_INTERNAL_NAMES = [
    "R_DATUM",
    "R_UDATE",
    "SY_DATLO",
    "SY_TIMLO",
    "DATE_FROM",
    "SY-DATUM",
]
# Fixed-option parameters: when present in 03, 04 must have a "[PARAM_NAME] Options:" subsection
_04_PARAMS_REQUIRING_OPTIONS = [
    "DATE_REF_FLD",
    "TIME_REF_FLD",
    "DURATION_UNIT",
    "TIME_DIFF_UNIT",  # same unit semantics as duration when present without DURATION_UNIT
]
# Additional globally enforced option-bearing parameters (code-first extraction).
_04_EXTRA_OPTIONS_PARAMS = [
    "MANAGE_IN_UTC",
    "STATE_COLOR",
    "STATUS",
    "QSTATE",
    "AGGR_LEVEL",
    "AGGLEVEL",
    # USER_FLD / USR_FLD: dedicated rules + verify (_04_user_fld_guideline_issues), not generic Options.
]
# Status selector family: each present parameter must have options documented.
_04_STATUS_SELECTOR_PARAMS = [
    "STAT_WAIT",
    "STAT_TRANSIT",
    "STAT_OK",
    "STAT_ERROR",
    "STAT_INCONS",
    "STAT_FUTURE",
    "STAT_RETRY",
    "STAT_DIRECT",
    "STAT_ACTIVE",
]
# Debit/credit style indicators: when present in 03, 04 must document **PARAM Options:** (see prompts).
_04_DEBIT_CREDIT_INDICATOR_PARAM_NAMES = frozenset(
    {
        "SHKZG",
        "BEKNZ",
        "KZZUAB",
        "SHKZU",
        "DRCRK",  # rare naming
    }
)
# Generic/filler phrases: if a phrase appears more than max_occurrences times in 04, verification fails (principle: no useless repetition)
_04_GENERIC_PHRASE_MAX_OCCURRENCES = [
    ("set to focus on", 5),
    ("set to narrow by", 5),
    ("when relevant to the data flow", 3),
    ("when needed", 5),
    ("when applicable", 3),
    ("refines scheduling/change/execution timestamp scope used for job-event selection", 2),
]
# Placeholder-like sentence templates that must not appear in 04 main explanations.
_04_FORBIDDEN_TEMPLATE_PATTERNS = [
    re.compile(
        r"in this .* monitoring function,\s*\*\*[A-Za-z0-9_/\-\s]+\*\*\s+is used with its business meaning",
        re.IGNORECASE,
    ),
]

# Prompt 06: BACKDAYS / DURATION / DURATION_UNIT / AGGLEVEL must use the standard
# "initial - treated as … by code" bullet shape when listed in 03 (no long prose).
_DEFAULT_VALUES_INITIAL_RUNTIME_PARAMS = frozenset(
    {"BACKDAYS", "DURATION", "DURATION_UNIT", "AGGLEVEL"}
)

# Substrings that must not appear in the effect tail (after initial -) for those parameters (too vague).
_06_VAGUE_INITIAL_EFFECT_SUBSTRINGS = (
    "framework applies its standard",
    "standard lookback",
    "falls back to its template default",
)

# Minimum length of the effect tail: "treated as … by code".
_06_INITIAL_RUNTIME_EFFECT_MIN_LEN = 20

# Required shape: treated as <value> by code (optional short trailing clause; no parentheses).
_06_INITIAL_RUNTIME_TREATED_AS_LINE = re.compile(
    r"(?is)^treated\s+as\s+.+\bby\s+code\b(?:\s+[^();]+)?\.?\s*$",
)

# Practical use cases: these parameters must use a single value, never low - high.
_06_NO_RANGE_PARAMS = frozenset(
    {
        "BACKDAYS",
        "FORWDAYS",
        "DURATION",
        "DURATION_UNIT",
        "DURATION_D",
        "DURATION_H",
        "DURATION_M",
    }
)
_06_PARAM_ASSIGNMENT_LINE_RE = re.compile(r"^([A-Za-z0-9_]+)\s*=\s*(.+)$")
_06_VALUE_LOOKS_LIKE_RANGE_RE = re.compile(r"[\u2013\u2014\-]\s*\S")


def _06_iter_use_case_blocks(text06: str) -> list[dict[str, str | int | list[str]]]:
    """Yield use-case chunks from the Practical Example section: num, purpose, code lines."""
    chunk = _06_practical_section_text(text06)
    if not chunk.strip():
        return []
    pat = re.compile(r"(?m)^\*\*Use Case\s+(\d+)\s*:[^\n]*\*\*[ \t]*$")
    matches = list(pat.finditer(chunk))
    blocks: list[dict[str, str | int | list[str]]] = []
    for i, m in enumerate(matches):
        body_start = m.end()
        body_end = matches[i + 1].start() if i + 1 < len(matches) else len(chunk)
        body = chunk[body_start:body_end]
        fence_pos = body.find("```")
        if fence_pos == -1:
            continue
        before_fence = body[:fence_pos]
        after_first_fence = body[fence_pos + 3 :]
        fence_end = after_first_fence.find("```")
        code_body = after_first_fence[:fence_end] if fence_end >= 0 else after_first_fence
        purpose = ""
        pm = re.search(r"\*\*Purpose:\*\*\s*(.+)", before_fence, re.DOTALL)
        if pm:
            purpose = pm.group(1).strip()
        code_lines = [
            ln.strip()
            for ln in code_body.splitlines()
            if ln.strip() and "=" in ln and re.match(r"^[A-Za-z0-9_]+\s*=", ln.strip())
        ]
        blocks.append(
            {
                "num": int(m.group(1)),
                "purpose": purpose,
                "code_lines": code_lines,
            }
        )
    return blocks


def _06_no_range_time_param_issues(text06: str) -> list[str]:
    """Use-case code blocks must not use ranges on time/duration selection parameters."""
    issues: list[str] = []
    for block in _06_iter_use_case_blocks(text06):
        num = block["num"]
        for line in block["code_lines"]:
            m = _06_PARAM_ASSIGNMENT_LINE_RE.match(line)
            if not m:
                continue
            param = m.group(1).upper()
            if param not in _06_NO_RANGE_PARAMS:
                continue
            value = m.group(2).strip()
            if _06_VALUE_LOOKS_LIKE_RANGE_RE.search(value):
                issues.append(
                    f"06_response.md: Use Case {num}: **{param}** must be a single value, "
                    f"not a range (found `{line}`)."
                )
    return issues


def _06_duration_unit_f_purpose_issues(
    text06: str, params_in_03: set[str], unused: set[str] | None = None
) -> list[str]:
    """When DURATION_UNIT and DURATION are in Parameters: require F example + exactly-N-days Purpose."""
    skip = unused or set()
    active = {p for p in params_in_03 if p.upper() not in skip}
    if "DURATION_UNIT" not in active or "DURATION" not in active:
        return []
    issues: list[str] = []
    blocks = _06_iter_use_case_blocks(text06)
    has_f_example = False
    for block in blocks:
        num = block["num"]
        purpose = str(block["purpose"])
        dur_val: str | None = None
        unit_f = False
        for line in block["code_lines"]:
            um = re.match(r"^DURATION_UNIT\s*=\s*(\S+)\s*$", line, re.IGNORECASE)
            if um and um.group(1).upper() == "F":
                unit_f = True
            dm = re.match(r"^DURATION\s*=\s*(\d+)\s*$", line, re.IGNORECASE)
            if dm:
                dur_val = dm.group(1)
        if not unit_f:
            continue
        has_f_example = True
        if dur_val is None:
            issues.append(
                f"06_response.md: Use Case {num}: with DURATION_UNIT = F, **DURATION** must be a "
                "single positive integer (e.g. DURATION = 7), not a range."
            )
            continue
        if not re.search(
            rf"(?i)exactly\s+{re.escape(dur_val)}\s+(?:full\s+)?days?\s+ago",
            purpose,
        ):
            issues.append(
                f"06_response.md: Use Case {num}: **Purpose:** must state that the scope is "
                f"exactly {dur_val} full days ago when DURATION_UNIT = F and DURATION = {dur_val}."
            )
    if not has_f_example:
        issues.append(
            "06_response.md: at least one use case must include DURATION_UNIT = F with a "
            "single-value DURATION when DURATION_UNIT and DURATION are in the Parameters table."
        )
    return issues


def _06_default_values_section_chunk(text06: str) -> str:
    """Markdown between ### Default Values and ### Practical Example of Parameter Configuration."""
    m = re.search(r"^###\s+Default Values\s*$", text06, re.MULTILINE)
    if not m:
        return ""
    chunk = text06[m.end() :]
    m2 = re.search(r"^###\s+Practical Example of Parameter Configuration\s*$", chunk, re.MULTILINE)
    if m2:
        chunk = chunk[: m2.start()]
    return chunk


def _06_default_values_covers_param(dv_chunk: str, param: str) -> bool:
    """True if Default Values has a bullet for param (explicit value or initial-runtime effect)."""
    for line in dv_chunk.splitlines():
        s = line.strip()
        mat = re.match(rf"^\-\s+\*\*{re.escape(param)}\*\*\s+\-\s+(.+)$", s)
        if not mat:
            continue
        rest = mat.group(1).strip()
        if rest.lower().startswith("initial"):
            # "initial - treated …" or legacy "initial — …" (en/em/hyphen after initial)
            if re.match(r"(?i)^initial\s*[\u2013\u2014-]\s*\S", rest):
                return True
        else:
            return True
    return False


def _06_initial_runtime_default_bullets_issues(
    text03: str, text06: str, unused: set[str] | None = None
) -> list[str]:
    """Enforce prompt 06: BACKDAYS, DURATION, DURATION_UNIT, AGGLEVEL when in 03 (active only)."""
    ordered, _ = _param_names_ordered_from_03_table(text03)
    active = [p for p in ordered if p.upper() not in (unused or set())]
    needed = _DEFAULT_VALUES_INITIAL_RUNTIME_PARAMS.intersection(set(active))
    if not needed:
        return []
    dv = _06_default_values_section_chunk(text06)
    if not dv.strip():
        return [
            "06_response.md: Default Values must include initial-runtime bullets for parameters present in 03 "
            f"({', '.join(sorted(needed))}). See prompts/PROMPT_Default_Values_and_Practical_Examples_section.md."
        ]
    issues: list[str] = []
    for p in sorted(needed):
        if not _06_default_values_covers_param(dv, p):
            issues.append(
                f"06_response.md: Default Values must document **{p}** (explicit default or "
                f"'- **{p}** - initial - treated as … by code') when {p} appears in 03; see prompt 06."
            )
    return issues


def _06_initial_effect_text_for_param(dv_chunk: str, param: str) -> str | None:
    """Return text after initial + dash (hyphen/en/em) for param's bullet, or None if not initial-runtime."""
    for line in dv_chunk.splitlines():
        s = line.strip()
        mat = re.match(rf"^\-\s+\*\*{re.escape(param)}\*\*\s+\-\s+(.+)$", s)
        if not mat:
            continue
        rest = mat.group(1).strip()
        m2 = re.match(r"(?i)^initial\s*[\u2013\u2014-]\s*(.+)$", rest)
        if m2:
            return m2.group(1).strip()
    return None


def _06_initial_runtime_effect_clarity_issues(
    text03: str, text06: str, unused: set[str] | None = None
) -> list[str]:
    """
    Enforce standard shape: `- **PARAM** - initial - treated as … by code` (concise; no parentheses).
    """
    ordered, _ = _param_names_ordered_from_03_table(text03)
    active = [p for p in ordered if p.upper() not in (unused or set())]
    needed = _DEFAULT_VALUES_INITIAL_RUNTIME_PARAMS.intersection(set(active))
    if not needed:
        return []
    dv = _06_default_values_section_chunk(text06)
    issues: list[str] = []
    for p in sorted(needed):
        effect = _06_initial_effect_text_for_param(dv, p)
        if effect is None:
            continue
        el = effect.lower()
        if not _06_INITIAL_RUNTIME_TREATED_AS_LINE.match(effect.strip()):
            issues.append(
                f"06_response.md: **{p}** Default Values bullet must match "
                f"`- **{p}** - initial - treated as <value> by code` "
                "(hyphen, en dash, or em dash allowed between `initial` and `treated`; "
                "no parentheses or long prose after `by code`). "
                "See prompts/PROMPT_Default_Values_and_Practical_Examples_section.md."
            )
            continue
        if len(effect.strip()) < _06_INITIAL_RUNTIME_EFFECT_MIN_LEN:
            issues.append(
                f"06_response.md: **{p}** treated-as line is too short (min {_06_INITIAL_RUNTIME_EFFECT_MIN_LEN} "
                "characters). See prompts/PROMPT_Default_Values_and_Practical_Examples_section.md."
            )
            continue
        if "(" in effect or ")" in effect:
            issues.append(
                f"06_response.md: **{p}** Default Values bullet must not use parentheses in the "
                "`treated as … by code` tail; keep a single short clause only."
            )
            continue
        if len(effect.strip()) > 120:
            issues.append(
                f"06_response.md: **{p}** Default Values bullet is too long; keep the "
                "`initial - treated as … by code` tail under ~120 characters with no wordy explanation."
            )
            continue
        vague_hit = next((frag for frag in _06_VAGUE_INITIAL_EFFECT_SUBSTRINGS if frag in el), None)
        if vague_hit is not None:
            issues.append(
                f"06_response.md: **{p}** initial-runtime text must not use vague wording ({vague_hit!r})."
            )
    return issues


def _expand_03_parameter_cell(cell: str) -> list[str]:
    """
    Turn one Parameters table Field cell into ordered SAP parameter names.
    Supports: 'P1 / P2'; ranges 'TAB1_ATTR1-5'; 'TAB1_ATTR1_V-5_V' (same as TAB1_ATTR1_V through TAB1_ATTR5_V).
    """
    cell = cell.strip().strip("`")
    if not cell:
        return []
    out: list[str] = []
    for raw_part in re.split(r"\s*/\s*", cell):
        part = raw_part.strip()
        if not part:
            continue
        m_v = re.match(r"^(.+?)(\d+)(_V)-(\d+)_V$", part, re.IGNORECASE)
        if m_v:
            stem, lo, hi = m_v.group(1), int(m_v.group(2)), int(m_v.group(4))
            if lo <= hi:
                for k in range(lo, hi + 1):
                    out.append(f"{stem}{k}_V")
            continue
        m = re.match(r"^(.+?)(\d+)-(\d+)$", part)
        if m:
            stem, lo, hi = m.group(1), int(m.group(2)), int(m.group(3))
            if lo <= hi:
                for k in range(lo, hi + 1):
                    out.append(f"{stem}{k}")
            continue
        out.append(part)
    return out


def _param_names_ordered_from_03_table(text03: str) -> tuple[list[str], int]:
    """
    Parse 03 markdown table rows. Returns (flattened parameter names in row order, number of table rows).
    """
    param_names_ordered: list[str] = []
    param_rows = [l for l in text03.splitlines() if re.match(r"^\|\s*\d+\s*\|", l)]
    for row in param_rows:
        parts = [p.strip() for p in row.split("|")]
        if len(parts) < 3:
            continue
        param_names_ordered.extend(_expand_03_parameter_cell(parts[2]))
    return param_names_ordered, len(param_rows)


def _03_field_cells_in_table_order(text03: str) -> list[str]:
    """Field column values from numbered Parameters Reference Table rows (one entry per table row)."""
    names: list[str] = []
    for line in text03.splitlines():
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 4:
            continue
        if not parts[1].isdigit():
            continue
        fld = parts[2].strip().strip("`")
        if fld.lower() in ("field", "#"):
            continue
        names.append(fld)
    return names


def _03_parameters_sorted_alphabetically_issues(text03: str) -> list[str]:
    """Parameters Reference Table rows must be sorted A–Z by Field (case-insensitive)."""
    names = _03_field_cells_in_table_order(text03)
    if len(names) < 2:
        return []
    for i in range(len(names) - 1):
        if names[i].lower() > names[i + 1].lower():
            return [
                "03_response.md: Parameters Reference Table must list fields in alphabetical order (A–Z) by "
                f"Field; found {names[i]!r} before {names[i + 1]!r}."
            ]
    return []


def _04_debit_credit_indicator_params_from_03(text03: str, params_in_03: set[str]) -> set[str]:
    """
    Names of parameters treated as debit/credit indicators for Options documentation.
    Uses a canonical name list plus the Parameters Reference Table Description column in 03.
    """
    out: set[str] = set()
    for row in text03.splitlines():
        parts = [p.strip() for p in row.split("|")]
        if len(parts) < 5:
            continue
        if not parts[1].isdigit():
            continue
        pname = parts[2].strip().strip("`")
        if pname not in params_in_03:
            continue
        if pname in _04_DEBIT_CREDIT_INDICATOR_PARAM_NAMES:
            out.add(pname)
            continue
        desc = parts[3].lower()
        if ("debit" in desc and "credit" in desc) or "dr/cr" in desc or "dr / cr" in desc:
            out.add(pname)
    return out


def _serial_series_from_03_param_names(param_names_ordered: list[str]) -> list[tuple[str, str]]:
    """Detect serial-number series (same prefix + consecutive indices). Return list of (first_name, last_name) per series."""
    def parse(name: str) -> tuple[str, int] | None:
        m = re.match(r"^(.+?)(\d+)$", name)
        if m:
            return (m.group(1), int(m.group(2)))
        return None

    series_list: list[tuple[str, str]] = []
    run: list[tuple[str, str, int]] = []  # (name, prefix, index)

    for name in param_names_ordered:
        p = parse(name)
        if p is None:
            if len(run) >= 2:
                series_list.append((run[0][0], run[-1][0]))
            run = []
            continue
        prefix, idx = p
        if run and run[-1][1] == prefix and run[-1][2] + 1 == idx:
            run.append((name, prefix, idx))
        else:
            if len(run) >= 2:
                series_list.append((run[0][0], run[-1][0]))
            run = [(name, prefix, idx)]

    if len(run) >= 2:
        series_list.append((run[0][0], run[-1][0]))
    return series_list


def _serial_series_members_to_skip(ordered_names: list[str], series: list[tuple[str, str]]) -> set[str]:
    """Parameter names that are merged into a **FIRST - LAST** block (all except the first of each series)."""
    skip: set[str] = set()
    present = set(ordered_names)
    for first, last in series:
        if first not in present or last not in present:
            continue
        started = False
        for n in ordered_names:
            if n == first:
                started = True
                continue
            if not started:
                continue
            if n == last:
                skip.add(n)
                break
            skip.add(n)
    return skip


def _serial_group_heading_present(text: str, first: str, last: str) -> bool:
    """True if 04 groups this serial series: slash (table1/table2 style), hyphen, or en-dash."""
    return (
        f"**{first}/{last}**" in text
        or f"**{first} - {last}**" in text
        or f"**{first}\u2013{last}**" in text
    )


def _abbrev_pair_heading_in_text(text: str, f1: str, l1: str, f2: str, l2: str) -> bool:
    """True if text contains abbreviated **STEMa-b / STEM2a-b** matching both ends (e.g. TAB1_ATTR1-5 / TAB2_ATTR1-5)."""
    m1a = re.match(r"^(.+?)(\d+)$", f1)
    m1b = re.match(r"^(.+?)(\d+)$", l1)
    m2a = re.match(r"^(.+?)(\d+)$", f2)
    m2b = re.match(r"^(.+?)(\d+)$", l2)
    if not (m1a and m1b and m2a and m2b):
        return False
    if m1a.group(1) != m1b.group(1) or m2a.group(1) != m2b.group(1):
        return False
    n1, n2 = int(m1a.group(2)), int(m1b.group(2))
    if int(m2a.group(2)) != n1 or int(m2b.group(2)) != n2:
        return False
    abbrev = f"**{m1a.group(1)}{n1}-{n2} / {m2a.group(1)}{n1}-{n2}**"
    return abbrev in text


def _abbrev_pair_heading_v_in_text(text: str, f1: str, l1: str, f2: str, l2: str) -> bool:
    """Abbreviated **TAB1_ATTR1_V-5_V / TAB2_ATTR1_V-5_V** style."""
    m1a = re.match(r"^(.+?)(\d+)(_V)$", f1, re.IGNORECASE)
    m1b = re.match(r"^(.+?)(\d+)(_V)$", l1, re.IGNORECASE)
    m2a = re.match(r"^(.+?)(\d+)(_V)$", f2, re.IGNORECASE)
    m2b = re.match(r"^(.+?)(\d+)(_V)$", l2, re.IGNORECASE)
    if not (m1a and m1b and m2a and m2b):
        return False
    if m1a.group(1) != m1b.group(1) or m2a.group(1) != m2b.group(1):
        return False
    n1, n2 = int(m1a.group(2)), int(m1b.group(2))
    if int(m2a.group(2)) != n1 or int(m2b.group(2)) != n2:
        return False
    suf = m1a.group(3)
    abbrev = f"**{m1a.group(1)}{n1}-{n2}{suf} / {m2a.group(1)}{n1}-{n2}{suf}**"
    return abbrev in text


def _parallel_tab12_heading_ok(text: str, f1: str, l1: str, f2: str, l2: str) -> bool:
    """Combined TAB1/TAB2 block, abbreviated range, or two separate serial-style headings."""
    if f"**{f1} - {l1} / {f2} - {l2}**" in text:
        return True
    if f"**{f1}\u2013{l1} / {f2}\u2013{l2}**" in text:
        return True
    if f"**{f1}/{l1} / {f2}/{l2}**" in text:
        return True
    if _abbrev_pair_heading_in_text(text, f1, l1, f2, l2):
        return True
    if f1.upper().endswith("_V"):
        if _abbrev_pair_heading_v_in_text(text, f1, l1, f2, l2):
            return True
    return _serial_group_heading_present(text, f1, l1) and _serial_group_heading_present(text, f2, l2)


# When an EI has TAB1_* and TAB2_* slot groups 1–5, 04 must document them as one combined block (or two grouped blocks).
_PARALLEL_TAB_SLOT_GROUPS: list[tuple[str, str, int, int, str]] = [
    ("TAB1_ATTR", "TAB2_ATTR", 1, 5, ""),
    ("TAB1_ATTR", "TAB2_ATTR", 1, 5, "_V"),
    ("TAB1_FLD", "TAB2_FLD", 1, 5, ""),
    ("TAB1_FLD", "TAB2_FLD", 1, 5, "_V"),
    ("TAB1_KEY", "TAB2_KEY", 1, 5, ""),
    ("TAB1_KEY", "TAB2_KEY", 1, 5, "_V"),
]


def _parallel_slot_param_names(stem1: str, stem2: str, lo: int, hi: int, suffix: str) -> tuple[list[str], list[str]]:
    if suffix:
        a = [f"{stem1}{k}{suffix}" for k in range(lo, hi + 1)]
        b = [f"{stem2}{k}{suffix}" for k in range(lo, hi + 1)]
    else:
        a = [f"{stem1}{k}" for k in range(lo, hi + 1)]
        b = [f"{stem2}{k}" for k in range(lo, hi + 1)]
    return a, b


def _04_options_heading_present(text04: str, param_name: str) -> bool:
    """True when 04 contains '**PARAM Options:**' heading."""
    return re.search(rf"^\*\*{re.escape(param_name)}\s+Options:\*\*\s*$", text04, re.MULTILINE) is not None


def _04_options_block_text(text04: str, param_name: str) -> str:
    """Return text inside '**PARAM Options:**' block up to next bold heading."""
    m = re.search(
        rf"^\*\*{re.escape(param_name)}\s+Options:\*\*\s*$([\s\S]*?)(?=^\*\*[^*\n].*?\*\*\s*$|\Z)",
        text04,
        re.MULTILINE,
    )
    return m.group(1) if m else ""


def _04_option_line_documents_literal(searchable: str, lit: str) -> bool:
    """
    True when a line documents lit as VALUE — explanation (em dash or hyphen).
    Markdown may use a leading '- ' (list); Word export renders options as plain lines.
    """
    return (
        re.search(
            rf"(?mi)^\s*(?:-\s+)?{re.escape(lit)}\s+[—\-]\s+\S",
            searchable,
        )
        is not None
    )


def _manifest_document_title(*, name_val: str, id_val: str, basename: str) -> str:
    """Document H1/title: 'Exception Indicator: Name ( ID)' with a space after '(' per house style."""
    if name_val and id_val:
        return f"Exception Indicator: {name_val} ( {id_val})"
    if name_val:
        return f"Exception Indicator: {name_val}"
    if id_val:
        return f"Exception Indicator: ( {id_val})"
    return f"Exception Indicator: {basename}"


def _extract_main_sentence(chunk: str) -> str:
    """
    Extract first narrative sentence from a parameter block chunk
    (ignoring blank lines, option bullets, and nested bold headings).
    """
    for line in chunk.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("**"):  # options/connection headings
            break
        if s.startswith("- "):  # bullets in options blocks
            continue
        return s
    return ""


def _normalize_sentence_for_comparison(sentence: str) -> str:
    """
    Normalize sentence for duplication checks:
    - lowercase, collapse whitespace
    - remove markdown bold markers
    - remove parameter-like tokens (ALL_CAPS / underscores) so template-like
      wording with swapped parameter names is still detected
    """
    s = sentence.replace("**", "")
    s = re.sub(r"\b[A-Z][A-Z0-9_]{2,}\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _04_reused_main_sentence_issues(text04: str) -> list[str]:
    """
    Detect verbatim reuse of main-description sentences across different parameter blocks in 04.
    Allows grouped serial-range headers naturally, but flags duplicated first narrative sentence across blocks.
    """
    issues: list[str] = []
    heading_re = re.compile(r"^\*\*([A-Za-z0-9_/\-\s]+)\*\*\s*\(", re.MULTILINE)
    headings = list(heading_re.finditer(text04))
    if not headings:
        return issues

    sentence_to_params: dict[str, list[str]] = {}
    for idx, m in enumerate(headings):
        param = m.group(1).strip()
        start = m.end()
        end = headings[idx + 1].start() if idx + 1 < len(headings) else len(text04)
        chunk = text04[start:end]
        first_sentence = _extract_main_sentence(chunk)
        if not first_sentence:
            continue
        key = _normalize_sentence_for_comparison(first_sentence)
        # Ignore very short sentences to reduce false positives.
        if len(key) < 35:
            continue
        sentence_to_params.setdefault(key, []).append(param)

    for sentence, params in sentence_to_params.items():
        unique_params = sorted(set(params))
        if len(unique_params) >= 2:
            issues.append(
                "04_response.md: repeated main-description sentence across parameters "
                f"{', '.join(unique_params[:6])}{'...' if len(unique_params) > 6 else ''}. "
                "Use parameter-specific wording; do not reuse identical main explanation sentences."
            )
    return issues


def _04_confusable_param_differentiation_issues(text04: str) -> list[str]:
    """
    Ensure confusable parameter families are not explained with effectively
    identical main sentences.
    """
    issues: list[str] = []
    heading_re = re.compile(r"^\*\*([A-Za-z0-9_/\-\s]+)\*\*\s*\(", re.MULTILINE)
    headings = list(heading_re.finditer(text04))
    if not headings:
        return issues

    main_sentence_by_param: dict[str, str] = {}
    for idx, m in enumerate(headings):
        param = m.group(1).strip()
        start = m.end()
        end = headings[idx + 1].start() if idx + 1 < len(headings) else len(text04)
        chunk = text04[start:end]
        first_sentence = _extract_main_sentence(chunk)
        if first_sentence:
            main_sentence_by_param[param] = _normalize_sentence_for_comparison(first_sentence)

    families = [
        ("MSGNO", "MSGTP", "MSGTXT"),
        ("DATE_REF_FLD", "TIME_REF_FLD"),
    ]
    for fam in families:
        present = [p for p in fam if p in main_sentence_by_param]
        if len(present) < 2:
            continue
        vals = [main_sentence_by_param[p] for p in present if main_sentence_by_param[p]]
        if len(vals) >= 2 and len(set(vals)) == 1:
            issues.append(
                "04_response.md: confusable parameters share the same normalized main explanation "
                f"({', '.join(present)}). Differentiate each parameter's business role explicitly."
            )
    return issues


def _05_time_filter_clarity_issues(
    text05: str, params_in_03: set[str], unused: set[str] | None = None
) -> list[str]:
    """
    Enforce clear, plain-language relationship description for date fallback vs duration
    when these parameters exist in the EI.
    """
    issues: list[str] = []
    skip = unused or set()
    active = {p for p in params_in_03 if p.upper() not in skip}
    has_backdays = "BACKDAYS" in active
    has_duration = "DURATION" in active and "DURATION_UNIT" in active
    has_explicit_date = any(p in active for p in {"DATUM", "SND_DATE", "WAIT_DATE", "STAT_DATE"})
    if not (has_backdays and has_duration and has_explicit_date):
        return issues

    t = text05.lower()
    has_explicit_date_text = ("explicit date" in t) or ("snd_date" in t) or ("datum" in t)
    has_backdays_fallback_text = ("backdays" in t) and any(k in t for k in ["fallback", "if no", "not provided", "empty"])
    has_duration_additional_filter_text = ("duration" in t) and ("duration_unit" in t) and any(
        k in t for k in ["additional filter", "second filter", "then filter", "after date", "age filter"]
    )

    if not has_explicit_date_text:
        issues.append(
            "05_response.md: when date and duration parameters exist, explain explicit date parameters "
            "(e.g. DATUM/SND_DATE) in plain language."
        )
    if not has_backdays_fallback_text:
        issues.append(
            "05_response.md: explain that BACKDAYS is fallback when explicit dates are not provided."
        )
    if not has_duration_additional_filter_text:
        issues.append(
            "05_response.md: explain that DURATION + DURATION_UNIT is an additional/second filter after date selection."
        )
    return issues


def _05_combined_effect_block_issues(text05: str) -> list[str]:
    """Reject a global **Combined effect:** (or similar) closing block in Parameter Relationships."""
    issues: list[str] = []
    if re.search(r"^\s*\*\*Combined effect:\*\*", text05, re.MULTILINE | re.IGNORECASE):
        issues.append(
            "05_response.md: remove the **Combined effect:** block; explain how parameters combine "
            "within each relationship group only (no global closing summary)."
        )
    return issues


def _manifest_output_basename() -> str | None:
    """Stem used for Code_/Structure_ files in this pipeline run (from manifest.txt)."""
    mf = RUN_DIR / "manifest.txt"
    if not mf.exists():
        return None
    for line in mf.read_text(encoding="utf-8").splitlines():
        if line.startswith("output_basename="):
            return line.split("=", 1)[1].strip()
    return None


def _code_path_and_text_for_verify() -> tuple[str | None, str | None]:
    """
    Load main EI ABAP source for verify-time inference (e.g. BACKDAYS date anchor).
    Prefers Code_<manifest_basename>.txt; otherwise a single Code_*.txt in input/.
    """
    base = _manifest_output_basename()
    if base:
        for prefix in ("Code_", "Code _"):
            p = INPUT_DIR / f"{prefix}{base}.txt"
            if p.exists():
                t = p.read_text(encoding="utf-8", errors="replace")
                return str(p), t
        r07 = RUN_DIR / "07_response.md"
        if r07.exists():
            text07 = r07.read_text(encoding="utf-8", errors="replace")
            m = re.search(r"```abap\s*\n([\s\S]*?)```", text07, re.IGNORECASE)
            if m and m.group(1).strip():
                return str(r07), m.group(1)
    codes = sorted(INPUT_DIR.glob("Code_*.txt")) + sorted(INPUT_DIR.glob("Code _*.txt"))
    if len(codes) == 1:
        p = codes[0]
        return str(p), p.read_text(encoding="utf-8", errors="replace")
    return None, None


_DATE_RANGE_TABLE_NAMES = (
    "R_DATUM",
    "R_CREDATE",
    "R_ERDAT",
    "R_ERSDA",
    "R_BEDAT",
    "R_MODDA",
    "R_UDATE",
)


def _infer_date_anchor_field_from_code(code: str) -> str | None:
    """
    Best-effort: SAP field whose values are compared via a standard R_* date range table.
    Matches TABLE~FIELD IN R_* and FIELD IN R_* patterns. Returns None when unclear—do not guess.
    """
    if not code.strip():
        return None
    range_names = _DATE_RANGE_TABLE_NAMES
    seen: set[str] = set()
    for rng in range_names:
        for m in re.finditer(
            rf"~([A-Za-z][A-Za-z0-9_]*)\s+IN\s+{rng}\b",
            code,
            re.IGNORECASE,
        ):
            fld = m.group(1).upper()
            if len(fld) >= 4 and fld not in seen:
                seen.add(fld)
                return fld
    for rng in range_names:
        m = re.search(
            rf"\b([A-Za-z][A-Za-z0-9_]{{2,}})\s+IN\s+{rng}\b",
            code,
            re.IGNORECASE,
        )
        if m and m.group(1).upper() not in ("AND", "OR", "NOT"):
            return m.group(1).upper()
    return None


def _infer_backdays_anchor_field_from_code(code: str) -> str | None:
    """Alias for shared date-range inference used by BACKDAYS anchor lines."""
    return _infer_date_anchor_field_from_code(code)


def _infer_forwdays_anchor_field_from_code(code: str) -> str | None:
    """
    SAP date field for FORWDAYS anchor lines. When FORWDAYS only rewrites BACKDAYS
    (LV_BACKDAYS = LV_FORWDAYS …), the effective filter field is the same as BACKDAYS.
    """
    if not code.strip():
        return None
    if re.search(r"LV_BACKDAYS\s*=\s*LV_FORWDAYS\b", code, re.IGNORECASE):
        return _infer_date_anchor_field_from_code(code)
    return _infer_date_anchor_field_from_code(code)


def _04_parameter_block_chunk_for_param(text04: str, param: str) -> str:
    """Body text under **PARAM** ... up to (but not including) the next **SomeName** heading line."""
    m = re.search(
        rf"(?m)^\*\*{re.escape(param)}\*\*[^\n]*\n([\s\S]*?)(?=^\*\*[A-Za-z0-9_])",
        text04,
    )
    if m:
        return m.group(1)
    m_last = re.search(
        rf"(?m)^\*\*{re.escape(param)}\*\*[^\n]*\n([\s\S]*)$",
        text04,
    )
    return m_last.group(1) if m_last else ""


def _04_chunk_stops_at_subsection(line: str) -> bool:
    s = line.strip()
    if not s:
        return False
    if re.match(r"^\*\*[^*]+ Options:\*\*", s, re.IGNORECASE):
        return True
    if re.search(r"\bConnection:\*\*\s*$", s, re.IGNORECASE):
        return True
    return False


def _04_main_explanation_for_dictionary_compare(chunk: str, pname: str) -> str:
    """
    Main explanation text used for dictionary verbatim checks (excludes Options/Connection subsections).
    """
    u = pname.strip().upper()
    lines: list[str] = []
    for line in chunk.splitlines():
        if _04_chunk_stops_at_subsection(line):
            break
        lines.append(line)
    text = "\n".join(lines).strip()
    if u == "BACKDAYS":
        for sent in (_BACKDAYS_WINDOW_SENTENCE_EN_DASH, _BACKDAYS_WINDOW_SENTENCE_ASCII_DASH):
            text = text.replace(sent, "")
        text = re.sub(
            r"Backdays is based on [^\n.]+\.?\s*",
            "",
            text,
            flags=re.IGNORECASE,
        )
        return text.strip()
    if u == "FORWDAYS":
        for sent in (_FORWDAYS_WINDOW_SENTENCE_EN_DASH, _FORWDAYS_WINDOW_SENTENCE_ASCII_DASH):
            text = text.replace(sent, "")
        text = re.sub(
            r"Forwdays is based on [^\n.]+\.?\s*",
            "",
            text,
            flags=re.IGNORECASE,
        )
        return text.strip()
    if u in ("USER_FLD", "USR_FLD"):
        if text.startswith(_USER_FLD_DRL_FIXED_MARKDOWN):
            text = text[len(_USER_FLD_DRL_FIXED_MARKDOWN) :].lstrip("\n")
        if re.search(r"(?im)^Explicit values from the supplied ABAP:\s*$", text):
            text = re.split(r"(?im)^Explicit values from the supplied ABAP:\s*$", text, maxsplit=1)[0]
        return text.strip()
    text = re.sub(r"(?m)^\*\*Not in use\*\*\s*$", "", text)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def _serial_group_member_names(first: str, last: str, ordered_names: list[str]) -> list[str]:
    mf = re.match(r"^(.+?)(\d+)$", first)
    ml = re.match(r"^(.+?)(\d+)$", last)
    if not mf or not ml or mf.group(1) != ml.group(1):
        return [first]
    prefix = mf.group(1)
    n0, n1 = int(mf.group(2)), int(ml.group(2))
    members = []
    for p in ordered_names:
        m = re.match(rf"^{re.escape(prefix)}(\d+)$", p, re.IGNORECASE)
        if m and n0 <= int(m.group(1)) <= n1:
            members.append(p)
    return members or [first]


def _members_from_04_heading(
    heading: str, ordered_names: list[str], serial_series: list[tuple[str, str]]
) -> list[str]:
    heading = heading.strip()
    if " - " in heading and " / " not in heading:
        fr, lr = [x.strip() for x in heading.split(" - ", 1)]
        for a, b in serial_series:
            if fr == a and lr == b:
                return _serial_group_member_names(fr, lr, ordered_names)
    tok = re.match(r"^([A-Za-z0-9_]+)", heading)
    return [tok.group(1)] if tok else [heading]


def _04_append_not_in_use_if_unused(parts: list[str], unused: set[str], members: list[str]) -> None:
    if members and all(m.strip().upper() in unused for m in members):
        parts.append("")
        parts.append(NOT_IN_USE_LINE)


def _04_insert_not_in_use_lines(
    text04: str, unused: set[str], ordered_names: list[str]
) -> str:
    if not unused:
        return text04
    serial_series = _serial_series_from_03_param_names(ordered_names)
    lines = text04.splitlines()
    out: list[str] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        hm = re.match(r"^\*\*([^*]+)\*\*\s*\(", line)
        if hm:
            members = _members_from_04_heading(hm.group(1), ordered_names, serial_series)
            out.append(line)
            i += 1
            block_lines: list[str] = []
            while i < len(lines):
                nl = lines[i]
                if re.match(r"^\*\*[^*]+\*\*", nl):
                    break
                block_lines.append(nl)
                i += 1
            opts_idx = None
            for j, bl in enumerate(block_lines):
                if re.match(r"^\*\*.+ Options:\*\*", bl.strip()):
                    opts_idx = j
                    break
            main_part = block_lines if opts_idx is None else block_lines[:opts_idx]
            opts_part = [] if opts_idx is None else block_lines[opts_idx:]
            out.extend(main_part)
            if members and all(m.upper() in unused for m in members):
                if not any(NOT_IN_USE_LINE in bl for bl in main_part):
                    if out and out[-1].strip():
                        out.append("")
                    out.append(NOT_IN_USE_LINE)
            out.extend(opts_part)
            continue
        out.append(line)
        i += 1
    return "\n".join(out)


def _resolve_unused_params_set(paths: dict[str, Path] | None = None) -> set[str]:
    cached = read_unused_params_file(UNUSED_PARAMS_RUN_FILE)
    if cached:
        return cached
    if paths and paths.get("code") and paths.get("params"):
        return analyze_unused_params(
            paths["code"], paths["params"], search_dirs=[INPUT_DIR, INPUT_DIR / "old"]
        )
    discovered = _discover_inputs(assume_yes=True)
    if discovered and discovered.get("code") and discovered.get("params"):
        return analyze_unused_params(
            discovered["code"],
            discovered["params"],
            search_dirs=[INPUT_DIR, INPUT_DIR / "old"],
        )
    return set()


def _04_block_has_not_in_use(chunk: str) -> bool:
    return bool(re.search(r"(?m)^\*\*Not in use\*\*\s*$", chunk))


def _04_not_in_use_marker_issues(text04: str, text03: str, unused: set[str]) -> list[str]:
    issues: list[str] = []
    ordered, _ = _param_names_ordered_from_03_table(text03)
    serial_series = _serial_series_from_03_param_names(ordered)
    serial_skip = _serial_series_members_to_skip(ordered, serial_series)
    checked_blocks: set[str] = set()
    for pname in ordered:
        if pname in serial_skip:
            continue
        members = [pname]
        block_key = pname
        for fr, lr in serial_series:
            if pname == fr:
                members = _serial_group_member_names(fr, lr, ordered)
                block_key = f"{fr}-{lr}"
                break
        if block_key in checked_blocks:
            continue
        checked_blocks.add(block_key)
        chunk = _04_parameter_block_chunk_for_param(text04, pname)
        if not chunk.strip():
            continue
        has_marker = _04_block_has_not_in_use(chunk)
        all_unused = bool(members) and all(m.upper() in unused for m in members)
        if all_unused and not has_marker:
            label = f"**{pname}**" if len(members) == 1 else f"serial group **{members[0]} - {members[-1]}**"
            issues.append(
                f"04_response.md: unused parameter block {label} must include {NOT_IN_USE_LINE} "
                "after its main description."
            )
        elif not all_unused and has_marker:
            label = f"**{pname}**" if len(members) == 1 else f"serial group **{members[0]} - {members[-1]}**"
            issues.append(
                f"04_response.md: active parameter block {label} must not include {NOT_IN_USE_LINE}."
            )
    return issues


def _text_mentions_param_token(text: str, param: str) -> bool:
    u = param.upper()
    if re.search(rf"\*\*{re.escape(param)}\*\*", text, re.IGNORECASE):
        return True
    return bool(re.search(rf"(?<![A-Za-z0-9_]){re.escape(u)}(?![A-Za-z0-9_])", text, re.IGNORECASE))


def _05_unused_param_presence_issues(text05: str, unused: set[str]) -> list[str]:
    issues: list[str] = []
    for p in sorted(unused):
        if _text_mentions_param_token(text05, p):
            issues.append(
                f"05_response.md: unused parameter {p!r} must not appear in Parameter Relationships "
                f"(documented with {NOT_IN_USE_LINE} in section 04 only)."
            )
    return issues


def _06_unused_param_presence_issues(text06: str, unused: set[str]) -> list[str]:
    issues: list[str] = []
    dv = _06_default_values_section_chunk(text06)
    for p in sorted(unused):
        if _06_default_values_covers_param(dv, p):
            issues.append(
                f"06_response.md: unused parameter {p!r} must not appear in Default Values."
            )
    for block in _06_iter_use_case_blocks(text06):
        for line in block.get("code_lines", []):
            m = _06_PARAM_ASSIGNMENT_LINE_RE.match(str(line).strip())
            if m and m.group(1).upper() in unused:
                issues.append(
                    f"06_response.md: Use Case {block['num']}: unused parameter {m.group(1)!r} "
                    "must not appear in practical configuration examples."
                )
    return issues


# Mandatory verbatim BACKDAYS wording in 04 (user-specified; do not paraphrase).
_BACKDAYS_WINDOW_SENTENCE_EN_DASH = (
    "BACKDAYS defines the historical monitoring window by specifying how many days backward from today "
    "to retrieve records. 0 \u2013 today, 1 \u2013 today + yesterday etc."
)
_BACKDAYS_WINDOW_SENTENCE_ASCII_DASH = (
    "BACKDAYS defines the historical monitoring window by specifying how many days backward from today "
    "to retrieve records. 0 - today, 1 - today + yesterday etc."
)
_BACKDAYS_ANCHOR_DATE_REF_SENTENCE_EXACT = "Backdays is based on DATE_REF_FLD field."

# Mandatory verbatim FORWDAYS wording in 04 (mirror of BACKDAYS §3a).
_FORWDAYS_WINDOW_SENTENCE_EN_DASH = (
    "FORWDAYS defines the historical monitoring window by specifying how many days forward from today "
    "to retrieve records. 0 \u2013 today, 1 \u2013 today + tomorrow etc."
)
_FORWDAYS_WINDOW_SENTENCE_ASCII_DASH = (
    "FORWDAYS defines the historical monitoring window by specifying how many days forward from today "
    "to retrieve records. 0 - today, 1 - today + tomorrow etc."
)
_FORWDAYS_ANCHOR_DATE_REF_SENTENCE_EXACT = "Forwdays is based on DATE_REF_FLD field."

# DURATION_UNIT Options subsection (single source of truth; matches PROMPT_Parameter_Configuration_Guidelines_section.md).
_DURATION_UNIT_OPTIONS_HEADING = "**DURATION_UNIT Options:**"
_DURATION_UNIT_OPTION_LINES: tuple[str, ...] = (
    "- H: Hours",
    "- M: Minutes",
    "- D: Days",
    "- F: Full days for specific day filtering",
)


def _append_duration_unit_options_block(parts: list[str]) -> None:
    parts.append(_DURATION_UNIT_OPTIONS_HEADING)
    parts.extend(_DURATION_UNIT_OPTION_LINES)
    parts.append("")


# USER_FLD / DRL: mandatory narrative in section 04 (do not paraphrase). See PROMPT_Parameter_Configuration_Guidelines_section.md §3b.
_USER_FLD_DRL_FIXED_MARKDOWN = (
    "The USER_FLD parameter serves a dual purpose in work process monitoring: filtering results and enabling "
    "dynamic notification routing through the Dynamic Recipient List (DRL) mechanism. Instead of notifying "
    "pre-assigned users, the DRL mechanism provides a flexible, role-based notification system that automatically "
    "determines the appropriate recipients based on the monitoring results.\n"
    "How DRL Works:\n"
    "When USER_FLD is specified, the system extracts values from that field in the monitoring result set\n"
    "These extracted values are then used as recipient addresses for alert notifications\n"
    "This creates a dynamic, role-based notification mechanism that adapts based on the actual data being monitored\n"
    "The mechanism is much more flexible than defining and constantly updating lists of specific pre-assigned users"
)
_USER_FLD_DRL_VERIFY_PHRASES = [
    "dynamic recipient list (drl)",
    "how drl works:",
    "when user_fld is specified, the system extracts values from that field in the monitoring result set",
    "these extracted values are then used as recipient addresses for alert notifications",
]

def _date_ref_fld_literals_from_code(code: str) -> list[str]:
    """
    Extract explicit DATE_REF_FLD field-name literals from ABAP.
    Primary target: CASE LV_DATE_REF_FLD / WHEN '...'.
    """
    if not code.strip():
        return []
    found: list[str] = []
    for m in re.finditer(
        r"\bCASE\s+(?:LV_)?DATE_REF_FLD\b([\s\S]*?)\bENDCASE\.",
        code,
        re.IGNORECASE,
    ):
        body = m.group(1)
        for when_line in re.findall(r"(?im)^\s*WHEN\b[^\n]*", body):
            for q in re.finditer(r"['\"]([A-Za-z][A-Za-z0-9_]*)['\"]", when_line):
                found.append(q.group(1).upper())
    # Also support direct literal assignment forms when present.
    for m in re.finditer(r"\bDATE_REF_FLD\b\s*=\s*['\"]([A-Za-z][A-Za-z0-9_]*)['\"]", code, re.IGNORECASE):
        found.append(m.group(1).upper())
    for m in re.finditer(r"\bLV_\w*DATE_REF_FLD\w*\b\s*=\s*['\"]([A-Za-z][A-Za-z0-9_]*)['\"]", code, re.IGNORECASE):
        found.append(m.group(1).upper())
    out: list[str] = []
    seen: set[str] = set()
    for t in found:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out[:60]


def _aggr_level_literals_from_code(code: str) -> list[str]:
    """Extract explicit AGGR_LEVEL/AGGLEVEL literals from ABAP assignments and comparisons."""
    if not code.strip():
        return []
    found: list[str] = []
    for m in re.finditer(r"\b(?:LV_)?AGGR?_LEVEL\b\s*=\s*['\"]([A-Za-z0-9_]{1,20})['\"]", code, re.IGNORECASE):
        found.append(m.group(1).upper())
    for m in re.finditer(
        r"\b(?:LV_)?AGGR?_LEVEL\b\s*(?:<>|=|EQ|NE)\s*['\"]([A-Za-z0-9_]{1,20})['\"]",
        code,
        re.IGNORECASE,
    ):
        found.append(m.group(1).upper())
    for m in re.finditer(r"\bCASE\s+(?:LV_)?AGGR?_LEVEL\b([\s\S]*?)\bENDCASE\.", code, re.IGNORECASE):
        body = m.group(1)
        for when_line in re.findall(r"(?im)^\s*WHEN\b[^\n]*", body):
            for q in re.finditer(r"['\"]([A-Za-z0-9_]{1,20})['\"]", when_line):
                found.append(q.group(1).upper())
    out: list[str] = []
    seen: set[str] = set()
    for t in found:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out[:20]


def _aggr_level_literal_expl(lit: str) -> str:
    """Short stable explanation for known AGGR_LEVEL values."""
    lu = lit.upper()
    if lu == "T":
        return "Total aggregation across all servers."
    if lu == "S":
        return "Server-level aggregation."
    return "Code-derived aggregation mode."


def _user_fld_literal_values_from_code(code: str) -> list[str]:
    """
    Strict: SAP-like field tokens only where ABAP clearly assigns into USER_FLD or assigns USER_FLD from a literal.
    """
    if not code.strip():
        return []
    found: list[str] = []
    for m in re.finditer(r"\bUSER_FLD\b\s*=\s*['\"]([A-Z][A-Z0-9_]*)['\"]", code, re.IGNORECASE):
        found.append(m.group(1).upper())
    for m in re.finditer(r"\bLV_\w*USER_FLD\w*\b\s*=\s*['\"]([A-Z][A-Z0-9_]*)['\"]", code, re.IGNORECASE):
        found.append(m.group(1).upper())
    for m in re.finditer(
        r"['\"]([A-Z][A-Z0-9_]{2,29})['\"]\s+TO\s+\bUSER_FLD\b", code, re.IGNORECASE
    ):
        found.append(m.group(1).upper())
    out: list[str] = []
    seen: set[str] = set()
    for t in found:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out[:40]


def _04_user_fld_drl_present(chunk: str) -> bool:
    low = re.sub(r"\s+", " ", chunk).strip().lower()
    return all(p in low for p in _USER_FLD_DRL_VERIFY_PHRASES)


def _04_user_fld_guideline_issues(text04: str, params_in_03: set[str]) -> list[str]:
    """
    When USER_FLD or USR_FLD is in Parameters, 04 must include the mandatory DRL narrative (verbatim intent)
    and optional code-derived literals. See PROMPT_Parameter_Configuration_Guidelines_section.md §3b.
    """
    issues: list[str] = []
    names = [p for p in ("USER_FLD", "USR_FLD") if p in params_in_03]
    if not names:
        return issues
    for nm in ("USER_FLD", "USR_FLD"):
        if re.search(rf"^\*\*{nm}\s+Sample values:\*\*\s*$", text04, re.MULTILINE):
            issues.append(
                f"04_response.md: do not use '**{nm} Sample values:**' (USER_FLD / USR_FLD use §3b DRL text "
                "and optional ABAP literals only)."
            )
        if re.search(rf"^\*\*{nm}\s+Options:\*\*\s*$", text04, re.MULTILINE):
            issues.append(
                f"04_response.md: do not use '**{nm} Options:**' (USER_FLD / USR_FLD use §3b DRL narrative "
                "and optional ABAP literals only, not a generic Options subsection)."
            )
    _cp, code = _code_path_and_text_for_verify()
    literals = _user_fld_literal_values_from_code(code) if code else []

    for param in names:
        chunk = _04_parameter_block_chunk_for_param(text04, param)
        if not chunk.strip():
            issues.append(
                f"04_response.md: {param} appears in 03 but no **{param}** parameter block was found in 04."
            )
            continue
        low = re.sub(r"\s+", " ", chunk).strip().lower()
        for phrase in _USER_FLD_DRL_VERIFY_PHRASES:
            if phrase not in low:
                issues.append(
                    f"04_response.md: {param} block must include the mandatory USER_FLD / DRL narrative "
                    f"(missing distinctive phrase: {phrase!r}). See prompts/PROMPT_Parameter_Configuration_Guidelines_section.md §3b."
                )
        if "no fixed user_fld value list" in low:
            issues.append(
                f"04_response.md: {param} block must not use the deprecated fallback "
                "'No fixed USER_FLD value list is defined in the available code for this EI.' "
                "See §3b: after the mandatory DRL text, add code literals only if present; otherwise add nothing further."
            )
        for lit in literals:
            if not re.search(rf"\b{re.escape(lit)}\b", chunk, re.IGNORECASE):
                issues.append(
                    f"04_response.md: {param} block must list explicit ABAP literal {lit!r} for USER_FLD "
                    "(found in Code file on a line referencing USER_FLD). Add it after the mandatory DRL text."
                )
    return issues


def _04_backdays_window_sentence_present(chunk: str) -> bool:
    """True if chunk contains the exact BACKDAYS window sentence (en-dash or ASCII hyphen variants)."""
    return _BACKDAYS_WINDOW_SENTENCE_EN_DASH in chunk or _BACKDAYS_WINDOW_SENTENCE_ASCII_DASH in chunk


def _04_forwdays_window_sentence_present(chunk: str) -> bool:
    """True if chunk contains the exact FORWDAYS window sentence (en-dash or ASCII hyphen variants)."""
    return _FORWDAYS_WINDOW_SENTENCE_EN_DASH in chunk or _FORWDAYS_WINDOW_SENTENCE_ASCII_DASH in chunk


def _04_backdays_wording_issues(text04: str, params_in_03: set[str]) -> list[str]:
    """
    When BACKDAYS is in 03, the BACKDAYS block in 04 must contain the verbatim window sentence; an anchor
    line is required only for DATE_REF_FLD or when a field can be inferred from the Code file—otherwise
    omit any 'Backdays is based on …' line. See PROMPT_Parameter_Configuration_Guidelines_section.md.
    """
    issues: list[str] = []
    if "BACKDAYS" not in params_in_03:
        return issues
    chunk = _04_parameter_block_chunk_for_param(text04, "BACKDAYS")
    if not chunk.strip():
        issues.append(
            "04_response.md: BACKDAYS appears in 03 but no **BACKDAYS** parameter block was found in 04."
        )
        return issues
    if not _04_backdays_window_sentence_present(chunk):
        issues.append(
            "04_response.md: BACKDAYS block must contain this sentence exactly (en-dash or hyphen-minus OK "
            "only for the two dashes): "
            f"'{_BACKDAYS_WINDOW_SENTENCE_EN_DASH}' "
            f"or '{_BACKDAYS_WINDOW_SENTENCE_ASCII_DASH}'."
        )
    if "DATE_REF_FLD" in params_in_03:
        if _BACKDAYS_ANCHOR_DATE_REF_SENTENCE_EXACT not in chunk:
            issues.append(
                "04_response.md: When DATE_REF_FLD is in Parameters, BACKDAYS block must contain exactly: "
                f"'{_BACKDAYS_ANCHOR_DATE_REF_SENTENCE_EXACT}'"
            )
    else:
        _cp, code = _code_path_and_text_for_verify()
        inferred = _infer_backdays_anchor_field_from_code(code) if code else None
        if inferred:
            required = f"Backdays is based on {inferred}"
            if required not in chunk:
                issues.append(
                    "04_response.md: Without DATE_REF_FLD in Parameters, BACKDAYS block must contain exactly "
                    f"(no extra words): '{required}' (FIELD inferred from Code file)."
                )
        else:
            if re.search(r"Backdays is based on\b", chunk, re.IGNORECASE):
                issues.append(
                    "04_response.md: BACKDAYS block must not include a 'Backdays is based on …' line when the "
                    "anchor field cannot be inferred from the Code file (omit it), unless DATE_REF_FLD is in Parameters."
                )
    return issues


def _04_forwdays_wording_issues(text04: str, params_in_03: set[str]) -> list[str]:
    """
    When FORWDAYS is in 03, the FORWDAYS block in 04 must contain the verbatim forward window sentence;
    an anchor line mirrors BACKDAYS (DATE_REF_FLD or code-inferred FIELD). See prompt §3c.
    """
    issues: list[str] = []
    if "FORWDAYS" not in params_in_03:
        return issues
    chunk = _04_parameter_block_chunk_for_param(text04, "FORWDAYS")
    if not chunk.strip():
        issues.append(
            "04_response.md: FORWDAYS appears in 03 but no **FORWDAYS** parameter block was found in 04."
        )
        return issues
    if not _04_forwdays_window_sentence_present(chunk):
        issues.append(
            "04_response.md: FORWDAYS block must contain this sentence exactly (en-dash or hyphen-minus OK "
            "only for the two dashes): "
            f"'{_FORWDAYS_WINDOW_SENTENCE_EN_DASH}' "
            f"or '{_FORWDAYS_WINDOW_SENTENCE_ASCII_DASH}'."
        )
    if "DATE_REF_FLD" in params_in_03:
        if _FORWDAYS_ANCHOR_DATE_REF_SENTENCE_EXACT not in chunk:
            issues.append(
                "04_response.md: When DATE_REF_FLD is in Parameters, FORWDAYS block must contain exactly: "
                f"'{_FORWDAYS_ANCHOR_DATE_REF_SENTENCE_EXACT}'"
            )
    else:
        _cp, code = _code_path_and_text_for_verify()
        inferred = _infer_forwdays_anchor_field_from_code(code) if code else None
        if inferred:
            required = f"Forwdays is based on {inferred}"
            if required not in chunk:
                issues.append(
                    "04_response.md: Without DATE_REF_FLD in Parameters, FORWDAYS block must contain exactly "
                    f"(no extra words): '{required}' (FIELD inferred from Code file)."
                )
        else:
            if re.search(r"Forwdays is based on\b", chunk, re.IGNORECASE):
                issues.append(
                    "04_response.md: FORWDAYS block must not include a 'Forwdays is based on …' line when the "
                    "anchor field cannot be inferred from the Code file (omit it), unless DATE_REF_FLD is in Parameters."
                )
    return issues


def _04_date_ref_fld_guideline_issues(text04: str, params_in_03: set[str]) -> list[str]:
    """
    When DATE_REF_FLD exists and ABAP has explicit DATE_REF_FLD literals, 04 must list them in the DATE_REF_FLD block.
    """
    issues: list[str] = []
    if "DATE_REF_FLD" not in params_in_03:
        return issues
    chunk = _04_parameter_block_chunk_for_param(text04, "DATE_REF_FLD")
    if not chunk.strip():
        issues.append(
            "04_response.md: DATE_REF_FLD appears in 03 but no **DATE_REF_FLD** parameter block was found in 04."
        )
        return issues
    # DATE_REF_FLD literals are documented in the Options subsection; validate against both
    # the block body and the options text.
    opt = _04_options_block_text(text04, "DATE_REF_FLD")
    searchable = f"{chunk}\n{opt}"
    _cp, code = _code_path_and_text_for_verify()
    literals = _date_ref_fld_literals_from_code(code) if code else []
    if literals:
        if "Explicit DATE_REF_FLD literal found in the supplied ABAP" in searchable:
            issues.append(
                "04_response.md: DATE_REF_FLD literals must use lines like 'VALUE — short explanation' "
                "(optional leading '- ' in markdown; do not use the old 'Explicit DATE_REF_FLD literal found...' wording)."
            )
        if any(
            g in opt
            for g in (
                "Use a field name from the EI's source structure",
                "Values are system-specific; choose the field the ABAP selection uses for the primary date axis.",
            )
        ):
            issues.append(
                "04_response.md: DATE_REF_FLD has explicit ABAP literals; replace generic fallback guidance with "
                "one line per literal in 'VALUE — short explanation' format (optional '- ' list prefix in markdown)."
            )
    for lit in literals:
        if not re.search(rf"\b{re.escape(lit)}\b", searchable, re.IGNORECASE):
            issues.append(
                f"04_response.md: DATE_REF_FLD block must list explicit ABAP literal {lit!r} "
                "(found in Code file via DATE_REF_FLD CASE/WHEN or assignment)."
            )
            continue
        if not _04_option_line_documents_literal(searchable, lit):
            issues.append(
                f"04_response.md: DATE_REF_FLD literal {lit!r} must be documented as "
                f"'{lit} — <short explanation>' (optional leading '- ' in markdown; em dash or hyphen before text)."
            )
    return issues


def _04_duration_unit_options_issues(text04: str, params_in_03: set[str]) -> list[str]:
    """When DURATION_UNIT is in 03, Options block must match _DURATION_UNIT_OPTION_LINES verbatim."""
    issues: list[str] = []
    if "DURATION_UNIT" not in params_in_03:
        return issues
    chunk = _04_parameter_block_chunk_for_param(text04, "DURATION_UNIT")
    if not chunk.strip():
        issues.append(
            "04_response.md: DURATION_UNIT appears in 03 but no **DURATION_UNIT** parameter block was found in 04."
        )
        return issues
    opt = _04_options_block_text(text04, "DURATION_UNIT")
    if not opt.strip():
        issues.append(
            "04_response.md: DURATION_UNIT appears in 03 but no **DURATION_UNIT Options:** block was found in 04."
        )
        return issues
    opt_norm = opt.replace("\r\n", "\n")
    for line in _DURATION_UNIT_OPTION_LINES:
        if line not in opt_norm:
            issues.append(
                "04_response.md: DURATION_UNIT Options must include this line exactly "
                f"(see pipeline.py _DURATION_UNIT_OPTION_LINES): {line!r}"
            )
    return issues


def _04_aggr_level_guideline_issues(text04: str, params_in_03: set[str]) -> list[str]:
    """
    When AGGR_LEVEL/AGGLEVEL exists and ABAP has explicit literals, 04 must list them in Options.
    """
    issues: list[str] = []
    present = [p for p in ("AGGR_LEVEL", "AGGLEVEL") if p in params_in_03]
    if not present:
        return issues
    _cp, code = _code_path_and_text_for_verify()
    literals = _aggr_level_literals_from_code(code) if code else []
    if not literals:
        return issues
    for param in present:
        block = _04_options_block_text(text04, param)
        if not block.strip():
            issues.append(
                f"04_response.md: {param} has code-derived literals but no **{param} Options:** block was found."
            )
            continue
        for lit in literals:
            if not _04_option_line_documents_literal(block, lit):
                issues.append(
                    f"04_response.md: {param} literal {lit!r} must be documented in options as "
                    f"'{lit} — <short explanation>' (optional leading '- ' in markdown)."
                )
    return issues


def _04_params_dictionary_canonical_issues(text04: str, text03: str) -> list[str]:
    """
    When input/params_dictionary.xlsx contains an entry for a parameter listed in 03, the main
    explanation in 04 must match that dictionary text verbatim (no appended EI/table clauses).
    """
    param_canon = _load_params_dictionary_explanations()
    if not param_canon:
        return []
    issues: list[str] = []
    param_names_ordered, _ = _param_names_ordered_from_03_table(text03)
    seen_upper: set[str] = set()
    n_mand_en = _normalize_text_for_dictionary_compare(_BACKDAYS_WINDOW_SENTENCE_EN_DASH)
    n_mand_ascii = _normalize_text_for_dictionary_compare(_BACKDAYS_WINDOW_SENTENCE_ASCII_DASH)
    n_fwd_en = _normalize_text_for_dictionary_compare(_FORWDAYS_WINDOW_SENTENCE_EN_DASH)
    n_fwd_ascii = _normalize_text_for_dictionary_compare(_FORWDAYS_WINDOW_SENTENCE_ASCII_DASH)
    for pname in param_names_ordered:
        u = pname.strip().upper()
        if u in seen_upper:
            continue
        seen_upper.add(u)
        canon = param_canon.get(u)
        if not canon:
            continue
        chunk = _04_parameter_block_chunk_for_param(text04, pname)
        if not chunk.strip():
            continue
        n_exp = _normalize_text_for_dictionary_compare(canon)
        if not n_exp:
            continue
        if u == "BACKDAYS":
            if not _04_backdays_window_sentence_present(chunk):
                issues.append(
                    "04_response.md: **BACKDAYS** must include the mandatory monitoring-window sentence "
                    "(see PROMPT_Parameter_Configuration_Guidelines_section.md §3a)."
                )
            n_extra = _normalize_text_for_dictionary_compare(
                _04_main_explanation_for_dictionary_compare(chunk, "BACKDAYS")
            )
            if n_exp in (n_mand_en, n_mand_ascii):
                if n_extra:
                    issues.append(
                        "04_response.md: **BACKDAYS** must not add text beyond the mandatory window sentence "
                        "and optional anchor when the dictionary entry is only that sentence."
                    )
            elif n_extra != n_exp:
                issues.append(
                    "04_response.md: dictionary entry for "
                    f"{pname!r} must be the only extra main text before the mandatory BACKDAYS sentence "
                    "(verbatim; no appended clauses)."
                )
            continue
        if u == "FORWDAYS":
            if not _04_forwdays_window_sentence_present(chunk):
                issues.append(
                    "04_response.md: **FORWDAYS** must include the mandatory forward monitoring-window sentence "
                    "(see PROMPT_Parameter_Configuration_Guidelines_section.md §3c)."
                )
            n_extra = _normalize_text_for_dictionary_compare(
                _04_main_explanation_for_dictionary_compare(chunk, "FORWDAYS")
            )
            if n_exp in (n_fwd_en, n_fwd_ascii):
                if n_extra:
                    issues.append(
                        "04_response.md: **FORWDAYS** must not add text beyond the mandatory window sentence "
                        "and optional anchor when the dictionary entry is only that sentence."
                    )
            elif n_extra != n_exp:
                issues.append(
                    "04_response.md: dictionary entry for "
                    f"{pname!r} must be the only extra main text before the mandatory FORWDAYS sentence "
                    "(verbatim; no appended clauses)."
                )
            continue
        if u in ("USER_FLD", "USR_FLD"):
            if not _04_user_fld_drl_present(chunk):
                issues.append(
                    "04_response.md: **USER_FLD** / **USR_FLD** must include the mandatory DRL narrative "
                    "(see PROMPT_Parameter_Configuration_Guidelines_section.md §3b)."
                )
            n_dict = _normalize_text_for_dictionary_compare(
                _04_main_explanation_for_dictionary_compare(chunk, pname)
            )
            if n_dict != n_exp:
                issues.append(
                    "04_response.md: dictionary entry for "
                    f"{pname!r} must appear verbatim after the DRL block (no paraphrase or extra wording)."
                )
            continue
        n_main = _normalize_text_for_dictionary_compare(
            _04_main_explanation_for_dictionary_compare(chunk, pname)
        )
        if n_main != n_exp:
            issues.append(
                "04_response.md: dictionary entry exists for "
                f"{pname!r} in input/params_dictionary.xlsx; the main explanation must be that text "
                "verbatim only (no appended or prepended sentences; Options/Connection subsections may follow)."
            )
    return issues


def _04_dictionary_mark_when_unchecked_issues(text04: str, text03: str) -> list[str]:
    """
    When input/checked params.txt is active (non-empty), any parameter that appears in
    params_dictionary.xlsx but is NOT listed in checked params.txt must wrap its dictionary
    main paragraph in <mark>...</mark> (yellow in Word).
    """
    checked = _load_checked_params_set()
    if checked is None:
        return []
    param_canon = _load_params_dictionary_explanations()
    if not param_canon:
        return []
    issues: list[str] = []
    param_names_ordered, _ = _param_names_ordered_from_03_table(text03)
    seen_upper: set[str] = set()
    for pname in param_names_ordered:
        u = pname.strip().upper()
        if u in seen_upper:
            continue
        seen_upper.add(u)
        if pname in ("USER_FLD", "USR_FLD"):
            continue
        if u not in param_canon:
            continue
        if u in checked:
            continue
        chunk = _04_parameter_block_chunk_for_param(text04, pname)
        if not chunk.strip():
            continue
        # BACKDAYS: dictionary text matches the mandatory window sentence — do not require a
        # duplicate <mark> copy when the block already contains the verbatim sentence (see §3a).
        if u == "BACKDAYS" and _04_backdays_window_sentence_present(chunk):
            canon_b = param_canon.get("BACKDAYS") or ""
            n_c = _normalize_text_for_dictionary_compare(canon_b)
            n_en = _normalize_text_for_dictionary_compare(_BACKDAYS_WINDOW_SENTENCE_EN_DASH)
            n_ascii = _normalize_text_for_dictionary_compare(_BACKDAYS_WINDOW_SENTENCE_ASCII_DASH)
            if n_c and (n_c == n_en or n_c == n_ascii):
                continue
        if u == "FORWDAYS" and _04_forwdays_window_sentence_present(chunk):
            canon_f = param_canon.get("FORWDAYS") or ""
            n_c = _normalize_text_for_dictionary_compare(canon_f)
            n_en = _normalize_text_for_dictionary_compare(_FORWDAYS_WINDOW_SENTENCE_EN_DASH)
            n_ascii = _normalize_text_for_dictionary_compare(_FORWDAYS_WINDOW_SENTENCE_ASCII_DASH)
            if n_c and (n_c == n_en or n_c == n_ascii):
                continue
        if "<mark>" not in chunk.lower():
            issues.append(
                "04_response.md: parameter "
                f"{pname!r} is in input/params_dictionary.xlsx but not in input/checked params.txt; "
                "wrap the dictionary main paragraph in <mark>...</mark> (see PROMPT_Parameter_Configuration_Guidelines_section.md rule 0)."
            )
    return issues


def _04_dictionary_mark_when_checked_issues(text04: str, text03: str) -> list[str]:
    """
    When checked params.txt is active, parameters in both the dictionary and the checked list
    must not wrap the main explanation in <mark> (yellow highlight is for unchecked dictionary only).
    """
    checked = _load_checked_params_set()
    if checked is None:
        return []
    param_canon = _load_params_dictionary_explanations()
    if not param_canon:
        return []
    issues: list[str] = []
    param_names_ordered, _ = _param_names_ordered_from_03_table(text03)
    seen_upper: set[str] = set()
    for pname in param_names_ordered:
        u = pname.strip().upper()
        if u in seen_upper:
            continue
        seen_upper.add(u)
        if u not in param_canon or u not in checked:
            continue
        chunk = _04_parameter_block_chunk_for_param(text04, pname)
        if not chunk.strip():
            continue
        main = _04_main_explanation_for_dictionary_compare(chunk, pname)
        searchable = main if main else chunk
        if "<mark>" in searchable.lower():
            issues.append(
                "04_response.md: parameter "
                f"{pname!r} is in input/params_dictionary.xlsx and input/checked params.txt; "
                "do not wrap the dictionary main paragraph in <mark>...</mark> (see PROMPT_Parameter_Configuration_Guidelines_section.md rule 0)."
            )
    return issues


def verify_responses() -> list[str]:
    """Check that response files exist and 04_response.md meets Parameter Configuration Guidelines rules. Return list of error strings."""
    errors = []
    for num, _, _ in SECTION_SPEC:
        r = RUN_DIR / f"{num}_response.md"
        if not r.exists():
            errors.append(f"Missing response file: {r.name}")
    if errors:
        return errors

    # Duplicate ### Title + ## Title (same text) breaks Word (two headings); prompts use one level per section.
    for num, _, _ in SECTION_SPEC:
        r = RUN_DIR / f"{num}_response.md"
        text = _strip_bom_and_zwsp(r.read_text(encoding="utf-8"))
        for msg in _duplicate_md_heading_issues(text, num):
            errors.append(f"{r.name}: {msg}")

    t03a: str | None = None
    r03_alpha = RUN_DIR / "03_response.md"
    if r03_alpha.exists():
        t03a = _strip_bom_and_zwsp(r03_alpha.read_text(encoding="utf-8"))
        for msg in _03_parameters_sorted_alphabetically_issues(t03a):
            errors.append(msg)

    r01 = RUN_DIR / "01_response.md"
    r02 = RUN_DIR / "02_response.md"
    r07 = RUN_DIR / "07_response.md"
    text03_for_0102: str | None = t03a
    text07_for_0102: str | None = None
    if r07.exists():
        text07_for_0102 = _strip_bom_and_zwsp(r07.read_text(encoding="utf-8"))
    forbidden_0102 = _01_02_forbidden_name_tokens(text03_for_0102, text07_for_0102)
    if r01.exists():
        for msg in _01_02_business_language_issues(
            _strip_bom_and_zwsp(r01.read_text(encoding="utf-8")),
            "01_response.md",
            forbidden_0102,
        ):
            errors.append(msg)
    if r02.exists():
        text02 = _strip_bom_and_zwsp(r02.read_text(encoding="utf-8"))
        for msg in _02_suggested_resolution_forbidden_subsection_issues(text02):
            errors.append(msg)
        for msg in _01_02_business_language_issues(
            text02,
            "02_response.md",
            forbidden_0102,
        ):
            errors.append(msg)

    # 04_response.md: forbidden phrases and parameter count match
    r04 = RUN_DIR / "04_response.md"
    r03 = RUN_DIR / "03_response.md"
    if r04.exists():
        text04 = _strip_bom_and_zwsp(r04.read_text(encoding="utf-8"))
        lines04 = text04.splitlines()
        for phrase in _04_FORBIDDEN_PHRASES:
            for i, line in enumerate(lines04, 1):
                if phrase.lower() in line.lower():
                    errors.append(f"04_response.md line {i}: forbidden phrase '{phrase}' (do not use 'output only' / 'not a filter' wording)")
                    break  # one error per phrase type

        for name in _04_FORBIDDEN_INTERNAL_NAMES:
            for i, line in enumerate(lines04, 1):
                if name in line:
                    errors.append(
                        f"04_response.md line {i}: forbidden internal name '{name}' "
                        "(date/time params: business meaning only; no R_DATUM, SY_DATLO, DATE_FROM, etc.)"
                    )
                    break

        # Generic/filler phrases: fail if same phrase appears too often (no useless repetition)
        text04_lower = text04.lower()
        for phrase, max_occ in _04_GENERIC_PHRASE_MAX_OCCURRENCES:
            count = text04_lower.count(phrase.lower())
            if count > max_occ:
                errors.append(
                    f"04_response.md: generic phrase '{phrase}' appears {count} times (max {max_occ}). "
                    "Use function- or parameter-specific content; see prompt rule 'No generic or useless sentences'."
                )
        for pat in _04_FORBIDDEN_TEMPLATE_PATTERNS:
            if pat.search(text04):
                errors.append(
                    "04_response.md: placeholder/template sentence pattern detected "
                    "(e.g., '... is used with its business meaning ...'). "
                    "Use parameter-specific explanations with concrete EI behavior."
                )
                break
        for msg in _04_reused_main_sentence_issues(text04):
            errors.append(msg)
        for msg in _04_confusable_param_differentiation_issues(text04):
            errors.append(msg)

        if r03.exists():
            text03 = _strip_bom_and_zwsp(r03.read_text(encoding="utf-8"))
            param_names_ordered, _row_count = _param_names_ordered_from_03_table(text03)
            expected_count = len(param_names_ordered)
            params_in_03 = set(param_names_ordered)
            for msg in _04_backdays_wording_issues(text04, params_in_03):
                errors.append(msg)
            for msg in _04_forwdays_wording_issues(text04, params_in_03):
                errors.append(msg)
            for msg in _04_date_ref_fld_guideline_issues(text04, params_in_03):
                errors.append(msg)
            for msg in _04_duration_unit_options_issues(text04, params_in_03):
                errors.append(msg)
            for msg in _04_aggr_level_guideline_issues(text04, params_in_03):
                errors.append(msg)
            for msg in _04_params_dictionary_canonical_issues(text04, text03):
                errors.append(msg)
            for msg in _04_dictionary_mark_when_unchecked_issues(text04, text03):
                errors.append(msg)
            for msg in _04_dictionary_mark_when_checked_issues(text04, text03):
                errors.append(msg)
            for msg in _04_user_fld_guideline_issues(text04, params_in_03):
                errors.append(msg)
            # Serial-number series in 03 must be grouped in 04 (e.g. UVK01–UVK05 as one entry)
            for first, last in _serial_series_from_03_param_names(param_names_ordered):
                if not _serial_group_heading_present(text04, first, last):
                    gh = f"**{first} - {last}**"
                    gs = f"**{first}/{last}**"
                    errors.append(
                        f"04_response.md: serial-number series {first}..{last} must be grouped into one entry "
                        f'(e.g. "{gh}", "{gs}", or en-dash between names)'
                    )
            # Parallel TAB1_* / TAB2_* slot groups (1–5): one combined guideline block or two grouped blocks each
            for stem1, stem2, lo, hi, suf in _PARALLEL_TAB_SLOT_GROUPS:
                a, b = _parallel_slot_param_names(stem1, stem2, lo, hi, suf)
                if not all(n in params_in_03 for n in a + b):
                    continue
                f1, l1 = a[0], a[-1]
                f2, l2 = b[0], b[-1]
                if not _parallel_tab12_heading_ok(text04, f1, l1, f2, l2):
                    ex = f"**{f1} - {l1} / {f2} - {l2}**"
                    ab = f"**{stem1}{lo}-{hi}{suf} / {stem2}{lo}-{hi}{suf}**"
                    errors.append(
                        f"04_response.md: parallel table parameters {f1}..{l1} and {f2}..{l2} must be grouped "
                        f'(e.g. "{ex}" or abbreviated "{ab}")'
                    )
            # Fixed-option params: **BASE Options:** or **BASE1 Options:** when BASE or BASE1 exists
            for base in _04_PARAMS_REQUIRING_OPTIONS:
                has_suffixed = any(re.match(rf"^{re.escape(base)}\d+$", p) for p in params_in_03)
                if has_suffixed:
                    if not re.search(rf"\*\*{re.escape(base)}\d+\s+Options:\*\*", text04):
                        errors.append(
                            f"04_response.md: parameters like {base}1 must have an Options subsection "
                            f'(e.g. "**{base}1 Options:**" per suffixed parameter)'
                        )
                elif base in params_in_03:
                    marker = f"**{base} Options:**"
                    if marker not in text04:
                        errors.append(
                            f"04_response.md: {base} must have an Options subsection "
                            f"(fixed-option parameter; add '{marker}' with possible values)"
                        )
            # Global mandatory option parameters.
            for p in _04_EXTRA_OPTIONS_PARAMS:
                if p in params_in_03 and not _04_options_heading_present(text04, p):
                    errors.append(
                        f"04_response.md: {p} must have an Options subsection "
                        f"(add '**{p} Options:**' with values)."
                    )
            for p in _04_STATUS_SELECTOR_PARAMS:
                if p in params_in_03 and not _04_options_heading_present(text04, p):
                    errors.append(
                        f"04_response.md: {p} must have an Options subsection "
                        f"(status selector values must be documented)."
                    )
            for p in sorted(_04_debit_credit_indicator_params_from_03(text03, params_in_03)):
                if not _04_options_heading_present(text04, p):
                    errors.append(
                        f"04_response.md: {p} is a debit/credit indicator parameter; add '**{p} Options:**' "
                        "with possible values (code-derived first; else standard SAP S/H debit-credit letters unless "
                        "domain differs; see Parameter Configuration Guidelines prompt)."
                    )
            # MANAGE_IN_UTC must document both UTC ('X') and local/empty mode.
            if "MANAGE_IN_UTC" in params_in_03 and _04_options_heading_present(text04, "MANAGE_IN_UTC"):
                b = _04_options_block_text(text04, "MANAGE_IN_UTC").lower()
                has_x = re.search(r"\b\*\*?x\*\*?\b|\bx\b", b) is not None
                has_local_empty = any(k in b for k in ["empty", "space", "local time", "initial", "blank"])
                if not (has_x and has_local_empty):
                    errors.append(
                        "04_response.md: MANAGE_IN_UTC Options must include both 'X' (UTC mode) "
                        "and empty/blank/local-time mode."
                    )
            unused_04 = _resolve_unused_params_set()
            for msg in _04_not_in_use_marker_issues(text04, text03, unused_04):
                errors.append(msg)
            # Check 04 IMPORTANT line matches expanded parameter count from 03
            match = re.search(
                r"(?:ALL|defines)\s+(\d+)\s+parameters", text04, re.IGNORECASE
            )
            if match and expected_count > 0:
                n_in_04 = int(match.group(1))
                if n_in_04 != expected_count:
                    errors.append(
                        f"04_response.md IMPORTANT line says {n_in_04} parameters but 03 expands to {expected_count} parameters"
                    )
            elif expected_count > 0 and not re.search(
                r"IMPORTANT.*\d+.*parameters", text04, re.IGNORECASE
            ):
                errors.append(
                    "04_response.md missing IMPORTANT line with parameter count (e.g. defines N parameters)"
                )

    # 05_response.md formatting: title must be H3; "How parameter combinations..." must be normal text (not heading)
    r05 = RUN_DIR / "05_response.md"
    if r05.exists():
        text05 = _strip_bom_and_zwsp(r05.read_text(encoding="utf-8"))
        if not re.search(r"^###\s+Parameter Relationship(s)?\s*$", text05, re.MULTILINE):
            errors.append("05_response.md: heading must be H3: '### Parameter Relationship' (or '### Parameter Relationships').")
        if re.search(r"^#{1,6}\s+How parameter combinations work together\s*$", text05, re.MULTILINE | re.IGNORECASE):
            errors.append("05_response.md: 'How parameter combinations work together' must be normal text, not a heading.")
        unused_05 = _resolve_unused_params_set()
        if r03.exists():
            text03 = _strip_bom_and_zwsp(r03.read_text(encoding="utf-8"))
            param_names_ordered_05, _ = _param_names_ordered_from_03_table(text03)
            params_in_03_05 = set(param_names_ordered_05)
            for msg in _05_time_filter_clarity_issues(text05, params_in_03_05, unused_05):
                errors.append(msg)
        for msg in _05_combined_effect_block_issues(text05):
            errors.append(msg)
        for msg in _05_unused_param_presence_issues(text05, unused_05):
            errors.append(msg)

    # 06_response.md: title/heading formatting + practical examples rules
    r06 = RUN_DIR / "06_response.md"
    if r06.exists():
        text06 = _strip_bom_and_zwsp(r06.read_text(encoding="utf-8"))
        if "Default Values (if available)" in text06:
            errors.append(
                "06_response.md: rename heading to '### Default Values' (do not use 'Default Values (if available)')."
            )
        if not re.search(r"^###\s+Default Values\s*$", text06, re.MULTILINE):
            errors.append("06_response.md: heading must be '### Default Values'.")
        if not re.search(r"^###\s+Practical Example of Parameter Configuration\s*$", text06, re.MULTILINE):
            errors.append("06_response.md: heading must be '### Practical Example of Parameter Configuration'.")
        if re.search(r"^#{1,6}\s+Use Case\s+\d+\s*:", text06, re.MULTILINE):
            errors.append("06_response.md: Use Case titles must be normal text (not markdown headings).")
        if re.search(r"^\s*`\s*$", text06, re.MULTILINE):
            errors.append("06_response.md: malformed code fence '`' found; use triple backticks for each example block.")
        for msg in _06_default_values_format_issues(text06):
            errors.append(msg)
        unused_06 = _resolve_unused_params_set()
        r03 = RUN_DIR / "03_response.md"
        if r03.exists():
            text03_06 = _strip_bom_and_zwsp(r03.read_text(encoding="utf-8"))
            for msg in _06_initial_runtime_default_bullets_issues(text03_06, text06, unused_06):
                errors.append(msg)
            for msg in _06_initial_runtime_effect_clarity_issues(text03_06, text06, unused_06):
                errors.append(msg)
        for msg in _06_unused_param_presence_issues(text06, unused_06):
            errors.append(msg)
        param_counts = _06_practical_example_param_counts(text06)
        bold_use_cases, non_bold_use_cases = _06_use_case_title_counts(text06)
        if non_bold_use_cases > 0:
            errors.append(
                "06_response.md: Use Case titles must be bold lines "
                "(format: '**Use Case N: ...**')."
            )
        if bold_use_cases == 0:
            errors.append("06_response.md: missing use-case titles; add at least one '**Use Case N: ...**' line.")
        for msg in _06_purpose_before_code_issues(text06):
            errors.append(msg)
        purpose_count = _06_purpose_count(text06)
        if bold_use_cases > 0 and purpose_count < bold_use_cases:
            errors.append(
                f"06_response.md: found {bold_use_cases} use case title(s) but only {purpose_count} "
                "Purpose paragraph(s). Add one '**Purpose:** ...' per use case before its ``` parameter block "
                "(after a blank line following each **Use Case N: …** line)."
            )
        for i, n in enumerate(param_counts):
            if n < 2:
                errors.append(
                    f"06_response.md: Use Case {i + 1} has {n} parameter(s) in its code block; "
                    "each use case must have at least 2 parameters."
                )
        if param_counts and not any(3 <= n <= 5 for n in param_counts) and not any(n >= 6 for n in param_counts):
            # Allow 6+ as "rich" too; require at least one use case with 3+ params
            if not any(n >= 3 for n in param_counts):
                errors.append(
                    "06_response.md: No use case has 3–5 (or more) parameters. "
                    "At least one practical configuration example must include 3–5 parameters in its code block."
                )
        for msg in _06_no_range_time_param_issues(text06):
            errors.append(msg)
        if r03.exists():
            text03_06b = _strip_bom_and_zwsp(r03.read_text(encoding="utf-8"))
            ordered_06, _ = _param_names_ordered_from_03_table(text03_06b)
            for msg in _06_duration_unit_f_purpose_issues(text06, set(ordered_06), unused_06):
                errors.append(msg)

    return errors


# Sections where the canonical Markdown heading is ## (top-level document sections).
_HEADING_PREFER_H2_SECTIONS = frozenset({"01", "02", "07"})
# Sections 03–06 use ### subsections (under an implicit/benchmark "Parameters" block).


def _duplicate_md_heading_issues(md: str, section_num: str) -> list[str]:
    """Detect back-to-back `### Title` and `## Title` with the same title; return human-readable errors."""
    prefer_h2 = section_num in _HEADING_PREFER_H2_SECTIONS
    lines = md.splitlines()
    issues: list[str] = []
    i = 0
    while i < len(lines):
        s = lines[i].strip()
        if s.startswith("### ") and not s.startswith("####"):
            title_h3 = s[4:].strip()
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j < len(lines):
                s2 = lines[j].strip()
                if s2.startswith("## ") and not s2.startswith("###"):
                    title_h2 = s2[3:].strip()
                    if title_h3 == title_h2:
                        fix = (
                            "remove the `###` line (keep `##` only)"
                            if prefer_h2
                            else "remove the `##` line (keep `###` only)"
                        )
                        issues.append(
                            f"duplicate heading '{title_h3}' (~line {i + 1}: `###` then `##`); {fix}"
                        )
        i += 1
    return issues


def _dedupe_duplicate_md_section_headings(md: str, section_num: str) -> str:
    """Resolve `### Title` + `## Title` duplicate: keep the level required for this section (see prompts / benchmark)."""
    prefer_h2 = section_num in _HEADING_PREFER_H2_SECTIONS
    lines = md.splitlines(keepends=True)
    out: list[str] = []
    i = 0
    while i < len(lines):
        line = lines[i]
        s = line.strip()
        if s.startswith("### ") and not s.startswith("####"):
            title_h3 = s[4:].strip()
            j = i + 1
            while j < len(lines) and not lines[j].strip():
                j += 1
            if j < len(lines):
                s2 = lines[j].strip()
                if s2.startswith("## ") and not s2.startswith("###"):
                    title_h2 = s2[3:].strip()
                    if title_h3 == title_h2:
                        if prefer_h2:
                            i = j
                            continue
                        out.append(line)
                        i = j + 1
                        continue
        out.append(line)
        i += 1
    return "".join(out)


# 01/02: business-oriented sections must not name input parameters or output fields (see prompts 01/02).
_01_02_GLOBAL_FORBIDDEN_NAMES = frozenset(
    {
        "T_DATA",
        "R_DATUM",
        "R_UDATE",
        "R_UDATE_REPET",
        "R_AEDAT",
        "R_BEDAT",
        "R_DURATION",
        "SY_DATLO",
        "SY_TIMLO",
        "SY_DATUM",
        "SY_UZEIT",
        "DATE_FROM",
        "DATE_REF_FLD",
        "TIME_REF_FLD",
        "DURATION_UNIT",
        "FORWDAYS",
        "REPET_BACKDAYS",
        "UDATE_REPET",
        "USER_FLD",
        "USR_FLD",
    }
)


def _output_field_names_from_07_table(text07: str) -> list[str]:
    """Field Name column from the EI Function Structure markdown table in 07."""
    names: list[str] = []
    for line in text07.splitlines():
        if not line.strip().startswith("|"):
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 4:
            continue
        if parts[1].lower() in ("structure name", "---") or parts[2].lower() == "field name":
            continue
        fld = parts[2].strip().strip("`")
        if fld and fld.lower() not in ("field name",):
            names.append(fld)
    return names


def _01_02_forbidden_name_tokens(
    text03: str | None,
    text07: str | None,
) -> set[str]:
    """Union of 03 parameter names, 07 output field names, and global blocklist for 01/02 checks."""
    tokens: set[str] = set(_01_02_GLOBAL_FORBIDDEN_NAMES)
    if text03:
        for name in _param_names_ordered_from_03_table(text03)[0]:
            u = name.strip().upper()
            if u:
                tokens.add(u)
    if text07:
        for name in _output_field_names_from_07_table(text07):
            u = name.strip().upper()
            if u:
                tokens.add(u)
    return tokens


def _01_02_business_language_issues(
    md: str,
    response_name: str,
    forbidden_tokens: set[str],
) -> list[str]:
    """
    01_response.md / 02_response.md: no parameter or output-field tokens from 03/07;
    business language only (see PROMPT_General_Overview_section.md and
    PROMPT_Problem_Description_and_Suggested_Resolution_section.md).
    """
    if not forbidden_tokens:
        return []
    issues: list[str] = []
    # Longest names first so nested tokens (e.g. DURATION vs DURATION_UNIT) report the specific match.
    ordered = sorted(forbidden_tokens, key=len, reverse=True)
    seen_lines: set[int] = set()
    for line_no, line in enumerate(md.splitlines(), 1):
        for token in ordered:
            if re.search(rf"\*\*{re.escape(token)}\*\*", line, re.IGNORECASE):
                if line_no not in seen_lines:
                    seen_lines.add(line_no)
                    issues.append(
                        f"{response_name} line {line_no}: business sections must not name parameters or "
                        f"output fields ({token!r}); use business language — see "
                        "prompts/PROMPT_General_Overview_section.md and "
                        "prompts/PROMPT_Problem_Description_and_Suggested_Resolution_section.md."
                    )
                break
            if re.search(rf"\b{re.escape(token)}\b", line, re.IGNORECASE):
                if line_no not in seen_lines:
                    seen_lines.add(line_no)
                    issues.append(
                        f"{response_name} line {line_no}: business sections must not name parameters or "
                        f"output fields ({token!r}); use business language — see "
                        "prompts/PROMPT_General_Overview_section.md and "
                        "prompts/PROMPT_Problem_Description_and_Suggested_Resolution_section.md."
                    )
                break
    return issues


_02_SUGGESTED_RES_FORBIDDEN_SUBSECTION = re.compile(
    r"^\*\*(Training|User training|Education|User enablement|Process Improvements)\b",
    re.MULTILINE | re.IGNORECASE,
)


def _02_suggested_resolution_forbidden_subsection_issues(md: str) -> list[str]:
    """02_response.md: under Suggested Resolution, disallow extra bold subsections (e.g. **Training**)."""
    issues: list[str] = []
    m = re.search(r"^###\s+Suggested Resolution\s*$", md, re.MULTILINE | re.IGNORECASE)
    if not m:
        return issues
    chunk = md[m.end() :]
    m2 = re.search(r"^###\s+\S", chunk, re.MULTILINE)
    if m2:
        chunk = chunk[: m2.start()]
    match = _02_SUGGESTED_RES_FORBIDDEN_SUBSECTION.search(chunk)
    if match:
        line_end = chunk.find("\n", match.start())
        preview = chunk[match.start() : line_end if line_end != -1 else len(chunk)].strip()[:100]
        issues.append(
            "02_response.md: Suggested Resolution must use only **Immediate Response**, **System Assessment**, "
            "**Corrective Actions**—no separate **Training** / **Education** / **User enablement** / **Process Improvements** subsection; "
            f"fold enablement into Corrective Actions. Found: {preview!r}"
        )
    return issues


def _06_practical_example_param_counts(text: str) -> list[int]:
    """Parse 06_response.md Practical Configuration Examples; return list of parameter counts per use case code block."""
    counts = []
    in_practical_section = False
    in_block = False
    for line in text.splitlines():
        if re.match(r"^###\s+Practical", line, re.IGNORECASE):
            in_practical_section = True
            continue
        if not in_practical_section:
            continue
        if line.strip().startswith("```"):
            if in_block:
                in_block = False
            else:
                in_block = True
                counts.append(0)
            continue
        if in_block:
            # Count lines that look like PARAM = value (parameter assignment)
            s = line.strip()
            if s and "=" in s and not s.startswith("#") and re.match(r"^[A-Za-z0-9_]+\s*=", s):
                counts[-1] += 1
    return counts


def _06_default_values_format_issues(text: str) -> list[str]:
    """
    Validate Default Values block format in 06_response.md.
    Required bullet styles:
    - Explicit: `- **PARAM** - value`
    - Initial-runtime (BACKDAYS/DURATION/DURATION_UNIT/AGGLEVEL): `- **PARAM** - initial - treated as … by code`
    """
    issues: list[str] = []
    m = re.search(r"^###\s+Default Values\s*$", text, re.MULTILINE)
    if not m:
        return issues
    chunk = text[m.end() :]
    m2 = re.search(r"^###\s+Practical Example of Parameter Configuration\s*$", chunk, re.MULTILINE)
    if m2:
        chunk = chunk[: m2.start()]
    lines = [ln.rstrip() for ln in chunk.splitlines() if ln.strip()]
    if not lines:
        issues.append("06_response.md: Default Values section is empty.")
        return issues
    if len(lines) == 1 and lines[0].lower().startswith("no default values"):
        return issues

    explicit_bullet_re = re.compile(r"^\-\s+\*\*[A-Za-z0-9_ ]+\*\*\s+\-\s+.+$")
    initial_runtime_bullet_re = re.compile(
        r"^\-\s+\*\*[A-Za-z0-9_ ]+\*\*\s+\-\s+initial\s*[\u2013\u2014-]\s*treated\s+as\s+.+\bby\s+code\b.*$"
    )
    for i, ln in enumerate(lines, 1):
        if not ln.startswith("- "):
            continue
        if "? Default" in ln or "— Default:" in ln or " - Default:" in ln:
            issues.append(
                "06_response.md: Default Values bullets must use '- **PARAM** - value' "
                "(do not use '? Default:' or '— Default:')."
            )
            break
        if "`" in ln:
            issues.append(
                "06_response.md: Default Values bullets must not use backticks around values; use plain text."
            )
            break
        if "(" in ln or ")" in ln or ";" in ln:
            issues.append(
                f"06_response.md: Default Values bullet #{i} is too wordy; keep only value-style text "
                "without parentheses/semicolon explanations."
            )
            break
        if len(ln.strip()) > 95:
            issues.append(
                f"06_response.md: Default Values bullet #{i} is too long; keep concise value-only wording."
            )
            break
        if initial_runtime_bullet_re.match(ln) is None and explicit_bullet_re.match(ln) is None:
            issues.append(
                f"06_response.md: Default Values bullet format invalid at bullet #{i}; "
                "use '- **PARAM** - value' or '- **PARAM** - initial - treated as … by code'."
            )
            break
    return issues


def _06_practical_section_text(text: str) -> str:
    """Return content of Practical Example section from 06_response.md."""
    m = re.search(r"^###\s+Practical Example of Parameter Configuration\s*$", text, re.MULTILINE)
    if not m:
        return ""
    return text[m.end() :]


# After **Use Case N: …** title line, **Purpose:** must start its own paragraph: at least one blank line
# (two line breaks) before the **Purpose:** line so Markdown/Word export keeps sub-blocks visually separate.
_06_PURPOSE_AFTER_TITLE_PATTERN = re.compile(
    r"^\s*(?:\r?\n[ \t]*){2,}\s*\*\*Purpose:\*\*",
    re.MULTILINE,
)


def _06_purpose_before_code_issues(text: str) -> list[str]:
    """
    Each practical use case must list **Purpose:** after the title and before the opening ``` fence.
    **Purpose:** must begin on a new line after a blank line (not directly under the title line).
    """
    issues: list[str] = []
    chunk = _06_practical_section_text(text)
    if not chunk.strip():
        return issues
    # Do not use \s* before $ here: \s would consume the newline after the title and break the
    # "blank line before **Purpose:**" check (body would start with only one \n).
    pat = re.compile(r"(?m)^\*\*Use Case\s+(\d+)\s*:[^\n]*\*\*[ \t]*$")
    matches = list(pat.finditer(chunk))
    if not matches:
        return issues
    for i, m in enumerate(matches):
        body_start = m.end()
        body_end = matches[i + 1].start() if i + 1 < len(matches) else len(chunk)
        body = chunk[body_start:body_end]
        num = m.group(1)
        fence_pos = body.find("```")
        if fence_pos == -1:
            issues.append(
                f"06_response.md: Use Case {num}: missing opening ``` code fence for the parameter block."
            )
            continue
        before_fence = body[:fence_pos]
        if "**Purpose:**" not in before_fence:
            issues.append(
                f"06_response.md: Use Case {num}: add **Purpose:** after the use case title and before the ``` parameter block."
            )
            continue
        if _06_PURPOSE_AFTER_TITLE_PATTERN.match(before_fence) is None:
            issues.append(
                f"06_response.md: Use Case {num}: put **Purpose:** on its own line after a blank line "
                "(end the **Use Case N: …** line, add one empty line, then **Purpose:** …) so the Purpose sub-block "
                "is separated from the title for export."
            )
    return issues


def _06_use_case_title_counts(text: str) -> tuple[int, int]:
    """Return (bold_use_cases, non_bold_use_cases) in practical section."""
    chunk = _06_practical_section_text(text)
    if not chunk:
        return (0, 0)
    bold = len(re.findall(r"^\*\*Use Case\s+\d+\s*:", chunk, re.MULTILINE))
    non_bold = len(re.findall(r"^(?!\*\*)Use Case\s+\d+\s*:", chunk, re.MULTILINE))
    return (bold, non_bold)


def _06_purpose_count(text: str) -> int:
    """Return number of **Purpose:** lines in practical section."""
    chunk = _06_practical_section_text(text)
    if not chunk:
        return 0
    return len(re.findall(r"^\*\*Purpose:\*\*", chunk, re.MULTILINE))


def _strip_bom_and_zwsp(text: str) -> str:
    """Remove BOM/zero-width chars that break markdown parsing when concatenated."""
    if not text:
        return text
    return text.replace("\ufeff", "").replace("\u200b", "")


def _normalize_response_markdown(text: str) -> str:
    """Normalize response markdown to be safe for assemble/export."""
    text = _strip_bom_and_zwsp(text)
    # Some generators accidentally emit single-backtick fence lines.
    lines = text.splitlines()
    lines = ["```" if line.strip() == "`" else line for line in lines]
    return "\n".join(lines) + ("\n" if text.endswith("\n") else "")


def _write_manifest_at(basename: str, title: str) -> None:
    """Write manifest.txt with output_basename and title."""
    manifest = RUN_DIR / "manifest.txt"
    manifest.write_text(f"output_basename={basename}\ntitle={title}\n", encoding="utf-8")


def _write_manifest(paths: dict) -> None:
    """Re-read Metadata (General/Metadata general sheet), compute basename and title, write manifest."""
    import openpyxl
    wb = openpyxl.load_workbook(paths["metadata"], read_only=True)
    ws = _metadata_sheet(wb)
    rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
    wb.close()
    id_val = str(rows[7][1] or "").strip() if len(rows) > 8 and len(rows[7]) > 1 else ""
    name_val = str(rows[8][1] or "").strip() if len(rows) >= 9 and len(rows[8]) > 1 else ""
    stem = _stem_from_path(paths["metadata"], PREFIXES["metadata"])
    basename = stem
    title = _manifest_document_title(name_val=name_val, id_val=id_val, basename=basename)
    _write_manifest_at(basename, title)


def _normalize_text_for_dictionary_compare(s: str) -> str:
    """Collapse whitespace and strip for exact dictionary match checks in section 04."""
    s = _strip_bom_and_zwsp(s)
    # Ignore <mark> wrappers used for dictionary-only (unchecked) highlighting in 04.
    s = re.sub(r"(?is)</?mark\b[^>]*>", "", s)
    s = re.sub(r"\s+", " ", s.strip())
    return s


def _load_params_dictionary_explanations(path: Path | None = None) -> dict[str, str]:
    """
    Load parameter name -> explanation from input/params_dictionary.xlsx (sheet 'dictionary' if present).
    Keys are uppercased Field names. Second column is the canonical explanation text.
    """
    p = path or PARAMS_DICTIONARY_PATH
    if not p.exists():
        return {}
    try:
        import openpyxl

        wb = openpyxl.load_workbook(p, read_only=True)
        sn = "dictionary" if "dictionary" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return {}
    if not rows:
        return {}
    out: dict[str, str] = {}
    start = 0
    if rows[0] and str(rows[0][0] or "").strip().lower() in ("parameter", "field", "param"):
        start = 1
    for row in rows[start:]:
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip().upper()
        if not key:
            continue
        raw = row[1] if len(row) > 1 else ""
        text = str(raw or "").strip()
        if not text:
            continue
        # Repair common mojibake from Excel/CSV round-trips
        text = text.replace("\ufffd", "'").replace("\u2019", "'").replace("\u2013", "-").replace("\u2014", "-")
        out[key] = text
    return out


def _load_checked_params_set() -> frozenset[str] | None:
    """
    Load uppercased parameter names from input/checked params.txt (one name per line).
    Returns None if the file is missing or empty → yellow highlighting is disabled (dictionary text still applies).
    Returns a frozenset if the file exists and has at least one non-comment, non-blank line.
    """
    p = CHECKED_PARAMS_PATH
    if not p.exists():
        return None
    try:
        raw = p.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return None
    names: set[str] = set()
    for line in raw.splitlines():
        s = line.split("#", 1)[0].strip().upper()
        if s:
            names.add(s)
    if not names:
        return None
    return frozenset(names)


def _04_wrap_dictionary_explanation(text: str, param_field: str, checked: frozenset[str] | None) -> str:
    """
    If checked-params is active and param_field is in params_dictionary but not in checked list,
    wrap explanation in <mark> for yellow highlight in Word (see md_to_docx). Otherwise return text as-is.
    """
    if not (text and text.strip()):
        return text
    if checked is None:
        return text
    key = str(param_field).strip().upper()
    if key in checked:
        return text
    inner = html.escape(text.strip(), quote=False)
    return f"<mark>{inner}</mark>"


# Rotating templates for auto-generated 04 (must stay varied enough for verify_responses reuse checks).
_AUTO_04_TEMPLATES = [
    "Narrows retrieved rows where {desc} ({fld}) must match the configured selection for this monitor.",
    "After data is read, lines are removed unless {desc} on {fld} still satisfies the active multivalued selection.",
    "Supports operational control by evaluating {desc} through {fld} for each candidate record.",
    "When populated, keeps the extract focused so {desc} ({fld}) aligns with the intended triage slice.",
    "Combines with related filters so {desc} on {fld} refines which records remain for duration or state checks.",
    "Uses {desc} from the source context so only records with {fld} inside declared values proceed.",
    "Aligns exception volume with the chosen scope by testing {desc} via {fld} before alert evaluation.",
    "Helps monitoring stay readable by requiring {desc} ({fld}) to match organizational or technical selectors when set.",
    "Interprets {desc} as part of the selection contract: open ranges follow framework defaults; restricted ranges apply strict matching on {fld}.",
    "Guards against oversized extracts when {desc} on {fld} is narrowed together with client, user, or session filters.",
    "For operations, {desc} on {fld} indicates whether a row belongs in the current monitoring pass versus historical noise.",
    "When tightened, {desc} ({fld}) removes rows that would otherwise dilute attention from failing or stuck cases.",
    "Pairs with duration logic: once {fld} passes list selection, elapsed time from the reference timestamp still must fit configured duration windows.",
    "Separates cross-client noise from in-scope work when {desc} on {fld} correlates with client or user attributes.",
    "Ensures reporting respects {desc} constraints carried by {fld}.",
    "Treats {desc} as a discriminator between similar rows that would otherwise look identical in a raw extract.",
    "When left open per framework rules, {fld} does not restrict {desc}; when set, only matching rows remain.",
    "Supports escalation where {desc} on {fld} signals ownership for follow-up between Basis and functional teams.",
    "Reduces false positives during peak windows by tightening {desc} through {fld} alongside state filters.",
    "Gives auditors traceable criteria because {desc} on {fld} is applied consistently before any alert flag is raised.",
    "Mirrors how administrators slice operational lists: {desc} ({fld}) is one lever that shapes which rows are comparable run over run.",
    "For distributed landscapes, {desc} on {fld} often anchors which application server or destination appears in results.",
    "When harmonized with related filters, {desc} on {fld} isolates the highest-risk record families.",
    "Captures edge cases where {desc} ({fld}) must be non-default to reproduce a customer-specific monitoring scenario.",
    "Works downstream of the initial read so {desc} on {fld} still participates in row-level deletion rules.",
    "Allows phased rollout: first widen {fld} for {desc}, then tighten thresholds once baseline noise is understood.",
    "Explains why two monitoring passes differ: only the pass with stricter {desc} on {fld} surfaces the disputed rows.",
    "Documents expected operator behavior—{desc} on {fld} should be set when that dimension is part of the control objective.",
    "Prevents accidental global scans when {desc} ({fld}) is meant to stay within a controlled application slice.",
    "Connects to alert semantics: rows removed for failing {desc} on {fld} never reach downstream filtering.",
    "Valuable when comparing health before and after a release—hold {desc} on {fld} constant while varying other filters.",
    "Improves readability of exported lists because {desc} ({fld}) columns stay aligned with the configured filter intent.",
    "When combined with destination discipline, {desc} on {fld} keeps both breadth and depth of the extract intentional.",
    "Reflects real administration where {desc} on {fld} is routinely restricted to a single productive client or object family.",
    "Stabilizes week-over-week metrics by fixing {desc} ({fld}) while allowing duration thresholds to move.",
    "Helps distinguish technical versus business attributes when {desc} on {fld} correlates with counters or status fields.",
]


def _auto_04_opener(i: int, fld: str, desc: str) -> str:
    h = int(hashlib.md5(fld.encode()).hexdigest(), 16)
    t = _AUTO_04_TEMPLATES[(i * 13 + h) % len(_AUTO_04_TEMPLATES)]
    return t.format(desc=(desc or fld).lower(), fld=fld)


def _load_parameter_rows_from_available_fields_xlsx(params_path: Path) -> list[list[str]]:
    import openpyxl

    wb = openpyxl.load_workbook(params_path, read_only=True)
    if "Parameters" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"No 'Parameters' sheet in {params_path}")
    ws = wb["Parameters"]
    rows: list[list[str]] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] is None or str(r[0]).strip() == "":
            break
        rows.append([str(c).strip() if c is not None else "" for c in r[:7]])
    wb.close()
    # Standard project convention: Parameters Reference Table is A–Z by Field (matches verify_responses).
    rows.sort(key=lambda row: (str(row[0]).strip().lower(),) if row else ("",))
    return rows


def _markdown_03_from_parameter_rows(rows: list[list[str]]) -> str:
    lines03 = [
        "### Parameters Reference Table",
        "",
        "This table lists all configurable input parameters. Users set values for these parameters to filter and control the EI's data selection and processing logic.",
        "",
        "| # | Field | Description | Type | Length | Decimal | Data Element | Domain |",
        "|---|-------|-------------|------|--------|---------|--------------|--------|",
    ]
    for i, r in enumerate(rows, start=1):
        lines03.append("| " + " | ".join([str(i)] + r) + " |")
    return "\n".join(lines03) + "\n"


def _markdown_04_from_parameter_rows(
    rows: list[list[str]],
    important_n: int,
    params_in_03: set[str],
    code_path: Path,
    params_path: Path,
) -> str:
    unused = analyze_unused_params(
        code_path, params_path, search_dirs=[INPUT_DIR, INPUT_DIR / "old"]
    )
    parts: list[str] = [
        "### Parameter Configuration Guidelines",
        "",
        f"IMPORTANT: This EI defines {important_n} parameters in the Parameters Reference Table. "
        "Configure parameters that affect selection and alerting; parameters marked "
        f"{NOT_IN_USE_LINE} are declared in the interface but do not change results for this EI.",
        "",
    ]
    code_text = ""
    if code_path.exists():
        try:
            code_text = code_path.read_text(encoding="utf-8", errors="replace")
        except OSError:
            code_text = ""
    date_ref_literals = _date_ref_fld_literals_from_code(code_text)
    aggr_level_literals = _aggr_level_literals_from_code(code_text)
    param_canon = _load_params_dictionary_explanations()
    checked_params = _load_checked_params_set()
    desc_by_field_upper: dict[str, str] = {}
    for row in rows:
        if not row or not row[0]:
            continue
        k = str(row[0]).strip().upper()
        d = str(row[1] or row[0]).strip()
        if k and k not in desc_by_field_upper:
            desc_by_field_upper[k] = d
    ordered_names = [str(row[0]).strip() for row in rows if row and row[0]]
    serial_series = _serial_series_from_03_param_names(ordered_names)
    serial_skip = _serial_series_members_to_skip(ordered_names, serial_series)
    for i, r in enumerate(rows):
        fld, desc = r[0], (r[1] or r[0]).strip()
        if fld in serial_skip:
            continue
        serial_head: tuple[str, str] | None = None
        for fr, lr in serial_series:
            if fld == fr:
                serial_head = (fr, lr)
                break
        if serial_head:
            fr, lr = serial_head
            parts.append(f"**{fr} - {lr}** ({desc})")
        else:
            parts.append(f"**{fld}** ({desc})")
        parts.append("")
        if fld == "BACKDAYS":
            mand = _BACKDAYS_WINDOW_SENTENCE_ASCII_DASH
            canon_b = param_canon.get("BACKDAYS")
            n_m = _normalize_text_for_dictionary_compare(mand)

            def _wrap_backdays(t: str) -> str:
                return _04_wrap_dictionary_explanation(t, "BACKDAYS", checked_params)

            if canon_b:
                n_c = _normalize_text_for_dictionary_compare(canon_b)
                if n_c == n_m:
                    parts.append(mand)
                elif n_m in n_c:
                    parts.append(_wrap_backdays(canon_b))
                else:
                    parts.append(_wrap_backdays(canon_b))
                    parts.append(mand)
            else:
                parts.append(mand)
            parts.append("")
            if "DATE_REF_FLD" in params_in_03:
                parts.append(_BACKDAYS_ANCHOR_DATE_REF_SENTENCE_EXACT)
            else:
                inf = _infer_backdays_anchor_field_from_code(code_text)
                if inf:
                    parts.append(f"Backdays is based on {inf}")
            parts.append("")
            continue
        if fld == "FORWDAYS":
            mand = _FORWDAYS_WINDOW_SENTENCE_ASCII_DASH
            canon_f = param_canon.get("FORWDAYS")
            n_m = _normalize_text_for_dictionary_compare(mand)

            def _wrap_forwdays(t: str) -> str:
                return _04_wrap_dictionary_explanation(t, "FORWDAYS", checked_params)

            if canon_f:
                n_c = _normalize_text_for_dictionary_compare(canon_f)
                if n_c == n_m:
                    parts.append(mand)
                elif n_m in n_c:
                    parts.append(_wrap_forwdays(canon_f))
                else:
                    parts.append(_wrap_forwdays(canon_f))
                    parts.append(mand)
            else:
                parts.append(mand)
            parts.append("")
            if "DATE_REF_FLD" in params_in_03:
                parts.append(_FORWDAYS_ANCHOR_DATE_REF_SENTENCE_EXACT)
            else:
                inf = _infer_forwdays_anchor_field_from_code(code_text)
                if inf:
                    parts.append(f"Forwdays is based on {inf}")
            parts.append("")
            continue
        if fld == "DURATION_UNIT":
            if param_canon.get("DURATION_UNIT"):
                parts.append(
                    _04_wrap_dictionary_explanation(
                        param_canon["DURATION_UNIT"], "DURATION_UNIT", checked_params
                    )
                )
            else:
                parts.append(
                    "Unit for elapsed time between each session's creation date and time and the evaluation clock."
                )
            parts.append("")
            _append_duration_unit_options_block(parts)
            continue
        if fld == "TIME_DIFF_UNIT":
            if param_canon.get("TIME_DIFF_UNIT"):
                parts.append(
                    _04_wrap_dictionary_explanation(
                        param_canon["TIME_DIFF_UNIT"], "TIME_DIFF_UNIT", checked_params
                    )
                )
            else:
                parts.append(
                    "Unit for time-difference comparisons used when the EI evaluates elapsed intervals between reference timestamps."
                )
            parts.append("")
            parts.append("**TIME_DIFF_UNIT Options:**")
            parts.append("- **H** — Hours.")
            parts.append("- **M** — Minutes.")
            parts.append("- **D** — Days.")
            parts.append("- **F** — Full-day style counting where applicable to the duration helper.")
            parts.append("")
            continue
        if fld == "STATE_COLOR":
            if param_canon.get("STATE_COLOR"):
                parts.append(
                    _04_wrap_dictionary_explanation(param_canon["STATE_COLOR"], "STATE_COLOR", checked_params)
                )
            else:
                parts.append(
                    "Filters lines by the derived color bucket used for severity-style triage in the monitor framework."
                )
            parts.append("")
            parts.append("**STATE_COLOR Options:**")
            parts.append("- **R** — Red (error or failed-style outcomes).")
            parts.append("- **G** — Green (successful outcomes).")
            parts.append("- **Y** — Yellow (warning or in-process outcomes).")
            parts.append("- Additional literals may exist where the framework extends the palette for neutral states.")
            parts.append("")
            continue
        if fld == "QSTATE":
            if param_canon.get("QSTATE"):
                parts.append(_04_wrap_dictionary_explanation(param_canon["QSTATE"], "QSTATE", checked_params))
            else:
                parts.append(
                    "Limits rows to the queue or processing state values you declare, so monitoring can target only selected outcome bands."
                )
            parts.append("")
            parts.append("**QSTATE Options:**")
            parts.append("- Use standard SAP status values configured for the monitored object type.")
            parts.append("- Code in this EI applies QSTATE as a selector but does not enumerate fixed literals inline.")
            parts.append("")
            continue
        if fld == "STATUS":
            if param_canon.get("STATUS"):
                parts.append(_04_wrap_dictionary_explanation(param_canon["STATUS"], "STATUS", checked_params))
            else:
                parts.append(
                    "Restricts the extract to the operational status values you configure for this EI's object type."
                )
            parts.append("")
            parts.append("**STATUS Options:**")
            parts.append("- Use status domain values defined for the underlying SAP object (see data element or domain in the system).")
            parts.append("- Code applies STATUS as a filter; literals are environment-specific.")
            parts.append("")
            continue
        if fld in ("USER_FLD", "USR_FLD"):
            parts.append(_USER_FLD_DRL_FIXED_MARKDOWN)
            parts.append("")
            if param_canon.get(fld):
                parts.append(_04_wrap_dictionary_explanation(param_canon[fld], fld, checked_params))
                parts.append("")
            lit = _user_fld_literal_values_from_code(code_text)
            if lit:
                parts.append("Explicit values from the supplied ABAP:")
                parts.append("")
                for t in lit:
                    parts.append(f"- **{t}** — Referenced in the ABAP on a line involving USER_FLD.")
                parts.append("")
            continue
        if fld == "MANAGE_IN_UTC":
            if param_canon.get("MANAGE_IN_UTC"):
                parts.append(
                    _04_wrap_dictionary_explanation(
                        param_canon["MANAGE_IN_UTC"], "MANAGE_IN_UTC", checked_params
                    )
                )
            else:
                parts.append(
                    "Controls whether reference timestamps for filtering and duration checks are interpreted in UTC or local time."
                )
            parts.append("")
            parts.append("**MANAGE_IN_UTC Options:**")
            parts.append("- **X** — UTC mode for the relevant timestamp comparisons.")
            parts.append("- Empty or blank — local time / framework default for the application server clock context.")
            parts.append("")
            continue
        if fld in _04_STATUS_SELECTOR_PARAMS:
            if param_canon.get(fld):
                parts.append(_04_wrap_dictionary_explanation(param_canon[fld], fld, checked_params))
            else:
                parts.append(
                    "Optional send-state selector: when set, the monitor applies this outcome flag together with the other active status dimensions."
                )
            parts.append("")
            parts.append(f"**{fld} Options:**")
            parts.append("- **X** — Restrict the extract to rows where this send or processing state is active for the object.")
            parts.append("- Empty or initial — Do not use this flag as a filter dimension.")
            parts.append("")
            continue
        if fld == "DATE_REF_FLD":
            if param_canon.get("DATE_REF_FLD"):
                parts.append(
                    _04_wrap_dictionary_explanation(
                        param_canon["DATE_REF_FLD"], "DATE_REF_FLD", checked_params
                    )
                )
            else:
                parts.append(
                    "Names the date field used as the reference for lookback and time-window filtering when explicit from/to dates are not set."
                )
            parts.append("")
            parts.append("**DATE_REF_FLD Options:**")
            if date_ref_literals:
                for lit in date_ref_literals:
                    d = param_canon.get(lit.upper()) or desc_by_field_upper.get(
                        lit.upper(), "Code-defined date reference field"
                    )
                    d = str(d).rstrip(".")
                    parts.append(f"- {lit} — {d}.")
            else:
                parts.append("- Use a field name from the EI's source structure that carries a valid calendar date for the monitored object.")
                parts.append("- Values are system-specific; choose the field the ABAP selection uses for the primary date axis.")
            parts.append("")
            continue
        if fld in ("AGGR_LEVEL", "AGGLEVEL"):
            if param_canon.get(fld):
                parts.append(_04_wrap_dictionary_explanation(param_canon[fld], fld, checked_params))
            else:
                parts.append(
                    "Controls aggregation granularity for work process totals versus server-specific breakdown."
                )
            parts.append("")
            parts.append(f"**{fld} Options:**")
            if aggr_level_literals:
                for lit in aggr_level_literals:
                    parts.append(f"- {lit} — {_aggr_level_literal_expl(lit)}")
            else:
                parts.append("- T — Total aggregation across all servers.")
                parts.append("- S — Server-level aggregation.")
            parts.append("")
            continue
        if fld == "TIME_REF_FLD":
            if param_canon.get("TIME_REF_FLD"):
                parts.append(
                    _04_wrap_dictionary_explanation(
                        param_canon["TIME_REF_FLD"], "TIME_REF_FLD", checked_params
                    )
                )
            else:
                parts.append(
                    "Names the time field paired with the date reference when the EI evaluates intraday boundaries."
                )
            parts.append("")
            parts.append("**TIME_REF_FLD Options:**")
            parts.append("- Use a time field from the same structure as DATE_REF_FLD or as defined in the EI code path.")
            parts.append("- Values follow SAP time representation (typically HHMMSS semantics in the domain).")
            parts.append("")
            continue
        canon = param_canon.get(str(fld).strip().upper())
        if canon:
            parts.append(_04_wrap_dictionary_explanation(canon, fld, checked_params))
        else:
            parts.append(_auto_04_opener(i, fld, desc))
        parts.append("")
    return _04_insert_not_in_use_lines("\n".join(parts), unused, ordered_names)


def _markdown_07_structure_and_abap(structure_path: Path, code_path: Path) -> str:
    import openpyxl

    code_txt = code_path.read_text(encoding="utf-8", errors="replace")
    wb = openpyxl.load_workbook(structure_path, read_only=True)
    sw = wb.active
    srows = list(sw.iter_rows(values_only=True))
    wb.close()
    hdr, sdata = None, []
    for j, row in enumerate(srows):
        if row and row[0] and "structure" in str(row[0]).lower() and "name" in str(row[0]).lower():
            hdr = [str(x or "").strip() for x in row]
            sdata = srows[j + 1 :]
            break
    if hdr is None:
        hdr = [str(x or "") for x in srows[0]]
        sdata = srows[1:]
    tbl = ["| " + " | ".join(hdr) + " |", "|" + "|".join(["---"] * len(hdr)) + "|"]
    for row in sdata:
        if not row or all(x is None or str(x).strip() == "" for x in row):
            continue
        cells = [str(x).replace("|", "\\|") if x is not None else "" for x in row[: len(hdr)]]
        tbl.append("| " + " | ".join(cells) + " |")
    struct_md = "\n".join(tbl)
    return (
        "## EI Function Structure\n\n"
        "This table lists all output fields returned by the EI. These fields contain the results of the EI's data retrieval and calculations.\n\n"
        + struct_md
        + "\n\n## ABAP Code\n\n```abap\n"
        + code_txt.rstrip()
        + "\n```\n"
    )


def generate_sections_03_04_07(paths: dict[str, Path]) -> int:
    """
    Build 03_response.md, 04_response.md, 07_response.md under RUN_DIR from Available fields, Structure, and Code paths.
    Returns expanded parameter count (for logging).
    """
    rows = _load_parameter_rows_from_available_fields_xlsx(paths["params"])
    text03 = _markdown_03_from_parameter_rows(rows)
    ordered, _ = _param_names_ordered_from_03_table(text03)
    important_n = len(ordered)
    unused = analyze_unused_params(
        paths["code"], paths["params"], search_dirs=[INPUT_DIR, INPUT_DIR / "old"]
    )
    write_unused_params_file(UNUSED_PARAMS_RUN_FILE, unused)
    text04 = _markdown_04_from_parameter_rows(
        rows, important_n, set(ordered), paths["code"], paths["params"]
    )
    text07 = _markdown_07_structure_and_abap(paths["structure"], paths["code"])
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    (RUN_DIR / "03_response.md").write_text(text03, encoding="utf-8")
    (RUN_DIR / "04_response.md").write_text(text04, encoding="utf-8")
    (RUN_DIR / "07_response.md").write_text(text07, encoding="utf-8")
    return important_n


def prepare(
    skip_verify: bool = False,
    assume_yes: bool = False,
    update_manifest_only: bool = False,
    generate_037: bool = False,
) -> None:
    """Verify (unless --skip-verify), clear run/, discover, read Metadata, write manifest and 7 prompt files. If update_manifest_only, only re-read Metadata and write manifest (do not clear response files)."""
    if update_manifest_only:
        paths = _discover_inputs(assume_yes=assume_yes)
        if paths is None:
            print("Could not discover input files.", file=sys.stderr)
            sys.exit(1)
        _write_manifest(paths)
        print("Manifest updated (title from Metadata). Run assemble to rebuild output.")
        return
    errs = verify(skip_verify=skip_verify, assume_yes=assume_yes)
    if errs:
        print("Verification failed:")
        for e in errs:
            print("  -", e)
        print("Fix the issues above or re-run with --skip-verify to ignore.")
        sys.exit(1)

    paths = _discover_inputs(assume_yes=assume_yes)
    if paths is None:
        print("Could not discover input files. Need one of each in input/: Code[_ ]*.txt, Structure[_ ]*.xlsx, Available fields[_ ]*.xlsx, Metadata[_ ]*.xlsx (or user declined to proceed).", file=sys.stderr)
        sys.exit(1)

    # Clear previous run: delete old response files (skip when only updating manifest)
    RUN_DIR.mkdir(parents=True, exist_ok=True)
    if not update_manifest_only:
        for num, _, _ in SECTION_SPEC:
            r = RUN_DIR / f"{num}_response.md"
            if r.exists():
                r.unlink()

    # Read Metadata (same sheet as verify: General / Metadata general / first sheet)
    import openpyxl
    wb = openpyxl.load_workbook(paths["metadata"], read_only=True)
    ws = _metadata_sheet(wb)
    rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
    wb.close()
    id_val = str(rows[7][1] or "").strip() if len(rows) > 8 and len(rows[7]) > 1 else ""
    name_val = str(rows[8][1] or "").strip() if len(rows) >= 9 and len(rows[8]) > 1 else ""
    stem = _stem_from_path(paths["metadata"], PREFIXES["metadata"])
    basename = stem
    title = _manifest_document_title(name_val=name_val, id_val=id_val, basename=basename)

    _write_manifest_at(basename, title)

    # Replacements for placeholder lines in prompts (escape backslashes for re.sub on Windows)
    def _repl(s):
        return str(s).replace("\\", "\\\\")

    def replace_placeholders(text: str) -> str:
        text = re.sub(r"\[Provide the structure file path[^\]]*\]", _repl(paths["structure"]), text)
        text = re.sub(r"\[Provide the output structure / fields file path[^\]]*\]", _repl(paths["structure"]), text)
        text = re.sub(r"\[Provide the Parameters sheet path[^\]]*\]", _repl(paths["params"]), text)
        text = re.sub(r"\[Provide the file path or paste the parameters table content here\]", _repl(paths["params"]), text)
        text = re.sub(r"\[Provide the code file path[^\]]*\]", _repl(paths["code"]), text)
        text = re.sub(r"\[Provide the ABAP source\]", _repl(paths["code"]), text)
        if paths["params_docx"]:
            text = re.sub(
                r"Selected parameters file[^\n]*\n[^\n]*",
                "Selected parameters file: " + _repl(paths["params_docx"]) + "\n",
                text,
                count=1,
            )
        else:
            text = re.sub(
                r"Selected parameters file[^\n]*\n[^\n]*",
                "Selected parameters file: (optional – omit if not in input)\n",
                text,
                count=1,
            )
        return text

    unused = analyze_unused_params(
        paths["code"], paths["params"], search_dirs=[INPUT_DIR, INPUT_DIR / "old"]
    )
    write_unused_params_file(UNUSED_PARAMS_RUN_FILE, unused)
    unused_prompt_block = format_unused_params_prompt_block(unused)
    if unused:
        print("Unused parameters (excluded from sections 05/06):", ", ".join(sorted(unused)))

    for num, template_name, _ in SECTION_SPEC:
        template_path = PROMPTS_DIR / template_name
        if not template_path.exists():
            print(f"Missing prompt template: {template_path}", file=sys.stderr)
            sys.exit(1)
        text = template_path.read_text(encoding="utf-8")
        text = replace_placeholders(text)
        if UNUSED_PARAMS_PROMPT_PLACEHOLDER in text:
            if num in ("05", "06"):
                text = text.replace(UNUSED_PARAMS_PROMPT_PLACEHOLDER, unused_prompt_block)
            else:
                text = text.replace(UNUSED_PARAMS_PROMPT_PLACEHOLDER, "")
        (RUN_DIR / f"{num}_prompt.txt").write_text(text, encoding="utf-8")

    if generate_037:
        n = generate_sections_03_04_07(paths)
        print(f"Generated 03_response.md, 04_response.md, 07_response.md from input ({n} parameters).")

    print("Prepare done. Output basename:", basename)
    if generate_037:
        print("Sections 03, 04, and 07 were auto-generated; add or refresh 01, 02, 05, and 06 in scripts/pipeline/run/ as needed.")
    print("In Cursor, send the instruction from scripts/pipeline/CURSOR_INSTRUCTION.txt")
    print("When the 7 response files are in scripts/pipeline/run/, run: python scripts/pipeline/pipeline.py verify  (optional)")
    print("Then run: python scripts/pipeline/pipeline.py assemble")


def assemble() -> None:
    """Read manifest and 7 response files, build one .md, write to output/, convert to .docx."""
    if not (RUN_DIR / "manifest.txt").exists():
        print("Run prepare first. No run/manifest.txt found.", file=sys.stderr)
        sys.exit(1)
    manifest = _strip_bom_and_zwsp((RUN_DIR / "manifest.txt").read_text(encoding="utf-8"))
    basename = ""
    title = ""
    for line in manifest.splitlines():
        if line.startswith("output_basename="):
            basename = line.split("=", 1)[1].strip()
        elif line.startswith("title="):
            title = line.split("=", 1)[1].strip()
    if not basename:
        print("manifest.txt missing output_basename=", file=sys.stderr)
        sys.exit(1)

    # Verify response files (04 rules: no "output only", parameter count match) before assembling
    verr = verify_responses()
    if verr:
        print("Response verification failed:")
        for e in verr:
            print("  -", e)
        print("Fix the issues above (e.g. edit 04_response.md) then run assemble again.")
        sys.exit(1)

    parts = []
    inserted_parameters_header = False
    for num, _, _ in SECTION_SPEC:
        if num == "03" and not inserted_parameters_header:
            # Parent section for 03-06 subsections in the final document.
            parts.append("## Parameters")
            inserted_parameters_header = True
        r = RUN_DIR / f"{num}_response.md"
        chunk = r.read_text(encoding="utf-8")
        chunk = _normalize_response_markdown(chunk)
        chunk = _dedupe_duplicate_md_section_headings(chunk, num)
        parts.append(chunk)

    full_md = f"# {title}\n\n" + "\n\n".join(parts)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_md = OUTPUT_DIR / f"Explanation_{basename}.md"
    out_md.write_text(full_md, encoding="utf-8")
    print("Wrote", out_md)

    out_docx = OUTPUT_DIR / f"Explanation_{basename}.docx"
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "md_to_docx",
            PIPELINE_DIR / "md_to_docx.py",
        )
        md2docx = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(md2docx)
        md2docx.convert_md_to_docx(out_md, out_docx)
        print("Wrote", out_docx)
    except ImportError as e:
        print("MD->DOCX converter unavailable (install markdown, beautifulsoup4). Skipping .docx.", e)
    except Exception as e:
        print("MD->DOCX conversion failed. Markdown is ready.", e)


def main() -> None:
    parser = argparse.ArgumentParser(description="EI Doc Pipeline: prepare | verify | assemble")
    parser.add_argument("mode", choices=["prepare", "verify", "assemble"], help="prepare (verify + write prompts), verify (check response files), or assemble (build .md + .docx)")
    parser.add_argument("--skip-verify", action="store_true", help="skip verification when running prepare")
    parser.add_argument("--yes", "-y", action="store_true", help="assume yes for 'proceed anyway?' prompts (e.g. similar file names)")
    parser.add_argument("--update-manifest", action="store_true", help="prepare: only re-read Metadata and write manifest (do not clear response files)")
    parser.add_argument(
        "--generate-037",
        action="store_true",
        help="prepare: after writing prompts, build 03/04/07 from Available fields, Structure, and Code (no separate script)",
    )
    args = parser.parse_args()
    if args.mode == "prepare":
        prepare(
            skip_verify=args.skip_verify,
            assume_yes=args.yes,
            update_manifest_only=args.update_manifest,
            generate_037=args.generate_037,
        )
    elif args.mode == "verify":
        verr = verify_responses()
        if verr:
            print("Response verification failed:")
            for e in verr:
                print("  -", e)
            sys.exit(1)
        print("Response verification passed.")
    else:
        assemble()


if __name__ == "__main__":
    main()
