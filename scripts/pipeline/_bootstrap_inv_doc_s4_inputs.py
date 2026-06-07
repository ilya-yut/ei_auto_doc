"""Bootstrap input files for SW_10_02_INV_DOC_S4 (same function/structure as INVENT_CNT)."""
import shutil
from pathlib import Path

import openpyxl

in_dir = Path(__file__).resolve().parents[2] / "input"
old = in_dir / "old"
stem = "INVENT_CNT"
struct_name = "/SKN/S_SW_10_02_INVENT_CNT"

CODE_PARAMS = [
    "AGG_LVL", "BACKDAYS", "BLDAT", "BUDAT", "BUKRS", "BWKEY", "COMP_OPERATOR", "DATUM",
    "DATE_REF_FLD", "DIFF_AMOUNT", "DURATION", "DURATION_UNIT", "GIDAT", "KTOPL", "LANGU",
    "LGORT", "LSTAT", "MANAGE_IN_UTC", "PRESENT_ZERO", "REF_FIELD1", "REF_FIELD2",
    "REF_TABNAME1", "REF_TABNAME2", "RESULT_COMP", "SOBKZ", "SPERR", "DSTAT", "SW_DEST",
    "USNAM", "USNAM_HD", "VGART", "WAERS", "WAERS_FR", "WERKS", "XBUFI", "ZLDAT", "ZSTAT",
]

def _find(name_glob: str):
    for base in (in_dir, old):
        hits = sorted(base.glob(name_glob))
        if hits:
            return hits[0]
    return None

af = _find("Available*INV_DOC*")
code_src_early = _find("Code_*INV_DOC*")

for pat in ["*INV_DOC_S4*", "Metadata _SKN_S_SW_10_02_INVENT_CNT*"]:
    for p in list(in_dir.glob(pat)):
        if p.is_file():
            dest = old / p.name
            if dest.exists():
                dest.unlink()
            shutil.move(str(p), str(dest))

if af is None:
    raise SystemExit("Available fields file not found")

wb_af = openpyxl.load_workbook(af, read_only=True)
ws_af = wb_af["Available Fields"]
fields = []
for r in ws_af.iter_rows(min_row=3, values_only=True):
    if r and r[0] and str(r[0]).strip() != "Field":
        fields.append(list(r)[:7])
wb_af.close()

sp = in_dir / f"Structure_SKN_S_SW_10_02_{stem}.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Structure"
ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
for f in fields:
    fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
    dtype = f"{typ}({ln})" if typ and ln else (typ or "")
    ws.append([struct_name, fname, desc or "", dtype, de or dom or ""])
wb.save(sp)

mp = in_dir / f"Metadata _SKN_S_SW_10_02_{stem}.xlsx"
src_meta = old / "Metadata _SKN_S_SW_10_02_INVENT_CNT.xlsx"
if not src_meta.exists():
  src_meta = next(old.glob("Metadata*INVENT_CNT*"), None)
shutil.copy(src_meta, mp)

code_src = code_src_early
if code_src is None:
    raise SystemExit("Code file not found")
dst = in_dir / f"Code_SKN_S_SW_10_02_{stem}.txt"
dst.write_text(code_src.read_text(encoding="utf-8"), encoding="utf-8")
if code_src.parent == in_dir:
    code_src.unlink()

ap2 = in_dir / f"Available fields_SKN_S_SW_10_02_{stem}.xlsx"
shutil.copy(af, ap2)
if af.parent == in_dir and af != ap2:
    af.unlink()

wb = openpyxl.load_workbook(ap2)
ws = wb["Parameters"]
oldp = {}
name_map = {
    "REF_FIELD_NAME1": "REF_FIELD1",
    "REF_FIELD_NAME2": "REF_FIELD2",
}
for r in ws.iter_rows(min_row=3, values_only=True):
    if r and r[0]:
        key = str(r[0]).strip().upper()
        oldp[name_map.get(key, key)] = list(r)
if ws.max_row >= 3:
    ws.delete_rows(3, ws.max_row - 2)
ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
extras = {
    "BACKDAYS": ["Backdays", "INT4", "10", "0", "BACKDAYS", "BACKDAYS"],
    "SW_DEST": ["RFC Destination", "", "0", "0", "", ""],
    "DATE_REF_FLD": ["Date Reference Field", "CHAR", "30", "0", "NAME_FELD", "NAME_FELD"],
    "MANAGE_IN_UTC": ["'X' - Manage in UTC", "", "0", "0", "", ""],
    "AGG_LVL": ["Aggregation level", "CHAR", "10", "0", "", ""],
    "DIFF_AMOUNT": ["Difference amount threshold", "INT4", "10", "0", "", ""],
    "PRESENT_ZERO": ["Present zero differences", "CHAR", "1", "0", "", "XFELD"],
    "REF_TABNAME1": ["Reference table 1", "CHAR", "30", "0", "TABNAME", "AS4TAB"],
    "REF_TABNAME2": ["Reference table 2", "CHAR", "30", "0", "TABNAME", "AS4TAB"],
    "REF_FIELD1": ["Reference field 1", "CHAR", "30", "0", "NAME_FELD", "NAME_FELD"],
    "REF_FIELD2": ["Reference field 2", "CHAR", "30", "0", "NAME_FELD", "NAME_FELD"],
    "COMP_OPERATOR": ["Comparison operator", "CHAR", "2", "0", "BUCC_OPERATOR", "BUCC_OPERATOR"],
    "WAERS_FR": ["Currency from", "CUKY", "5", "0", "WAERS", "WAERS"],
    "DATUM": ["Reference Date", "DATS", "8", "0", "DATUM", "DATUM"],
    "USNAM_HD": ["User name header", "CHAR", "12", "0", "USNAM", "USNAM"],
}
for i, fld in enumerate(CODE_PARAMS, start=3):
    row = oldp.get(fld.upper()) or extras.get(fld)
    ws.cell(i, 1, fld)
    if row:
        for c in range(2, 8):
            if len(row) > c - 1 and row[c - 1] not in (None, ""):
                ws.cell(i, c, row[c - 1])
wb.save(ap2)
print(f"ready {len(CODE_PARAMS)} params {len(fields)} fields")
