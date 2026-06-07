"""Bootstrap input files for SW_10_07_BANK_VEND_DEF."""
import shutil
from pathlib import Path

import openpyxl

in_dir = Path(__file__).resolve().parents[2] / "input"
old = in_dir / "old"
stem = "BANK_VEND_DEF"
struct_name = "/SKN/S_SW_10_07_ONE_TIME_VEND"

CODE_PARAMS = [
    "BACKDAYS", "BANKL", "BANKN", "BANKS", "BUKRS", "CONVERT_KEY", "COUNTER", "DATUM",
    "DATE_REF_FLD", "DURATION", "DURATION_UNIT", "EMPFG", "FORWDAYS", "KTOKK", "LAND1",
    "LAUFD", "LAUFI", "LANGU", "LIFNR", "LNRZA", "LOEVM", "STKZN", "SW_DEST", "VALID_FROM",
    "VALID_TO", "VBUND", "XVORL", "XCPDK", "ZBNKL", "ZBNKN", "ZBNKS",
]

for pat in ["*VEND_ACC_MISS*", "Metadata _SKN_S_SW_10_06_VEND_ACC_MISS*"]:
    for p in in_dir.glob(pat):
        if p.is_file():
            dest = old / p.name
            if dest.exists():
                dest.unlink()
            shutil.move(str(p), str(dest))

af = next(in_dir.glob("Available*BANK*"))
wb_af = openpyxl.load_workbook(af, read_only=True)
ws_af = wb_af["Available Fields"]
fields = []
for r in ws_af.iter_rows(min_row=3, values_only=True):
    if r and r[0] and str(r[0]).strip() != "Field":
        fields.append(list(r)[:7])
wb_af.close()

sp = in_dir / f"Structure_SKN_S_SW_10_07_{stem}.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Structure"
ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
for f in fields:
    fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
    dtype = f"{typ}({ln})" if typ and ln else (typ or "")
    ws.append([struct_name, fname, desc or "", dtype, de or dom or ""])
wb.save(sp)

wrong = in_dir / "Structure_SKN_S_SW_10_07_ONE_TIME_VEND.xlsx"
if wrong.exists():
    wrong.unlink()

mp = in_dir / f"Metadata _SKN_S_SW_10_07_{stem}.xlsx"
shutil.copy(old / "Metadata _SKN_S_SW_10_06_PF_VENDOR.xlsx", mp)
wb = openpyxl.load_workbook(mp)
ws = wb.active
ws.cell(8, 2, "SW_10_07_BANK_VEND_D")
ws.cell(9, 2, "Bank in payment advise missing in Vendor's master")
wb.save(mp)

src = next(in_dir.glob("Code_Bank*"))
dst = in_dir / f"Code_SKN_S_SW_10_07_{stem}.txt"
dst.write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
src.unlink()

ap2 = in_dir / f"Available fields_SKN_S_SW_10_07_{stem}.xlsx"
shutil.copy(af, ap2)
af.unlink()
wb = openpyxl.load_workbook(ap2)
ws = wb["Parameters"]
oldp = {}
for r in ws.iter_rows(min_row=3, values_only=True):
    if r and r[0]:
        oldp[str(r[0]).strip().upper()] = list(r)
if ws.max_row >= 3:
    ws.delete_rows(3, ws.max_row - 2)
ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
extras = {
    "BACKDAYS": ["Backdays", "INT4", "10", "0", "BACKDAYS", "BACKDAYS"],
    "FORWDAYS": ["Forward Days", "INT4", "10", "0", "FORWDAYS", "FORWDAYS"],
    "SW_DEST": ["RFC Destination", "", "0", "0", "", ""],
    "DATE_REF_FLD": ["Date Reference Field", "CHAR", "30", "0", "NAME_FELD", "NAME_FELD"],
    "DURATION_UNIT": ["Duration Unit", "CHAR", "1", "0", "/SKN/E_SW_DURATION_UNIT", "/SKN/D_SW_DURATION_UNIT"],
    "CONVERT_KEY": ["'X' - Decompose Key Field", "", "0", "0", "", ""],
    "XCPDK": ["One-time account", "CHAR", "1", "0", "XCPDK", "XFELD"],
    "LOEVM": ["Deletion flag", "CHAR", "1", "0", "LOEVM", "XFELD"],
    "LANGU": ["Language for texts", "", "0", "0", "", ""],
    "DATUM": ["Reference Date", "DATS", "8", "0", "DATUM", "DATUM"],
    "COUNTER": ["Counter", "", "", "", "", ""],
    "VALID_FROM": ["Valid from", "DATS", "8", "0", "KOVON", "DATUM"],
    "VALID_TO": ["Valid To", "DATS", "8", "0", "KOBIS", "DATUM"],
    "XVORL": ["Indicator for proposal run", "CHAR", "1", "0", "XVORL", "XVORL"],
    "ZBNKS": ["Bank country", "CHAR", "3", "0", "DZBNKS", "LAND1"],
    "ZBNKN": ["Bank account", "CHAR", "18", "0", "DZBNKN", "BANKN"],
    "ZBNKL": ["Bank key", "CHAR", "15", "0", "DZBNKL", "BANKK"],
}
for i, fld in enumerate(CODE_PARAMS, start=3):
    row = oldp.get(fld.upper()) or extras.get(fld)
    ws.cell(i, 1, fld)
    if row:
        for c in range(2, 8):
            if len(row) > c - 1 and row[c - 1] not in (None, ""):
                ws.cell(i, c, row[c - 1])
wb.save(ap2)
print(f"ready {len(CODE_PARAMS)} params {len(fields)} output fields")
