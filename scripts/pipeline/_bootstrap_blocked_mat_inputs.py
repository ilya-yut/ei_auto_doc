"""Bootstrap input files for SW_10_02_BLOCKED_MAT."""
import shutil
from pathlib import Path

import openpyxl

in_dir = Path(__file__).resolve().parents[2] / "input"
old = in_dir / "old"
stem = "BLOCKED_MAT"
struct_name = "/SKN/S_SW_10_02_BLOCKED_MAT"

CODE_PARAMS = [
    "AENAM",
    "BACKDAYS",
    "DATUM",
    "LAEDA",
    "LANGU",
    "MATNR",
    "MSTAV",
    "NO_DATE_RESTRICTION",
    "SW_DEST",
    "VKORG",
    "VMSTA",
    "VTWEG",
]

for pat in ["*BANK_VEND_DEF*", "Metadata _SKN_S_SW_10_07_BANK_VEND_DEF*"]:
    for p in in_dir.glob(pat):
        if p.is_file():
            dest = old / p.name
            if dest.exists():
                dest.unlink()
            shutil.move(str(p), str(dest))

af = next(in_dir.glob("Available*BLOCKED*"))
wb_af = openpyxl.load_workbook(af, read_only=True)
ws_af = wb_af["Available Fields"]
fields = []
for r in ws_af.iter_rows(min_row=3, values_only=True):
    if r and r[0] and str(r[0]).strip() != "Field":
        fields.append(list(r)[:7])
wb_af.close()

sp = in_dir / f"Structure_SKN_S_SW_10_02_{stem}.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Structure"
ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
for f in fields:
    fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
    dtype = f"{typ}({ln})" if typ and ln else (typ or "")
    ws.append([struct_name, fname, desc or "", dtype, de or dom or ""])
wb.save(sp)

mp = in_dir / f"Metadata _SKN_S_SW_10_02_{stem}.xlsx"
shutil.copy(old / "Metadata _SKN_S_SW_10_06_PF_VENDOR.xlsx", mp)
wb = openpyxl.load_workbook(mp)
ws = wb.active
ws.cell(8, 2, "SW_10_02_BLOCKED_MAT")
ws.cell(9, 2, "MM: Blocked materials")
wb.save(mp)

src = next(in_dir.glob("Code_MM*BLOCKED*"))
dst = in_dir / f"Code_SKN_S_SW_10_02_{stem}.txt"
dst.write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
src.unlink()

ap2 = in_dir / f"Available fields_SKN_S_SW_10_02_{stem}.xlsx"
shutil.copy(af, ap2)
af.unlink()
wb = openpyxl.load_workbook(ap2)
ws = wb["Parameters"]
oldp = {}
for r in ws.iter_rows(min_row=3, values_only=True):
    if r and r[0]:
        oldp[str(r[0]).strip().upper()] = list(r)
if ws.max_row >= 3:
    ws.delete_rows(3, ws.max_row - 2)
ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
extras = {
    "BACKDAYS": ["Backdays", "INT4", "10", "0", "BACKDAYS", "BACKDAYS"],
    "NO_DATE_RESTRICTION": ["No date restriction", "CHAR", "1", "0", "", "XFELD"],
    "SW_DEST": ["RFC Destination", "", "0", "0", "", ""],
    "DATUM": ["Reference Date", "DATS", "8", "0", "DATUM", "DATUM"],
    "LANGU": ["Language for texts", "", "0", "0", "", ""],
}
for i, fld in enumerate(CODE_PARAMS, start=3):
    row = oldp.get(fld.upper()) or extras.get(fld)
    ws.cell(i, 1, fld)
    if row:
        for c in range(2, 8):
            if len(row) > c - 1 and row[c - 1] not in (None, ""):
                ws.cell(i, c, row[c - 1])
wb.save(ap2)
print(f"ready {len(CODE_PARAMS)} params {len(fields)} fields")
