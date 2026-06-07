"""Bootstrap input files for SW_10_02_PO_AGR_VAL (PO_AGREE_VALID)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
OLD = INPUT / "old"
STEM = "SKN_S_SW_10_02_PO_AGREE_VALID"
STRUCT_NAME = "/SKN/S_SW_10_02_PO_AGREE_VALID"
INDICATOR_ID = "SW_10_02_PO_AGR_VAL"
INDICATOR_NAME = "PO based on agreement with long validity"

CODE_PARAMS = [
    "BACKDAYS",
    "BEDAT",
    "DATUM",
    "EBELN",
    "PERIOD",
    "SW_DEST",
]

SRC_CODE_GLOB = "Code_*PO_AGR_VAL*"
SRC_AVAIL_GLOB = "Available*PO_AGR_VAL*"


def _find(glob_pat: str) -> Path | None:
    for base in (INPUT, OLD):
        hits = sorted(base.glob(glob_pat))
        if hits:
            return hits[0]
    return None


def main() -> None:
    OLD.mkdir(exist_ok=True)

    for pat in ["*MAT_DESC_PLANT*", "Metadata _SKN_S_SW_10_02_MAT_DESC_PLANT*"]:
        for p in list(INPUT.glob(pat)):
            if p.is_file():
                dest = OLD / p.name
                if dest.exists():
                    dest.unlink()
                shutil.move(str(p), str(dest))

    src_code = _find(SRC_CODE_GLOB)
    src_avail = _find(SRC_AVAIL_GLOB)
    if not src_code or not src_avail:
        raise SystemExit("Missing code or available-fields source for PO_AGR_VAL")

    code_canon = INPUT / f"Code_{STEM}.txt"
    avail_canon = INPUT / f"Available fields_{STEM}.xlsx"
    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    meta_path = INPUT / f"Metadata _{STEM}.xlsx"

    code_canon.write_text(src_code.read_text(encoding="utf-8"), encoding="utf-8")
    if src_code.parent == INPUT and src_code != code_canon:
        src_code.unlink()

    shutil.copy2(src_avail, avail_canon)
    if src_avail.parent == INPUT and src_avail.resolve() != avail_canon.resolve():
        src_avail.unlink()

    wb_af = openpyxl.load_workbook(avail_canon, read_only=True)
    ws_af = wb_af["Available Fields"]
    fields = []
    for r in ws_af.iter_rows(min_row=3, values_only=True):
        if r and r[0] and str(r[0]).strip() != "Field":
            fields.append(list(r)[:7])
    wb_af.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Structure"
    ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
    for f in fields:
        fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
        dtype = f"{typ}({ln})" if typ and ln else (typ or "")
        ws.append([STRUCT_NAME, fname, desc or "", dtype, de or dom or ""])
    wb.save(struct_path)
    wb.close()

    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "General"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    wb = openpyxl.load_workbook(avail_canon)
    ws = wb["Parameters"]
    oldp = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r and r[0]:
            oldp[str(r[0]).strip().upper()] = list(r)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
    extras = {
        "SW_DEST": ["RFC Destination", "", "0", "0", "", ""],
        "BACKDAYS": ["Backdays", "INT4", "10", "0", "BACKDAYS", "BACKDAYS"],
        "PERIOD": ["Maximum days between PO date and change date", "INT4", "10", "0", "", ""],
        "DATUM": ["Reference Date", "DATS", "8", "0", "DATUM", "DATUM"],
    }
    for i, fld in enumerate(CODE_PARAMS, start=3):
        row = oldp.get(fld.upper()) or extras.get(fld)
        ws.cell(i, 1, fld)
        if row:
            for c in range(2, 8):
                if len(row) > c - 1 and row[c - 1] not in (None, ""):
                    ws.cell(i, c, row[c - 1])
    wb.save(avail_canon)
    wb.close()

    print(f"ready {len(CODE_PARAMS)} params {len(fields)} fields")


if __name__ == "__main__":
    main()
