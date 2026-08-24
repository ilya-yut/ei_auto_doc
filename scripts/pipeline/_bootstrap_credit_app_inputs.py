"""Bootstrap input files for SW_10_01_CREDIT_APP (Credit Management Approvals / MD_CHNG_LOG)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
OLD = INPUT / "old"
STEM = "SKN_S_SW_10_06_MD_CHNG_LOG"
STRUCT_NAME = "/SKN/S_SW_10_06_MD_CHNG_LOG"
INDICATOR_ID = "SW_10_01_CREDIT_APP"
INDICATOR_NAME = "Credit Management Approvals"

FRAMEWORK_PARAMS = [
    "LANGU",
    "BACKDAYS",
    "MANAGE_IN_UTC",
    "SW_DEST",
]

# Selectable / framework params from ABAP DATA_SINGLE / DATA_MULTY plus sheet params.
CODE_PARAMS = [
    "ACT_CHNGNO",
    "BACKDAYS",
    "CHANGENR",
    "CHANGE_IND",
    "CHANGE_IND_DESC",
    "CHNGIND",
    "CHNGIND_DESC",
    "CONVERT_KEY",
    "CUKY_NEW",
    "CUKY_OLD",
    "DATUM",
    "DURATION_D",
    "FIELD_DESC",
    "FNAME",
    "FNAME_REP",
    "HEADER_ONLY",
    "KEY1",
    "KEY1_DS",
    "KEY1_V",
    "KEY2",
    "KEY2_DS",
    "KEY2_V",
    "KEY3",
    "KEY3_DS",
    "KEY3_V",
    "KEY4",
    "KEY4_DS",
    "KEY4_V",
    "KEY5",
    "KEY5_DS",
    "KEY5_V",
    "KEY6",
    "KEY6_DS",
    "KEY6_V",
    "KEY7",
    "KEY7_DS",
    "KEY7_V",
    "KEY8",
    "KEY8_DS",
    "KEY8_V",
    "KEY9",
    "KEY9_DS",
    "KEY9_V",
    "KEY10",
    "KEY10_DS",
    "KEY10_V",
    "LANGU",
    "MANAGE_IN_UTC",
    "NAME_FIRST",
    "NAME_LAST",
    "NAME_TEXT",
    "OBJECTCLAS",
    "OBJECTID",
    "OBJECT_DESC",
    "PLANCHNGNR",
    "REPET_BACKDAYS",
    "REPETITIVE",
    "SW_DEST",
    "TABKEY",
    "TABNAME",
    "TAB_DESC",
    "TCODE",
    "TEXT_CASE",
    "UDATE",
    "UDATE_REPET",
    "UNIT_NEW",
    "UNIT_OLD",
    "USERNAME",
    "UTIME",
    "VALUE_NEW",
    "VALUE_OLD",
    "WAS_PLANND",
]


def _find(*globs: str) -> Path | None:
    for base in (INPUT, OLD):
        for g in globs:
            hits = sorted(base.glob(g))
            if hits:
                return hits[0]
    return None


def main() -> None:
    OLD.mkdir(exist_ok=True)

    for pat in [
        "*OUTDT*",
        "*Outdated*",
        "Code_Credit*",
        "Available fields_Credit*",
        "Metadata *MD_CHNG*",
        "Metadata _SKN_S_SW_10_06_MD_CHNG*",
    ]:
        for p in list(INPUT.glob(pat)):
            if p.is_file() and p.parent == INPUT:
                dest = OLD / p.name
                if dest.exists():
                    dest.unlink()
                shutil.move(str(p), str(dest))

    src_code = _find(
        "Code_*CREDIT_APP*",
        "Code_*Credit Management Approvals*",
        "Code_*MD_CHNG_LOG*",
    )
    src_avail = _find(
        "Available*CREDIT_APP*",
        "Available*Credit Management Approvals*",
        "Available*MD_CHNG_LOG*",
    )
    if not src_code or not src_avail:
        raise SystemExit("missing code or available-fields source for CREDIT_APP")

    code_canon = INPUT / f"Code_{STEM}.txt"
    avail_canon = INPUT / f"Available fields_{STEM}.xlsx"
    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    meta_path = INPUT / f"Metadata _{STEM}.xlsx"

    code_canon.write_text(src_code.read_text(encoding="utf-8"), encoding="utf-8")
    if src_code.parent == INPUT and src_code.resolve() != code_canon.resolve():
        # already moved to old above for Credit* names; leave MD_CHNG if different
        pass

    shutil.copy2(src_avail, avail_canon)

    wb_af = openpyxl.load_workbook(avail_canon, read_only=True)
    ws_af = wb_af["Available Fields"]
    fields = []
    for r in ws_af.iter_rows(min_row=3, values_only=True):
        if r and r[0] and str(r[0]).strip() != "Field":
            fields.append(list(r)[:7])
    wb_af.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Structure"
    ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
    for f in fields:
        fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
        dtype = f"{typ}({ln})" if typ and ln else (typ or "")
        ws.append([STRUCT_NAME, fname, desc or "", dtype, de or dom or ""])
    wb.save(struct_path)
    wb.close()

    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "General"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    wb = openpyxl.load_workbook(avail_canon)
    ws = wb["Parameters"]
    oldp: dict[str, list] = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or not r[0]:
            continue
        key = str(r[0]).strip().upper()
        oldp[key] = list(r)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
    extras = {
        "LANGU": ["Language for texts", "LANG", "1", "0", "LANGU", "SPRAS"],
        "DATUM": ["Date", "DATS", "8", "0", "DATUM", "DATUM"],
        "SW_DEST": ["RFC destination", "CHAR", "32", "0", "RFCDEST", "RFCDEST"],
        "DURATION_D": ["Duration in days", "INT4", "10", "0", "/SKN/E_SW_DURATION_D", ""],
        "REPET_BACKDAYS": ["Repetition lookback days", "INT4", "10", "0", "", ""],
        "UDATE_REPET": ["Change date for repetition scan", "DATS", "8", "0", "CDDATUM", "DATUM"],
        "HEADER_ONLY": ["Header changes only", "CHAR", "1", "0", "", "XFELD"],
        "CONVERT_KEY": ["'X' - Decompose Key Field", "CHAR", "1", "0", "", "XFELD"],
        "FNAME_REP": ["Field name for repetition", "CHAR", "30", "0", "FIELDNAME", "FDNAME"],
        "MANAGE_IN_UTC": ["'X' - Manage in UTC", "CHAR", "1", "0", "", "XFELD"],
        "BACKDAYS": ["Backdays", "INT4", "10", "0", "", ""],
    }
    for i, fld in enumerate(CODE_PARAMS, start=3):
        row = oldp.get(fld.upper()) or extras.get(fld)
        ws.cell(i, 1, fld)
        if row:
            for c in range(2, 8):
                if len(row) > c - 1 and row[c - 1] not in (None, ""):
                    ws.cell(i, c, row[c - 1])
    wb.save(avail_canon)
    wb.close()

    print(f"ready {len(CODE_PARAMS)} params, {len(fields)} fields, metadata={meta_path.name}")


if __name__ == "__main__":
    main()
