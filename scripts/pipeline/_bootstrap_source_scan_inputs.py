"""
Bootstrap Structure / Available fields / Metadata for SKN_S_SW_01_01_SOURCE_SCAN
from Code_Scan ABAP Report Sources_SW_01_01_SOURCE_SCAN.txt.
Run from repo root: python scripts/pipeline/_bootstrap_source_scan_inputs.py
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
STEM = "SKN_S_SW_01_01_SOURCE_SCAN"
STRUCT_NAME = "/SKN/S_SW_01_01_SOURCE_SCAN"
INDICATOR_ID = "SW_01_01_SOURCE_SCAN"
INDICATOR_NAME = "Scan ABAP Report Sources"
SRC = INPUT / "Code_Scan ABAP Report Sources_SW_01_01_SOURCE_SCAN.txt"
CODE_CANON = INPUT / f"Code_{STEM}.txt"


def _add_multiline_block(lines: list[str], start_idx: int, names: set[str]) -> int:
    ln = lines[start_idx]
    rest = re.sub(r"^\s*\w+_MULTY:\s*", "", ln)
    mm0 = re.match(r"(\w+)", rest.strip())
    if mm0:
        names.add(mm0.group(1))
    j = start_idx + 1
    while j < len(lines):
        s = lines[j]
        if re.match(r"^\s*DATA:\s", s) or re.match(r"^\s*(DATA_|SELECT_|LV_)", s):
            break
        mm = re.match(r"^\s+(\w+)\s*[,]", s) or re.match(r"^\s+(\w+)\s+\S", s)
        if mm:
            names.add(mm.group(1))
        j += 1
    return j


def _parse_params_from_code(text: str) -> list[str]:
    names: set[str] = set()
    m = re.search(r"DATA_SINGLE:\s*([\s\S]*?)^\s*DATA_MULTY:", text, re.MULTILINE)
    if m:
        for ln in m.group(1).splitlines():
            t = ln.strip()
            if not t or t.startswith('"'):
                continue
            mm = re.match(r"^,?\s*(\w+)\s+", t)
            if mm:
                names.add(mm.group(1))
    for m in re.finditer(r"^\s*DATA_SINGLE:\s*(\w+)\s+\S", text, re.MULTILINE):
        names.add(m.group(1))
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        if re.match(r"^\s*DATA_MULTY:", lines[i]):
            i = _add_multiline_block(lines, i, names)
            continue
        if re.match(r"^\s*SELECT_MULTY:", lines[i]):
            i = _add_multiline_block(lines, i, names)
            continue
        i += 1
    i = 0
    while i < len(lines):
        ln = lines[i]
        if not re.match(r"^\s*SELECT_SINGLE:\s*", ln):
            i += 1
            continue
        tail = re.sub(r"^\s*SELECT_SINGLE:\s*", "", ln).strip()
        if tail:
            tok = tail.split(",")[0].strip()
            if tok.isidentifier():
                names.add(tok)
        i += 1
        while i < len(lines):
            s = lines[i]
            st = s.strip()
            if (
                not st
                or st.startswith('"---')
                or re.match(r"^\s*DATA_SINGLE:\s*SW_DEST", s)
                or re.match(r"^\s*DATA:\s", s)
            ):
                break
            if len(s) - len(s.lstrip()) < 10:
                break
            mm = re.match(r"^\s*(\w+)\s*,", s) or re.match(r"^\s*(\w+)\s*\.\s*$", s)
            if mm and mm.group(1).upper() not in ("ENDIF", "ENDLOOP", "ENDCASE", "ENDFUNCTION"):
                names.add(mm.group(1))
            if st.endswith("."):
                i += 1
                break
            i += 1
        continue
    if re.search(r"LV_BACKDAYS|^\s*BACKDAYS\b", text, re.MULTILINE):
        names.add("BACKDAYS")
    return sorted(names, key=str.upper)


def _guess_type(field: str) -> tuple[str, str, str, str]:
    u = field.upper()
    if u in ("BACKDAYS", "DURATION"):
        return "INT4", "10", "0", u
    if u in ("DURATION_UNIT", "LANGU"):
        return "CHAR", "1", "0", u
    if u == "SW_DEST":
        return "CHAR", "32", "0", "RFCDEST"
    if u in ("CREATEDON", "CDAT", "UDAT", "DATUM"):
        return "DATS", "8", "0", u
    if u in ("TRKORR",):
        return "CHAR", "20", "0", "TRKORR"
    if u in ("TRSTATUS", "TRFUNCTION"):
        return "CHAR", "1", "0", u
    if u in ("PGMID",):
        return "CHAR", "4", "0", "PGMID"
    if u in ("OBJTYPE",):
        return "CHAR", "4", "0", "TROBJTYPE"
    if u in ("OBJNAME", "INCLUDE"):
        return "CHAR", "40", "0", u
    if u == "STRING_SEARCH":
        return "CHAR", "255", "0", "/SKN/E_SW_SOURCE_SCAN_STRING"
    return "CHAR", "50", "0", u


def main() -> None:
    dest = CODE_CANON
    if SRC.exists():
        text = SRC.read_text(encoding="utf-8", errors="replace")
        if dest.resolve() != SRC.resolve():
            shutil.move(str(SRC), str(dest))
            text = dest.read_text(encoding="utf-8", errors="replace")
    elif dest.exists():
        text = dest.read_text(encoding="utf-8", errors="replace")
    else:
        raise SystemExit(f"Missing {SRC} and {dest}")

    params = _parse_params_from_code(text)

    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Structure"
    ws_s.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
    for fld in params:
        if fld == "SW_DEST":
            continue
        t, ln, dec, de = _guess_type(fld)
        ws_s.append([STRUCT_NAME, fld, fld, f"{t}({ln})", de])
    wb_s.save(struct_path)
    wb_s.close()

    avail_path = INPUT / f"Available fields_{STEM}.xlsx"
    wb_a = openpyxl.Workbook()
    ws_p = wb_a.create_sheet("Parameters", 0)
    ws_p.append(["Field", "Description", "Type", "Length", "Decimal", "Data Element", "Domain"])
    ws_p.append(["", "", "", "", "", "", ""])
    for fld in params:
        t, ln, dec, de = _guess_type(fld)
        ws_p.append([fld, fld.replace("_", " ").title(), t, ln, dec, de, de])
    wb_a.save(avail_path)
    wb_a.close()

    meta_path = INPUT / f"Metadata _{STEM}.xlsx"
    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "Metadata general"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    dup = INPUT / "Available fields_Scan ABAP Report Sources_SW_01_01_SOURCE_SCAN.xlsx"
    old_dir = INPUT / "old"
    old_dir.mkdir(exist_ok=True)
    if dup.exists():
        shutil.move(str(dup), str(old_dir / dup.name))

    print("Wrote:", struct_path.name, avail_path.name, meta_path.name)
    print("Code:", dest.name)
    print("Parameters count:", len(params))


if __name__ == "__main__":
    main()
