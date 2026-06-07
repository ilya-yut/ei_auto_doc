"""Detect EI parameters declared in the interface but not used in code or on the output sheet."""
from __future__ import annotations

import re
from pathlib import Path

NOT_IN_USE_LINE = "**Not in use**"

_RE_CALL_FM = re.compile(r"CALL\s+FUNCTION\s+'([^']+)'", re.IGNORECASE)


def _add_multiline_block(lines: list[str], start_idx: int, names: set[str]) -> int:
    ln = lines[start_idx]
    rest = re.sub(r"^\s*\w+_MULTY:\s*", "", ln)
    mm0 = re.match(r"(\w+)", rest.strip())
    if mm0:
        names.add(mm0.group(1).upper())
    j = start_idx + 1
    while j < len(lines):
        s = lines[j]
        if re.match(r"^\s*DATA:\s", s) or re.match(r"^\s*(DATA_|SELECT_|LV_)", s):
            break
        mm = re.match(r"^\s+(\w+)\s*[,]", s) or re.match(r"^\s+(\w+)\s+\S", s)
        if mm:
            names.add(mm.group(1).upper())
        j += 1
    return j


def parse_interface_params(text: str) -> list[str]:
    """Parameter names read via DATA_SINGLE/MULTY and SELECT_SINGLE/MULTY."""
    names: set[str] = set()
    m = re.search(r"DATA_SINGLE:\s*([\s\S]*?)^\s*DATA_MULTY:", text, re.MULTILINE)
    if m:
        for ln in m.group(1).splitlines():
            t = ln.strip()
            if not t or t.startswith('"'):
                continue
            mm = re.match(r"^,?\s*(\w+)\s+", t)
            if mm:
                names.add(mm.group(1).upper())
    for m in re.finditer(r"^\s*DATA_SINGLE:\s*(\w+)\s+\S", text, re.MULTILINE):
        names.add(m.group(1).upper())
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        if re.match(r"^\s*DATA_MULTY:", lines[i]) or re.match(r"^\s*SELECT_MULTY:", lines[i]):
            i = _add_multiline_block(lines, i, names)
            continue
        i += 1
    i = 0
    while i < len(lines):
        ln = lines[i]
        if not re.match(r"^\s*SELECT_SINGLE:\s*", ln):
            i += 1
            continue
        tail = re.sub(r"^\s*SELECT_SINGLE:\s*", "", ln).strip()
        if tail:
            tok = tail.split(",")[0].strip()
            if tok.isidentifier():
                names.add(tok.upper())
        i += 1
        while i < len(lines):
            s = lines[i]
            st = s.strip()
            if (
                not st
                or st.startswith('"---')
                or re.match(r"^\s*DATA_SINGLE:\s*SW_DEST", s)
                or re.match(r"^\s*DATA:\s", s)
            ):
                break
            if len(s) - len(s.lstrip()) < 10:
                break
            mm = re.match(r"^\s*(\w+)\s*,", s) or re.match(r"^\s*(\w+)\s*\.\s*$", s)
            if mm:
                names.add(mm.group(1).upper())
            i += 1
    if re.search(r"\bSW_DEST\b", text, re.IGNORECASE):
        names.add("SW_DEST")
    return sorted(names)


def params_from_parameters_sheet(params_xlsx: Path) -> list[str]:
    import openpyxl

    wb = openpyxl.load_workbook(params_xlsx, read_only=True)
    if "Parameters" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Parameters"]
    out: list[str] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or not r[0]:
            break
        f = str(r[0]).strip().upper()
        if f and f != "FIELD":
            out.append(f)
    wb.close()
    return out


def load_output_field_names(params_xlsx: Path) -> set[str]:
    import openpyxl

    wb = openpyxl.load_workbook(params_xlsx, read_only=True)
    ws = None
    for name in wb.sheetnames:
        if "available" in name.lower() and "field" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[1] if len(wb.sheetnames) > 1 else 0]
    out: set[str] = set()
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r and r[0] and str(r[0]).strip().upper() != "FIELD":
            out.add(str(r[0]).strip().upper())
    wb.close()
    return out


def _datum_param_used(code: str) -> bool:
    cu = code.upper()
    sm = re.search(
        r"SELECT_MULTY:\s*([\s\S]*?)(?:\n\s*(?:SELECT_SINGLE|DATA:|CONVERT_|RANGES|IF |CHECK ))",
        cu,
    )
    if not sm or "DATUM" not in sm.group(1):
        return False
    if re.search(r"IN\s+R_DATUM|READ\s+TABLE\s+R_DATUM|R_DATUM\[\]\s+IS\s+NOT\s+INITIAL", cu):
        return True
    if re.search(r"CASE\s+LV_DATE_REF_FLD", cu) and re.search(r"R_DATUM", cu):
        return True
    return False


def code_uses_param(code: str, param: str) -> bool:
    cu = code.upper()
    p = param.upper()

    if re.search(rf"T_DATA-{p}\s*=|<FS_DATA>-{p}\s*=", cu):
        return True
    if re.search(rf"DELETE\s+T_DATA\s+WHERE\s+{p}\b", cu):
        return True
    if re.search(rf"WHEN\s+'{p}'", cu):
        return True
    if re.search(rf"{p}\s+NOT\s+IN\s+R_{p}", cu):
        return True

    if p == "DATUM":
        return _datum_param_used(cu)

    if re.search(rf"\bLV_{p}\b", cu):
        if re.search(
            rf"IF\s+LV_{p}|LV_{p}\s+IS\s+NOT\s+INITIAL|LV_{p}\s*[=<>]|"
            rf"[-+]\s*LV_{p}\b|\bLV_{p}\s*[-+]|>\s*LV_{p}\b|<\s*LV_{p}\b",
            cu,
        ):
            return True

    if re.search(rf"\bR_{p}\b", cu):
        if re.search(
            rf"IN\s+R_{p}|IF\s+R_{p}|APPEND\s+RS_{p}\s+TO\s+R_{p}|"
            rf"DELETE\s+T_DATA\s+WHERE\s+{p}\b|~\s*{p}\s+IN\s+R_{p}|"
            rf"\b{p}\s+IN\s+R_{p}",
            cu,
        ):
            return True
        if re.search(rf"\bR_{p}\b\s*(\[\]|\(|=)", cu):
            return True

    return False


def _function_name_from_code(text: str) -> str | None:
    m = re.search(r"FUNCTION\s+(\S+)", text, re.IGNORECASE)
    if not m:
        return None
    return m.group(1).upper().rstrip(".")


def _find_code_files_for_function(fm_name: str, search_dirs: list[Path]) -> list[Path]:
    hits: list[Path] = []
    pat = re.compile(rf"FUNCTION\s+{re.escape(fm_name)}\b", re.IGNORECASE)
    for d in search_dirs:
        if not d.exists():
            continue
        for p in list(d.glob("Code_*.txt")) + list(d.glob("Code _*.txt")):
            try:
                t = p.read_text(encoding="utf-8", errors="replace")[:8000]
            except OSError:
                continue
            if pat.search(t):
                hits.append(p)
    return hits


def _collect_code_texts(main_code: str, search_dirs: list[Path]) -> list[str]:
    texts = [main_code]
    if not re.search(r"CALL\s+FUNCTION", main_code, re.IGNORECASE):
        return texts
    passes_t_select = "T_SELECT" in main_code.upper()
    seen_fm: set[str] = set()
    for m in _RE_CALL_FM.finditer(main_code):
        fm = m.group(1).upper().rstrip(".")
        if fm in seen_fm:
            continue
        seen_fm.add(fm)
        if not passes_t_select:
            continue
        for path in _find_code_files_for_function(fm, search_dirs):
            try:
                callee = path.read_text(encoding="utf-8", errors="replace")
            except OSError:
                continue
            if callee not in texts:
                texts.append(callee)
    return texts


def analyze_unused_params(
    code_path: Path,
    params_xlsx: Path,
    *,
    search_dirs: list[Path] | None = None,
) -> set[str]:
    """
    Return parameter names (uppercase) that are in the interface but neither on the
    output sheet nor actively used in ABAP (main + T_SELECT callees when found).
    """
    code = code_path.read_text(encoding="utf-8", errors="replace")
    output_fields = load_output_field_names(params_xlsx)
    sheet_params = params_from_parameters_sheet(params_xlsx)
    candidates = sheet_params if sheet_params else parse_interface_params(code)
    dirs = search_dirs or [code_path.parent, code_path.parent / "old"]
    if code_path.parent.name == "old":
        dirs = [code_path.parent.parent, code_path.parent]
    code_texts = _collect_code_texts(code, dirs)
    unused: set[str] = set()
    for p in candidates:
        pu = p.upper()
        if pu in output_fields:
            continue
        if any(code_uses_param(t, p) for t in code_texts):
            continue
        unused.add(pu)
    return unused


def write_unused_params_file(path: Path, unused: set[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = sorted(unused)
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def read_unused_params_file(path: Path) -> set[str]:
    if not path.exists():
        return set()
    out: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if s and not s.startswith("#"):
            out.add(s.upper())
    return out


def format_unused_params_prompt_block(unused: set[str]) -> str:
    if not unused:
        return (
            "**Unused parameters (pipeline analysis):** None. "
            "All interface parameters are used in code or appear on the output structure sheet."
        )
    names = ", ".join(sorted(unused))
    return (
        "**Unused parameters (pipeline analysis — exclude from this section):** "
        f"{names}. "
        "These parameters are documented in section 04 with **Not in use** and must not appear here."
    )
