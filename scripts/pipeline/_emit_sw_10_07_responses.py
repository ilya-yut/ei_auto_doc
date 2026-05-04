# One-off generator for pipeline run SKN_S_SW_10_07_FI_DOC_POSTED (01, 02, 04, 05, 06).
# Run from repo root: python scripts/pipeline/_emit_sw_10_07_responses.py

from __future__ import annotations

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
RUN = ROOT / "scripts" / "pipeline" / "run"

BACKDAYS_MANDATORY = (
    "BACKDAYS defines the historical monitoring window by specifying how many days backward from "
    "today to retrieve records. 0 - today, 1 - today + yesterday etc."
)
BACKDAYS_ANCHOR = "Backdays is based on DATE_REF_FLD field."


def _load_dict() -> dict[str, str]:
    """Match scripts/pipeline/pipeline.py _load_params_dictionary_explanations (col A + col B + repairs)."""
    wb = openpyxl.load_workbook(ROOT / "input" / "params_dictionary.xlsx", read_only=True, data_only=True)
    sn = "dictionary" if "dictionary" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out: dict[str, str] = {}
    start = 0
    if rows[0] and str(rows[0][0] or "").strip().lower() in ("parameter", "field", "param"):
        start = 1
    for row in rows[start:]:
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip().upper()
        if not key:
            continue
        raw = row[1] if len(row) > 1 else ""
        text = str(raw or "").strip()
        if not text:
            continue
        text = text.replace("\ufffd", "'").replace("\u2019", "'").replace("\u2013", "-").replace("\u2014", "-")
        out[key] = text
    return out


def _load_checked() -> set[str]:
    p = ROOT / "input" / "checked params.txt"
    if not p.exists():
        return set()
    s = set()
    for ln in p.read_text(encoding="utf-8").splitlines():
        t = ln.split("#", 1)[0].strip()
        if t:
            s.add(t.upper())
    return s


def _mark(u: str, text: str, checked: set[str], dict_key: str | None = None) -> str:
    key = (dict_key or u).upper()
    if not text:
        return ""
    if key in checked:
        return text
    return f"<mark>{text}</mark>"


def _main_from_dict(u: str, d: dict[str, str], checked: set[str]) -> str:
    body = d.get(u, "").strip()
    if not body:
        return ""
    return _mark(u, body, checked)


def write_01() -> None:
    text = """## General Overview

This Exception Indicator surfaces accounting documents whose posting timing sits in an earlier fiscal period than the business activity dates you treat as current, so finance controllers can see postings that may belong to a closed or prior reporting window. It joins document header and line perspectives, enriches amounts and master descriptions, and applies an optional age test so exception queues stay focused on material items.

This EI serves as an essential control for financial close and operational integrity by:
- Highlighting postings that can distort period comparability when activity dates and fiscal posting periods diverge
- Giving accounts payable and receivable teams a consolidated view of who posted what, in which company code, and with which document types when timing exceptions appear
- Supporting GL and subledger reconciliation by carrying line-level direction, amounts, and account context alongside header identifiers
- Enabling targeted follow-up on clusters tied to specific users, transaction codes, or reference numbers without manual table extracts
- Providing evidence-friendly output for internal control testing around retroactive or late-period postings

Organizations use this style of monitoring during month-end and year-end close, after reopening periods, and when investigating suspected backdating or cut-off errors. Results are intended to feed exception workflows before final sign-off on financial statements.

The routine reads data from standard FI document header and line sources (including secondary index paths when the line table is stored as a cluster structure) together with company-code directory attributes used for currency and chart-of-accounts context.
"""
    (RUN / "01_response.md").write_text(text, encoding="utf-8")


def write_02() -> None:
    text = """## Problem Description

Failure to monitor postings that land in prior fiscal periods while business dates suggest current-period activity creates multiple risks across financial reporting, operational control, and audit readiness.

**Financial Reporting and Close Risks**
- Period profit and balance sheet balances can shift without transparent explanation when late or reopened postings are not reviewed in time
- Management reports that rely on posting period slices may misstate trends if timing exceptions accumulate unidentified
- Statutory and management reporting deadlines compress remediation time once exceptions are discovered only during external review
- Cross-company views become inconsistent when some entities correct cut-off issues while others remain unaware of similar patterns

**Operational and Master Data Risks**
- Accounts teams may approve accruals or reversals while unaware that underlying documents still carry prior-period posting dates
- Document type, user, or transaction code concentrations can signal process breakdowns yet stay hidden without automated surfacing
- Line-level debit and credit imbalances or unusual posting keys may indicate training gaps or system integration defects tied to the same timing issue
- Vendor or customer subledger mismatches can linger when clearing documents post outside the expected fiscal window

**Management Visibility and Accountability Risks**
- Executives lose confidence in flash close metrics when unexplained prior-period postings appear late in the cycle
- Internal audit cannot efficiently sample risky populations without a repeatable exception list tied to fiscal period logic
- Escalations between shared service centers and local entities slow when nobody owns a consolidated view of timing outliers

## Suggested Resolution

**Immediate Response**
- Review each surfaced document for company code, fiscal year, document number, and posting date versus the business dates shown in the exception list
- Validate whether the posting was an authorized reopening, a legitimate correction, or an unintended booking using standard FI display transactions your organization permits
- Confirm user and transaction code context with the preparer before reversing or adjusting anything in production
- Capture business commentary where the posting was intentional so close committees can document exceptions

**System Assessment**
- Compare current results with the prior monitoring cycle after period status changes, transports, or automated posting jobs
- Examine concentrations by document type, user, or reference number to see if a single process drives most findings
- Revisit the configured fiscal-period boundary logic relative to your organization’s official close calendar when false positives cluster at month boundaries
- Check whether optional age filters are excluding immaterial noise or, conversely, hiding items that still breach policy thresholds

**Corrective Actions**
- Post corrective or reversal documents through your standard FI change process, with approvals where policy requires them
- Update training, desktop procedures, or scheduling for recurring jobs when root cause is procedural rather than data defect
- Tighten or relax monitoring parameters after root-cause review so the queue remains actionable for controllers and shared services
- Route repeat systemic issues into defect or change management when configuration or integration changes are required
- Retain monitoring extracts and resolution notes when regulators or auditors expect evidence of supervisory review
"""
    (RUN / "02_response.md").write_text(text, encoding="utf-8")


def write_04(d: dict[str, str], checked: set[str]) -> None:
    lines: list[str] = []
    lines.append("### Parameter Configuration Guidelines")
    lines.append("")
    lines.append(
        "IMPORTANT: Configure ALL 49 parameters listed in the Parameters Reference Table when tuning this EI; "
        "each influences which records are read, filtered, aged, and surfaced for alerting."
    )
    lines.append("")

    def blk(heading: str, desc: str, paras: list[str]) -> None:
        lines.append(f"**{heading}** ({desc})")
        lines.append("")
        for p in paras:
            if p:
                lines.append(p)
                lines.append("")
        lines.append("")

    # AEDAT
    blk("AEDAT", "Aedat", [_main_from_dict("AEDAT", d, checked)])

    # BACKDAYS
    back_paras = [BACKDAYS_MANDATORY, BACKDAYS_ANCHOR]
    c_b = d.get("BACKDAYS", "")
    if c_b and c_b.replace("\u2013", "-").strip() != BACKDAYS_MANDATORY:
        back_paras.append(_mark("BACKDAYS", c_b, checked))
    blk("BACKDAYS", "Backdays", back_paras)

    blk("BELNR", "Belnr", [_main_from_dict("BELNR", d, checked)])
    blk("BKTXT", "Bktxt", [_main_from_dict("BKTXT", d, checked)])
    blk("BLART", "Blart", [_main_from_dict("BLART", d, checked)])
    blk("BLDAT", "Bldat", [_main_from_dict("BLDAT", d, checked)])

    blk(
        "BSCHL",
        "Bschl",
        [
            "Posting key on the accounting line that controls how amounts post to debits or credits, tax handling, and special posting situations."
        ],
    )

    blk("BSTAT", "Bstat", [_main_from_dict("BSTAT", d, checked)])

    blk("BUDAT", "Budat", [_main_from_dict("BUDAT", d, checked)])
    blk("BUKRS", "Bukrs", [_main_from_dict("BUKRS", d, checked)])
    blk("BUZEI", "Buzei", [_main_from_dict("BUZEI", d, checked)])
    blk("CPUDT", "Cpudt", [_main_from_dict("CPUDT", d, checked)])

    # DATE_REF_FLD
    lines.append("**DATE_REF_FLD** (Date Ref Fld)")
    lines.append("")
    lines.append(_main_from_dict("DATE_REF_FLD", d, checked))
    lines.append("")
    lines.append("**DATE_REF_FLD Options:**")
    lines.append("- CPUDT — System entry date of the document header used when you want monitoring windows aligned to capture time.")
    lines.append("- BLDAT — Document date carried on the header for legal or external correspondence timing.")
    lines.append("- AEDAT — Last-changed date on the header when maintenance-driven windows matter more than creation.")
    lines.append("- UPDDT — Last update date on the header when you need windows keyed to the latest modification cycle.")
    lines.append("")

    blk(
        "DATUM",
        "Datum",
        [
            "Explicit calendar bounds for the monitoring pass; when populated, these ranges override the relative lookback built from BACKDAYS."
        ],
    )

    blk(
        "DMBE2 - DMBE3",
        "Dmbe2",
        [
            "Additional local-currency amount fields on the line used for parallel valuation views; set ranges when you need to narrow lines that carry non-zero values in those valuation buckets."
        ],
    )

    blk("DMBTR", "Dmbtr", [_main_from_dict("DMBTR", d, checked)])

    blk("DURATION", "Duration", [_main_from_dict("DURATION", d, checked)])

    # DURATION_UNIT
    lines.append("**DURATION_UNIT** (Duration Unit)")
    lines.append("")
    lines.append(_main_from_dict("DURATION_UNIT", d, checked))
    lines.append("")
    lines.append("**DURATION_UNIT Options:**")
    lines.append("- H — Hours.")
    lines.append("- M — Minutes.")
    lines.append("- D — Days.")
    lines.append("- F — Full-day counting for day-based age thresholds.")
    lines.append("")

    # FORWDAYS
    ftxt = d.get("FORWDAYS", "")
    fparas = []
    if ftxt:
        fparas.append(_mark("FORWDAYS", ftxt, checked))
    fparas.append(
        "When supplied together with BACKDAYS, extends the upper calendar bound forward from the evaluation day while still "
        "anchoring the lower bound from the backward interval; when BACKDAYS is initial and this value is set, the selection "
        "starts forward from the evaluation day instead."
    )
    blk("FORWDAYS", "Forwdays", fparas)

    blk(
        "GJAHR",
        "Gjahr",
        ["Fiscal year of the accounting document used to pair header and line rows and to scope year-specific reporting."],
    )

    blk("GRPID", "Grpid", [_main_from_dict("GRPID", d, checked)])

    blk(
        "GVTYP",
        "Gvtyp",
        ["Transaction type on the line that classifies how the line participates in consolidation or tax reporting when you filter special categories."],
    )

    blk("HKONT", "Hkont", [_main_from_dict("HKONT", d, checked)])

    blk(
        "HWAE2 - HWAE3",
        "Hwae2",
        [
            "Secondary and tertiary currency keys on the document header used when parallel currency translations are stored; restrict them when monitoring focuses on specific reporting currencies."
        ],
    )

    blk("HWAER", "Hwaer", [_main_from_dict("HWAER", d, checked)])

    blk(
        "KOART",
        "Koart",
        [
            "Account-type selector for cluster-based environments that tells the join whether customer, vendor, or general-ledger secondary index paths should supply line facts."
        ],
    )

    blk("KTOPL", "Ktopl", [_main_from_dict("KTOPL", d, checked)])

    blk(
        "KURS2 - KURS3",
        "Kurs2",
        [
            "Secondary and tertiary exchange rates on the header used with the parallel currency fields; narrow them when rate-driven false positives must be suppressed."
        ],
    )

    blk("KURSF", "Kursf", [_main_from_dict("KURSF", d, checked)])

    blk(
        "KZBTR",
        "Kzbtr",
        ["Quantity in the posting unit of measure on the line for operational postings that carry physical quantities alongside monetary amounts."],
    )

    blk("LANGU", "Langu", [_main_from_dict("LANGU", d, checked)])

    blk("MONAT", "Monat", [_main_from_dict("MONAT", d, checked)])

    blk(
        "PERIOD_CLOSING_DAY",
        "Period Closing Day",
        [
            "Calendar day within a month that defines how fiscal periods are split for the document-posting-period helper before header and line selection runs."
        ],
    )

    blk("SGTXT", "Sgtxt", [_main_from_dict("SGTXT", d, checked)])

    lines.append("**SHKZG** (Shkzg)")
    lines.append("")
    lines.append(_main_from_dict("SHKZG", d, checked))
    lines.append("")
    lines.append("**SHKZG Options:**")
    lines.append("- S — Line posts on the debit side of the account.")
    lines.append("- H — Line posts on the credit side of the account.")
    lines.append("")

    blk(
        "STBLG",
        "Stblg",
        ["Number of the reversal or referenced document on the header when you need to correlate cancelled postings with their follow-on documents."],
    )

    blk("SW_DEST", "Sw Dest", [_main_from_dict("SW_DEST", d, checked)])

    blk("TCODE", "Tcode", [_main_from_dict("TCODE", d, checked)])

    lines.append("**TIME_REF_FLD** (Time Ref Fld)")
    lines.append("")
    lines.append(
        "Identifies which time-of-day field should accompany the chosen document date attribute when the runtime measures elapsed age for each line."
    )
    lines.append("")
    if d.get("TIME_REF_FLD"):
        lines.append(_mark("TIME_REF_FLD", d["TIME_REF_FLD"], checked))
        lines.append("")
    lines.append("**TIME_REF_FLD Options:**")
    lines.append("- Use a time field that exists on the same structure as the document date reference you configured.")
    lines.append("- Values follow the SAP time representation used in your system for that field.")
    lines.append("")

    blk("UPDDT", "Upddt", [_main_from_dict("UPDDT", d, checked)])
    blk("USNAM", "Usnam", [_main_from_dict("USNAM", d, checked)])
    blk("WAERS", "Waers", [_main_from_dict("WAERS", d, checked)])

    blk(
        "WAERS_T001",
        "Waers T001",
        [
            "Company-code local currency from the financial directory row joined to the document; use it to align document currency rows with the official company-code currency."
        ],
    )

    blk("WRBTR", "Wrbtr", [_main_from_dict("WRBTR", d, checked)])
    blk("XBLNR", "Xblnr", [_main_from_dict("XBLNR", d, checked)])

    blk(
        "XREVERSAL",
        "Xreversal",
        [
            "Header-level reversal indicator used to include or exclude documents that represent reversal traffic in the exception population."
        ],
    )

    (RUN / "04_response.md").write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def write_05() -> None:
    text = """### Parameter Relationships

How parameter combinations work together

**Explicit calendar window versus relative lookback:** **DATUM** supplies explicit from-and-to calendar bounds for the monitoring pass. When **DATUM** is empty, **BACKDAYS** (and optionally **FORWDAYS**) builds the calendar window relative to the evaluation day before documents are read.

**Reference date axis:** **DATE_REF_FLD** chooses which header date attribute is mapped into that calendar window for each generated period slice, so the same BACKDAYS span can follow creation, document, change, or update dates depending on configuration.

**Age filter after dates:** **DURATION** with **DURATION_UNIT** is an additional filter applied after date-oriented selection: each candidate line keeps its place in the result only when the computed age from the reference date and clock fields still fits the configured duration band.

**Fiscal period boundary:** **PERIOD_CLOSING_DAY** works with the generated date and posting-date tables to shape how fiscal periods are derived for the selection pass, which indirectly constrains which header lines qualify before line facts are merged.

**Remote execution path:** **SW_DEST** must be populated so the remote join runs in the monitored system; other organizational filters such as **BUKRS**, **HKONT**, or **KOART** only affect which documents are returned once connectivity is established.

**Final selection:** Both the date window logic (explicit **DATUM** or **BACKDAYS**/**FORWDAYS**) and the **DURATION**/**DURATION_UNIT** age test must be satisfied before a row is treated as part of the final exception population for alerting.
"""
    (RUN / "05_response.md").write_text(text, encoding="utf-8")


def write_06() -> None:
    text = """### Default Values

- **PERIOD_CLOSING_DAY** - 15
- **BACKDAYS** - 10
- **DATE_REF_FLD** - CPUDT
- **DURATION_UNIT** - D
- **LANGU** - EN
- **DURATION** - initial - treated as empty range keeps rows by code

### Practical Example of Parameter Configuration

**Use Case 1: Company-wide prior-period posting scan**

**Purpose:** Keep month-end focused on all company codes while using the default creation-date reference and day-based aging.
```
BUKRS = 1000 - 1999
BACKDAYS = 14
DATE_REF_FLD = CPUDT
DURATION = 5 - 999999
DURATION_UNIT = D
```

**Use Case 2: Full-day age filter for high-risk accounts**

**Purpose:** Highlight only lines that are at least thirty full days old after the date window is applied.
```
HKONT = 200000 - 299999
BACKDAYS = 30
DURATION = 30
DURATION_UNIT = F
PERIOD_CLOSING_DAY = 25
```

**Use Case 3: Explicit close-week window**

**Purpose:** Anchor the run to a known reopening week instead of relative lookback alone.
```
DATUM = 20250325 - 20250331
BUKRS = 1000
BLART = SA - ZP
DURATION_UNIT = H
DURATION = 0 - 48
```

**Use Case 4: Vendor subledger slice with document-type control**

**Purpose:** Narrow to vendor account-type cluster paths while still applying language and posting-date filters.
```
KOART = K
BUDAT = 20250101 - 20250131
LANGU = EN
TCODE = FB60
```

**Use Case 5: Material document references and user accountability**

**Purpose:** Tie exceptions to external reference numbers and preparers for targeted follow-up.
```
XBLNR = INV2025*
USNAM = BATCH01 - BATCH99
CPUDT = 20250401 - 20250415
WRBTR = 10000 - 999999999
SW_DEST = PROD_FIN
```
"""
    (RUN / "06_response.md").write_text(text, encoding="utf-8")


def main() -> None:
    d = _load_dict()
    checked = _load_checked()
    write_01()
    write_02()
    write_04(d, checked)
    write_05()
    write_06()
    print("Wrote 01, 02, 04, 05, 06 to", RUN)


if __name__ == "__main__":
    main()
