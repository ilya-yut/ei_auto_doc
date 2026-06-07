"""Bootstrap input files for SW_10_02_INVENT_CNT."""
import shutil
from pathlib import Path

import openpyxl

in_dir = Path(__file__).resolve().parents[2] / "input"
old = in_dir / "old"
stem = "INVENT_CNT"
struct_name = "/SKN/S_SW_10_02_INVENT_CNT"

CODE_PARAMS = [
    "AGG_LVL", "BACKDAYS", "BLDAT", "BUDAT", "BUKRS", "BWKEY", "COMP_OPERATOR", "DATUM",
    "DATE_REF_FLD", "DIFF_AMOUNT", "DURATION", "DURATION_UNIT", "GIDAT", "KTOPL", "LANGU",
    "LGORT", "LSTAT", "MANAGE_IN_UTC", "PRESENT_ZERO", "REF_FIELD1", "REF_FIELD2",
    "REF_TABNAME1", "REF_TABNAME2", "RESULT_COMP", "SOBKZ", "SPERR", "DSTAT", "SW_DEST",
    "USNAM", "USNAM_HD", "VGART", "WAERS", "WAERS_FR", "WERKS", "XBUFI", "ZLDAT", "ZSTAT",
]

for pat in ["*BLOCKED_MAT*", "Metadata _SKN_S_SW_10_02_BLOCKED_MAT*"]:
    for p in in_dir.glob(pat):
        if p.is_file():
            dest = old / p.name
            if dest.exists():
                dest.unlink()
            shutil.move(str(p), str(dest))

af = next(in_dir.glob("Available*INV_CNT*"))
wb_af = openpyxl.load_workbook(af, read_only=True)
ws_af = wb_af["Available Fields"]
fields = []
for r in ws_af.iter_rows(min_row=3, values_only=True):
    if r and r[0] and str(r[0]).strip() != "Field":
        fields.append(list(r)[:7])
wb_af.close()

sp = in_dir / f"Structure_SKN_S_SW_10_02_{stem}.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Structure"
ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
for f in fields:
    fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
    dtype = f"{typ}({ln})" if typ and ln else (typ or "")
    ws.append([struct_name, fname, desc or "", dtype, de or dom or ""])
wb.save(sp)

wrong = in_dir / "Structure_SKN_S_SW_10_02_INVENT_CNT.xlsx"
# same name as sp - already overwritten

mp = in_dir / f"Metadata _SKN_S_SW_10_02_{stem}.xlsx"
shutil.copy(old / "Metadata _SKN_S_SW_10_06_PF_VENDOR.xlsx", mp)
wb = openpyxl.load_workbook(mp)
ws = wb.active
ws.cell(8, 2, "SW_10_02_INV_CNT_DOC")
ws.cell(9, 2, "Inventory count - Inventory Document level (IBLNR)")
wb.save(mp)

src = next(in_dir.glob("Code_Inventory*"))
dst = in_dir / f"Code_SKN_S_SW_10_02_{stem}.txt"
dst.write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
src.unlink()

ap2 = in_dir / f"Available fields_SKN_S_SW_10_02_{stem}.xlsx"
shutil.copy(af, ap2)
af.unlink()
wb = openpyxl.load_workbook(ap2)
ws = wb["Parameters"]
oldp = {}
for r in ws.iter_rows(min_row=3, values_only=True):
    if r and r[0]:
        oldp[str(r[0]).strip().upper()] = list(r)
if ws.max_row >= 3:
    ws.delete_rows(3, ws.max_row - 2)
ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
extras = {
    "BACKDAYS": ["Backdays", "INT4", "10", "0", "BACKDAYS", "BACKDAYS"],
    "SW_DEST": ["RFC Destination", "", "0", "0", "", ""],
    "DATE_REF_FLD": ["Date Reference Field", "CHAR", "30", "0", "NAME_FELD", "NAME_FELD"],
    "MANAGE_IN_UTC": ["'X' - Manage in UTC", "", "0", "0", "", ""],
    "AGG_LVL": ["Aggregation level", "CHAR", "10", "0", "", ""],
    "DIFF_AMOUNT": ["Difference amount threshold", "INT4", "10", "0", "", ""],
    "PRESENT_ZERO": ["Present zero differences", "CHAR", "1", "0", "", "XFELD"],
    "REF_TABNAME1": ["Reference table 1", "CHAR", "30", "0", "TABNAME", "AS4TAB"],
    "REF_TABNAME2": ["Reference table 2", "CHAR", "30", "0", "TABNAME", "AS4TAB"],
    "REF_FIELD1": ["Reference field 1", "CHAR", "30", "0", "NAME_FELD", "NAME_FELD"],
    "REF_FIELD2": ["Reference field 2", "CHAR", "30", "0", "NAME_FELD", "NAME_FELD"],
    "COMP_OPERATOR": ["Comparison operator", "CHAR", "2", "0", "BUCC_OPERATOR", "BUCC_OPERATOR"],
    "WAERS_FR": ["Currency from", "CUKY", "5", "0", "WAERS", "WAERS"],
    "DATUM": ["Reference Date", "DATS", "8", "0", "DATUM", "DATUM"],
    "USNAM_HD": ["User name header", "CHAR", "12", "0", "USNAM", "USNAM"],
}
for i, fld in enumerate(CODE_PARAMS, start=3):
    row = oldp.get(fld.upper()) or extras.get(fld)
    ws.cell(i, 1, fld)
    if row:
        for c in range(2, 8):
            if len(row) > c - 1 and row[c - 1] not in (None, ""):
                ws.cell(i, c, row[c - 1])
wb.save(ap2)
print(f"ready {len(CODE_PARAMS)} params {len(fields)} fields")
