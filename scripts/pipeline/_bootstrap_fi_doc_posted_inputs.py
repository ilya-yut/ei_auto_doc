"""
One-off bootstrap: create Structure / Available fields / Metadata for
SKN_S_SW_10_07_FI_DOC_POSTED from the Code_*.txt in input/.
Run from repo root: python scripts/pipeline/_bootstrap_fi_doc_posted_inputs.py
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
STEM = "SKN_S_SW_10_07_FI_DOC_POSTED"
STRUCT_NAME = "/SKN/S_SW_10_07_FI_DOC_POSTED"
INDICATOR_ID = "SW_10_07_FI_DOC_POSTED"
INDICATOR_NAME = "FI documents posted to previous fiscal period"


def _parse_params_from_code(text: str) -> list[str]:
    names: set[str] = set()
    for pat in (
        r"^\s*DATA_MULTY:\s+(\w+)\s+\S",
        r"^\s*DATA_SINGLE:\s+(\w+)\s",
        r"^\s*SELECT_MULTY:\s+(\w+)\.",
        r"^\s*SELECT_SINGLE:\s+(\w+)\.",
    ):
        for m in re.finditer(pat, text, re.MULTILINE):
            names.add(m.group(1))
    # Framework / RFC params still appear in selection
    for extra in ("SW_DEST",):
        if extra in text:
            names.add(extra)
    return sorted(names, key=str.upper)


def _guess_type(field: str) -> tuple[str, str, str, str]:
    """Type, Length, Decimal, Data element placeholder."""
    u = field.upper()
    if u in ("BACKDAYS", "FORWDAYS", "DURATION"):
        return "INT4", "10", "0", u
    if u == "PERIOD_CLOSING_DAY":
        return "NUMC", "2", "0", u
    if u in ("DURATION_UNIT", "LANGU", "SHKZG"):
        return "CHAR", "1", "0", u
    if u in ("DATE_REF_FLD", "TIME_REF_FLD"):
        return "CHAR", "30", "0", "NAME_FELD"
    if u == "SW_DEST":
        return "CHAR", "32", "0", "RFCDEST"
    return "CHAR", "50", "0", u


def main() -> None:
    old_code = INPUT / "Code_FI documents are posted to previous fiscal period_SW_10_07_FI_DOC_POST.txt"
    new_code = INPUT / f"Code_{STEM}.txt"
    if not old_code.exists() and new_code.exists():
        text = new_code.read_text(encoding="utf-8", errors="replace")
    elif old_code.exists():
        text = old_code.read_text(encoding="utf-8", errors="replace")
        shutil.move(str(old_code), str(new_code))
    else:
        raise SystemExit(f"Missing code file: {old_code} or {new_code}")

    params = _parse_params_from_code(text)

    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Structure"
    ws_s.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
    for fld in params:
        if fld == "SW_DEST":
            continue
        t, ln, dec, comp = _guess_type(fld)
        ws_s.append([STRUCT_NAME, fld, fld, f"{t}({ln})", comp])
    wb_s.save(struct_path)
    wb_s.close()

    avail_path = INPUT / f"Available fields_{STEM}.xlsx"
    wb_a = openpyxl.Workbook()
    ws_p = wb_a.create_sheet("Parameters", 0)
    ws_p.append(["Field", "Description", "Type", "Length", "Decimal", "Data Element", "Domain"])
    ws_p.append(["", "", "", "", "", "", ""])
    for fld in params:
        t, ln, dec, de = _guess_type(fld)
        ws_p.append([fld, fld.replace("_", " ").title(), t, ln, dec, de, de])
    std = wb_a["Sheet"]
    wb_a.remove(std)
    wb_a.save(avail_path)
    wb_a.close()

    meta_path = INPUT / f"Metadata _{STEM}.xlsx"
    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "Metadata general"
    for r in range(1, 12):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    print("Wrote:", struct_path.name, avail_path.name, meta_path.name)
    print("Renamed/moved code to:", new_code.name)
    print("Parameters count:", len(params))


if __name__ == "__main__":
    main()
