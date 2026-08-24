"""Bootstrap input files for SW_10_01_ORD_CHNG (Sales Order SLA Agreement Analysis - detailed)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
OLD = INPUT / "old"
STEM = "SKN_S_SW_10_01_ORD_CHNG"
STRUCT_NAME = "/SKN/S_SW_10_01_ORD_CHNG"
INDICATOR_ID = "SW_10_01_ORD_CHNG"
INDICATOR_NAME = "Sales Order SLA Agreement Analysis - detailed"

FRAMEWORK_PARAMS = [
    "LANGU",
    "BACKDAYS",
    "MANAGE_IN_UTC",
    "DURATION_UNIT",
    "DATE_REF_FLD",
    "SW_DEST",
]


def _find(*globs: str) -> Path | None:
    for base in (INPUT, OLD):
        for g in globs:
            hits = sorted(base.glob(g))
            if hits:
                return hits[0]
    return None


def main() -> None:
    OLD.mkdir(exist_ok=True)

    for pat in [
        "*MD_CHNG*",
        "*CREDIT_APP*",
        "*Credit Management*",
        "Code_Sales Order SLA*",
        "Available fields_Sales Order SLA*",
        "Metadata *ORD_CHNG*",
        "Metadata _SKN_S_SW_10_01_ORD_CHNG.xlsx",
        "Code_SKN_S_SW_10_01_ORD_CHNG.txt",
        "Available fields_SKN_S_SW_10_01_ORD_CHNG.xlsx",
    ]:
        for p in list(INPUT.glob(pat)):
            if p.is_file() and p.parent == INPUT:
                # Keep structure in place if it is already the canonical name
                if p.name == f"Structure_{STEM}.xlsx":
                    continue
                dest = OLD / p.name
                if dest.exists():
                    dest.unlink()
                shutil.move(str(p), str(dest))

    src_code = _find(
        "Code_*ORD_CHNG*",
        "Code_*Sales Order SLA*",
    )
    src_avail = _find(
        "Available*ORD_CHNG*",
        "Available*Sales Order SLA*",
    )
    if not src_code or not src_avail:
        raise SystemExit("missing code or available-fields source for ORD_CHNG")

    code_canon = INPUT / f"Code_{STEM}.txt"
    avail_canon = INPUT / f"Available fields_{STEM}.xlsx"
    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    meta_path = INPUT / f"Metadata _{STEM}.xlsx"

    # Prefer source still in input; otherwise restore from old
    if src_code.parent == OLD or src_code.resolve() != code_canon.resolve():
        code_canon.write_text(src_code.read_text(encoding="utf-8"), encoding="utf-8")
    if src_avail.resolve() != avail_canon.resolve():
        shutil.copy2(src_avail, avail_canon)

    if not struct_path.exists():
        # Try restore from old first
        old_struct = OLD / struct_path.name
        if old_struct.exists():
            shutil.copy2(old_struct, struct_path)
        else:
            wb_af = openpyxl.load_workbook(avail_canon, read_only=True)
            ws_af = wb_af["Available Fields"]
            fields = []
            for r in ws_af.iter_rows(min_row=3, values_only=True):
                if r and r[0] and str(r[0]).strip() != "Field":
                    fields.append(list(r)[:7])
            wb_af.close()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Structure"
            ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
            for f in fields:
                fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
                dtype = f"{typ}({ln})" if typ and ln else (typ or "")
                ws.append([STRUCT_NAME, fname, desc or "", dtype, de or dom or ""])
            wb.save(struct_path)
            wb.close()

    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "General"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    wb = openpyxl.load_workbook(avail_canon)
    ws = wb["Parameters"]
    oldp: dict[str, list] = {}
    ordered: list[str] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or not r[0]:
            continue
        key = str(r[0]).strip().upper()
        if key == "LANG":
            key = "LANGU"
            row = list(r)
            row[0] = "LANGU"
            oldp[key] = row
        else:
            oldp[key] = list(r)
        name = "LANGU" if key == "LANGU" else str(r[0]).strip()
        if name.upper() not in {x.upper() for x in ordered}:
            ordered.append(name)
    tail = [p for p in ordered if p.upper() not in {f.upper() for f in FRAMEWORK_PARAMS}]
    code_params: list[str] = []
    seen: set[str] = set()
    for p in FRAMEWORK_PARAMS + tail:
        u = p.upper()
        if u in seen:
            continue
        seen.add(u)
        code_params.append(p)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    ws["A1"] = f"Parameters, #of Fields = {len(code_params)}"
    extras = {
        "LANGU": ["Language for texts", "LANG", "1", "0", "LANGU", "SPRAS"],
        "BACKDAYS": ["Backdays", "", "0", "0", "", ""],
        "MANAGE_IN_UTC": ["'X' - Manage in UTC", "", "0", "0", "", ""],
        "DURATION_UNIT": ["Duration Unit", "CHAR", "1", "0", "/SKN/E_SW_DURATION_UNIT", "/SKN/D_SW_DURATION_UNIT"],
        "DATE_REF_FLD": ["Date reference field", "CHAR", "30", "0", "NAME_FELD", "FDNAME"],
        "SW_DEST": ["RFC destination", "CHAR", "32", "0", "RFCDEST", "RFCDEST"],
    }
    for i, fld in enumerate(code_params, start=3):
        row = oldp.get(fld.upper()) or extras.get(fld)
        ws.cell(i, 1, fld)
        if row:
            for c in range(2, 8):
                if len(row) > c - 1 and row[c - 1] not in (None, ""):
                    ws.cell(i, c, row[c - 1])
    wb.save(avail_canon)
    wb.close()

    print(f"ready {len(code_params)} params, metadata at {meta_path.name}")


if __name__ == "__main__":
    main()
