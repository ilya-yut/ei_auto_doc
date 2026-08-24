"""Bootstrap input files for SW_10_01_INV_L_POST (SD Billing Doc invoice list posting status)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
OLD = INPUT / "old"
STEM = "SKN_S_SW_10_01_BILL_STAT"
STRUCT_NAME = "/SKN/S_SW_10_01_BILL_STAT"
INDICATOR_ID = "SW_10_01_INV_L_POST"
INDICATOR_NAME = "SD Billing Doc.- invoice list only -post. stat."

CODE_PARAMS = [
    "MANAGE_IN_UTC",
    "LANGU",
    "BACKDAYS",
    "BP1_FUNCT",
    "BP2_FUNCT",
    "BP3_FUNCT",
    "DATE_REF_FLD",
    "DURATION_UNIT",
    "VBELN",
    "FKART",
    "FKTYP",
    "VBTYP",
    "VKORG",
    "VTWEG",
    "KDGRP",
    "BZIRK",
    "FKDAT",
    "ERDAT",
    "ERNAM",
    "AEDAT",
    "DATUM",
    "DURATION",
    "KUNRG",
    "KUNAG",
    "SPART",
    "BUCHK",
    "RELIK",
    "RRSTA",
    "BLOCK",
    "RFBSK",
    "BP1_CODE",
    "BP2_CODE",
    "BP3_CODE",
    "FKSTO",
    "BP1_NAME",
    "BP2_NAME",
    "BP3_NAME",
    "COSTA",
    "DUMMY",
    "DURATION_D",
    "ERZET",
    "FMSTK",
    "NETWR",
    "PAYER_DESC",
    "SOLDTO_DESC",
    "UVK01",
    "UVK02",
    "UVK03",
    "UVK04",
    "UVK05",
    "UVS01",
    "UVS02",
    "UVS03",
    "UVS04",
    "UVS05",
    "WAERK",
    "SW_DEST",
]


def main() -> None:
    OLD.mkdir(exist_ok=True)

    for pat in ["*CUST_CRDT*", "*Customer Credit*"]:
        for p in list(INPUT.glob(pat)):
            if p.is_file():
                dest = OLD / p.name
                if dest.exists():
                    dest.unlink()
                shutil.move(str(p), str(dest))

    src_code = next(INPUT.glob("Code_*INV_L_POST*"), None) or next(
        INPUT.glob("Code_*Billing Doc*"), None
    )
    src_avail = next(INPUT.glob("Available*INV_L_POST*"), None) or next(
        INPUT.glob("Available*Billing Doc*"), None
    )
    if not src_code or not src_avail:
        raise SystemExit("missing code or available-fields source for BILL_STAT")

    code_canon = INPUT / f"Code_{STEM}.txt"
    avail_canon = INPUT / f"Available fields_{STEM}.xlsx"
    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    meta_path = INPUT / f"Metadata _{STEM}.xlsx"

    code_canon.write_text(src_code.read_text(encoding="utf-8"), encoding="utf-8")
    if src_code.resolve() != code_canon.resolve() and src_code.parent == INPUT:
        src_code.unlink()

    shutil.copy2(src_avail, avail_canon)
    if src_avail.resolve() != avail_canon.resolve() and src_avail.parent == INPUT:
        src_avail.unlink()

    if not struct_path.exists():
        wb_af = openpyxl.load_workbook(avail_canon, read_only=True)
        ws_af = wb_af["Available Fields"]
        fields = []
        for r in ws_af.iter_rows(min_row=3, values_only=True):
            if r and r[0] and str(r[0]).strip() != "Field":
                fields.append(list(r)[:7])
        wb_af.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Structure"
        ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
        for f in fields:
            fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
            dtype = f"{typ}({ln})" if typ and ln else (typ or "")
            ws.append([STRUCT_NAME, fname, desc or "", dtype, de or dom or ""])
        wb.save(struct_path)
        wb.close()

    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "General"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    wb = openpyxl.load_workbook(avail_canon)
    ws = wb["Parameters"]
    oldp = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r and r[0]:
            key = str(r[0]).strip().upper()
            if key == "LANG":
                key = "LANGU"
            oldp[key] = list(r)
            if str(r[0]).strip().upper() == "LANG":
                oldp[key][0] = "LANGU"
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
    extras = {
        "LANGU": ["Language for texts", "LANG", "1", "0", "LANGU", "SPRAS"],
        "BACKDAYS": ["Backdays", "INT4", "10", "0", "BACKDAYS", "BACKDAYS"],
        "DATE_REF_FLD": ["Date reference field", "", "0", "0", "", ""],
        "DURATION_UNIT": ["Duration unit", "CHAR", "1", "0", "/SKN/E_SW_DURATION_UNIT", "/SKN/D_SW_DURATION_UNIT"],
        "DATUM": ["Monitoring date range", "DATS", "8", "0", "DATUM", "DATUM"],
        "MANAGE_IN_UTC": ["Manage in UTC", "CHAR", "1", "0", "", "XFELD"],
        "SW_DEST": ["RFC destination", "CHAR", "32", "0", "RFCDEST", "RFCDEST"],
        "PAYER_DESC": ["Payer description", "CHAR", "35", "0", "NAME1_GP", "NAME"],
        "SOLDTO_DESC": ["Sold-to description", "CHAR", "35", "0", "NAME1_GP", "NAME"],
    }
    for i, fld in enumerate(CODE_PARAMS, start=3):
        row = oldp.get(fld.upper()) or extras.get(fld)
        ws.cell(i, 1, fld)
        if row:
            for c in range(2, 8):
                if len(row) > c - 1 and row[c - 1] not in (None, ""):
                    ws.cell(i, c, row[c - 1])
    wb.save(avail_canon)
    wb.close()

    print(f"ready {len(CODE_PARAMS)} params, metadata at {meta_path.name}")


if __name__ == "__main__":
    main()
