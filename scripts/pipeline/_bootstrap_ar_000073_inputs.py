"""Bootstrap input files for SW_10_01_AR_000073 (Sales order items and header)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
OLD = INPUT / "old"
STEM = "SKN_ARG_200005_000073"
STRUCT_NAME = "/SKN/ARG_200005_000073"
INDICATOR_ID = "SW_10_01_AR_000073"
INDICATOR_NAME = "Sales order items and header"


def main() -> None:
    OLD.mkdir(exist_ok=True)

    for pat in ["*ADDR_CH*", "*changed addresses*"]:
        for p in list(INPUT.glob(pat)):
            if p.is_file():
                dest = OLD / p.name
                if dest.exists():
                    dest.unlink()
                shutil.move(str(p), str(dest))

    src_code = next(INPUT.glob("Code_*AR_000073*"), None) or next(
        INPUT.glob("Code_*Sales order*"), None
    )
    src_avail = next(INPUT.glob("Available*AR_000073*"), None) or next(
        INPUT.glob("Available*Sales order*"), None
    )
    if not src_code or not src_avail:
        raise SystemExit("missing code or available-fields source for AR_000073")

    code_canon = INPUT / f"Code_{STEM}.txt"
    avail_canon = INPUT / f"Available fields_{STEM}.xlsx"
    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    meta_path = INPUT / f"Metadata _{STEM}.xlsx"

    code_canon.write_text(src_code.read_text(encoding="utf-8"), encoding="utf-8")
    if src_code.resolve() != code_canon.resolve() and src_code.parent == INPUT:
        src_code.unlink()

    shutil.copy2(src_avail, avail_canon)
    if src_avail.resolve() != avail_canon.resolve() and src_avail.parent == INPUT:
        src_avail.unlink()

    old_struct = list(INPUT.glob("Structure_*000073*"))
    for p in old_struct:
        if p.resolve() != struct_path.resolve():
            p.unlink(missing_ok=True)

    wb_af = openpyxl.load_workbook(avail_canon, read_only=True)
    ws_af = wb_af["Available Fields"]
    fields = []
    for r in ws_af.iter_rows(min_row=3, values_only=True):
        if r and r[0] and str(r[0]).strip() != "Field":
            fields.append(list(r)[:7])
    wb_af.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Structure"
    ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
    for f in fields:
        fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
        dtype = f"{typ}({ln})" if typ and ln else (typ or "")
        ws.append([STRUCT_NAME, fname, desc or "", dtype, de or dom or ""])
    wb.save(struct_path)
    wb.close()

    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "General"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    print(f"ready {len(fields)} output fields, metadata at {meta_path.name}")


if __name__ == "__main__":
    main()
