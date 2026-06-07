"""
Bootstrap Metadata, canonical Code/Available fields names, and Structure rows for
SKN_S_SW_10_08_EMP_USER_CHK from HR Employee - SAP User Association Check inputs.
Run from repo root: python scripts/pipeline/_bootstrap_emp_user_chk_inputs.py
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
OLD = INPUT / "old"
STEM = "SKN_S_SW_10_08_EMP_USER_CHK"
STRUCT_NAME = "/SKN/S_SW_10_08_EMP_USER_CHK"
INDICATOR_ID = "SW_10_08_EMP_USR_CHK"
INDICATOR_NAME = "HR Employee - SAP User Association Check"
SRC_CODE = INPUT / "Code_HR_ Employee - SAP User Association Check_SW_10_08_EMP_USR_CHK.txt"
SRC_AVAIL = INPUT / "Available fields_HR_ Employee - SAP User Association Check_SW_10_08_EMP_USR_CHK.xlsx"
CODE_CANON = INPUT / f"Code_{STEM}.txt"
AVAIL_CANON = INPUT / f"Available fields_{STEM}.xlsx"
STRUCT_PATH = INPUT / f"Structure_{STEM}.xlsx"
META_PATH = INPUT / f"Metadata _{STEM}.xlsx"


def main() -> None:
    OLD.mkdir(exist_ok=True)

    if SRC_CODE.exists():
        if CODE_CANON.exists() and CODE_CANON.resolve() != SRC_CODE.resolve():
            shutil.move(str(CODE_CANON), str(OLD / CODE_CANON.name))
        shutil.move(str(SRC_CODE), str(CODE_CANON))
    elif not CODE_CANON.exists():
        raise SystemExit(f"Missing {SRC_CODE} and {CODE_CANON}")

    if SRC_AVAIL.exists():
        if AVAIL_CANON.exists() and AVAIL_CANON.resolve() != SRC_AVAIL.resolve():
            shutil.move(str(AVAIL_CANON), str(OLD / AVAIL_CANON.name))
        shutil.move(str(SRC_AVAIL), str(AVAIL_CANON))
    elif not AVAIL_CANON.exists():
        raise SystemExit(f"Missing {SRC_AVAIL} and {AVAIL_CANON}")

    wb_a = openpyxl.load_workbook(AVAIL_CANON, read_only=True)
    ws_p = wb_a["Parameters"]
    param_rows = []
    for row in ws_p.iter_rows(min_row=3, values_only=True):
        fld = str(row[0] or "").strip() if row else ""
        if not fld:
            continue
        desc = str(row[1] or fld).strip()
        typ = str(row[2] or "CHAR").strip() if row[2] else "CHAR"
        ln = str(row[3] or "50").strip() if row[3] is not None else "50"
        de = str(row[5] or fld).strip() if len(row) > 5 and row[5] else fld
        param_rows.append((fld, desc, typ, ln, de))
    wb_a.close()

    wb_s = openpyxl.Workbook()
    ws_s = wb_s.active
    ws_s.title = "Structure"
    ws_s.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
    for fld, desc, typ, ln, de in param_rows:
        ws_s.append([STRUCT_NAME, fld, desc, f"{typ}({ln})", de])
    wb_s.save(STRUCT_PATH)
    wb_s.close()

    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "Metadata general"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(META_PATH)
    wb_m.close()

    print("Wrote:", META_PATH.name, STRUCT_PATH.name)
    print("Code:", CODE_CANON.name)
    print("Available fields:", AVAIL_CANON.name)
    print("Structure/output fields:", len(param_rows))


if __name__ == "__main__":
    main()
