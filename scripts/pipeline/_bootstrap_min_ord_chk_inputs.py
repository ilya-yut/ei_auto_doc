"""Bootstrap input files for SW_10_01_MIN_ORD_CHK (Minimum order quantity violation)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
OLD = INPUT / "old"
STEM = "SKN_S_SW_10_01_MIN_ORD_CHK"
STRUCT_NAME = "/SKN/S_SW_10_01_MIN_ORD_CHK"
INDICATOR_ID = "SW_10_01_MIN_ORD_CHK"
INDICATOR_NAME = "Minimum order quantity violation"

CODE_PARAMS = [
    "SW_DEST",
    "BACKDAYS",
    "MANAGE_IN_UTC",
    "DATE_REF_FLD",
    "DIV_ALERT_CHK",
    "MATNR",
    "VBELN",
    "WERKS",
    "VBTYP",
    "AUART",
    "KUNNR",
    "PSTYV",
    "ERDAT",
    "AEDAT",
    "DATUM",
    "ABSTA",
    "AUMNG",
    "BESTA",
    "KLMENG",
    "KWMENG",
    "LAND1",
    "NAME1",
    "NAME2",
    "ORT01",
    "POSNR",
    "PSTLZ",
    "PSTYV",
    "TELF1",
    "VKORG",
    "VTWEG",
]


def main() -> None:
    OLD.mkdir(exist_ok=True)

    for pat in ["*BILL_STAT*", "*INV_L_POST*", "*Billing Doc*"]:
        for p in list(INPUT.glob(pat)):
            if p.is_file():
                dest = OLD / p.name
                if dest.exists():
                    dest.unlink()
                shutil.move(str(p), str(dest))

    src_code = next(INPUT.glob("Code_*MIN_ORD*"), None) or next(
        INPUT.glob("Code_*Minimum order*"), None
    )
    src_avail = next(INPUT.glob("Available*MIN_ORD*"), None) or next(
        INPUT.glob("Available*Minimum order*"), None
    )
    if not src_code or not src_avail:
        raise SystemExit("missing code or available-fields source for MIN_ORD_CHK")

    code_canon = INPUT / f"Code_{STEM}.txt"
    avail_canon = INPUT / f"Available fields_{STEM}.xlsx"
    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    meta_path = INPUT / f"Metadata _{STEM}.xlsx"

    code_canon.write_text(src_code.read_text(encoding="utf-8"), encoding="utf-8")
    if src_code.resolve() != code_canon.resolve() and src_code.parent == INPUT:
        src_code.unlink()

    shutil.copy2(src_avail, avail_canon)
    if src_avail.resolve() != avail_canon.resolve() and src_avail.parent == INPUT:
        src_avail.unlink()

    if not struct_path.exists():
        wb_af = openpyxl.load_workbook(avail_canon, read_only=True)
        ws_af = wb_af["Available Fields"]
        fields = []
        for r in ws_af.iter_rows(min_row=3, values_only=True):
            if r and r[0] and str(r[0]).strip() != "Field":
                fields.append(list(r)[:7])
        wb_af.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Structure"
        ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
        for f in fields:
            fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
            dtype = f"{typ}({ln})" if typ and ln else (typ or "")
            ws.append([STRUCT_NAME, fname, desc or "", dtype, de or dom or ""])
        wb.save(struct_path)
        wb.close()

    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "General"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    wb = openpyxl.load_workbook(avail_canon)
    ws = wb["Parameters"]
    oldp = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r and r[0]:
            oldp[str(r[0]).strip().upper()] = list(r)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
    extras = {
        "BACKDAYS": ["Backdays", "INT4", "10", "0", "BACKDAYS", "BACKDAYS"],
        "DATE_REF_FLD": ["Date reference field", "", "0", "0", "", ""],
        "DIV_ALERT_CHK": ["Divisibility alert check", "CHAR", "1", "0", "", "XFELD"],
        "DATUM": ["Monitoring date range", "DATS", "8", "0", "DATUM", "DATUM"],
        "MANAGE_IN_UTC": ["Manage in UTC", "CHAR", "1", "0", "", "XFELD"],
        "SW_DEST": ["RFC destination", "CHAR", "32", "0", "RFCDEST", "RFCDEST"],
    }
    for i, fld in enumerate(CODE_PARAMS, start=3):
        row = oldp.get(fld.upper()) or extras.get(fld)
        ws.cell(i, 1, fld)
        if row:
            for c in range(2, 8):
                if len(row) > c - 1 and row[c - 1] not in (None, ""):
                    ws.cell(i, c, row[c - 1])
    wb.save(avail_canon)
    wb.close()

    print(f"ready {len(CODE_PARAMS)} params, metadata at {meta_path.name}")


if __name__ == "__main__":
    main()
