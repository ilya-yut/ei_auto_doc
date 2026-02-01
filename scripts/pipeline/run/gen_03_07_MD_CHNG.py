"""Generate 03_response.md and 07_response.md for Customer's address details were changed (Master Data Change Log) EI."""
import openpyxl
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
INPUT_DIR = PROJECT_ROOT / "input"
RUN_DIR = Path(__file__).resolve().parent

params_path = INPUT_DIR / "Available fields_Master Data General Change Log_SW_10_06_ANY_MD_CHNG (1).xlsx"
struct_path = INPUT_DIR / "Structure_Customer's address details were changed_200010_000165_ARP__EI__SW_10_06_ANY_MD_CHNG.xlsx"
code_path = INPUT_DIR / "Code_Customer's address details were changed_200010_000165_ARP__EI__SW_10_06_ANY_MD_CHNG.txt"


def c(r, i):
    return str(r[i]).strip() if r and len(r) > i and r[i] is not None else ""


# --- 03: Parameters Reference Table ---
wb = openpyxl.load_workbook(params_path, read_only=True, data_only=True)
ws = wb["Parameters"]
rows = list(ws.iter_rows(min_row=3, max_row=250, values_only=True))
wb.close()

intro = """### Parameters Reference Table

This table lists all configurable input parameters. Users set values for these parameters to filter and control the EI's data selection and processing logic.

| # | Parameter | Description | Type | Length | Decimal | Data Element | Domain |
|---|-----------|-------------|------|--------|---------|--------------|--------|"""
out = [intro]
num = 1
for i, r in enumerate(rows):
    if not r or not c(r, 0):
        continue
    out.append(
        "| " + str(num) + " | " + c(r, 0) + " | " + c(r, 1) + " | " + c(r, 2) + " | " + c(r, 3) + " | " + c(r, 4) + " | " + c(r, 5) + " | " + c(r, 6) + " |"
    )
    num += 1
    i += 1

(RUN_DIR / "03_response.md").write_text("\n".join(out), encoding="utf-8")
print("Wrote 03_response.md, param count:", num - 1)

# --- 07: EI Function Structure + ABAP Code ---
wb = openpyxl.load_workbook(struct_path, read_only=True, data_only=True)
ws = wb.active
struct_rows = list(ws.iter_rows(min_row=1, max_row=200, values_only=True))
wb.close()

header = """## EI Function Structure

This table lists all output fields returned by the EI. These fields contain the results of the EI's data retrieval and calculations.

| Structure Name | Field Name | Description | Data Type | Component Type |
|----------------|------------|-------------|-----------|----------------|"""
lines = [header]
for row in struct_rows[1:]:
    if not row or not c(row, 0):
        continue
    lines.append("| " + c(row, 0) + " | " + c(row, 1) + " | " + c(row, 2) + " | " + c(row, 3) + " | " + c(row, 4) + " |")

lines.append("")
lines.append("## ABAP Code")
lines.append("")
lines.append("```abap")
lines.append(code_path.read_text(encoding="utf-8", errors="replace"))
lines.append("```")
(RUN_DIR / "07_response.md").write_text("\n".join(lines), encoding="utf-8")
print("Wrote 07_response.md")
