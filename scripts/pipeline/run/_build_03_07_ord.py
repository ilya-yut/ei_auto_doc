"""Build 03_response.md and 07_response.md for Credit Memo / ORD_VAL_TOT EI."""
import openpyxl
from pathlib import Path

INPUT = Path(__file__).resolve().parent.parent.parent.parent / "input"
RUN = Path(__file__).resolve().parent

# Stem from manifest: Credit Memo Monthly volume by Payer _ 1M in LC_200019_000012
avail = INPUT / "Available fields_Credit Memo Monthly volume by Payer _ 1M in LC_200019_000012.xlsx"
struct_path = INPUT / "Structure_Credit Memo Monthly volume by Payer _ 1M in LC_200019_000012.xlsx"
code_path = INPUT / "Code_Credit Memo Monthly volume by Payer _ 1M in LC_200019_000012.txt"


def c(r, i):
    return str(r[i]).strip() if r and len(r) > i and r[i] is not None else ""


wb = openpyxl.load_workbook(avail, read_only=True, data_only=True)
ws = wb["Parameters"]
rows = list(ws.iter_rows(min_row=3, max_row=200, values_only=True))
wb.close()
params = []
for r in rows:
    if not r or not c(r, 0):
        continue
    params.append((c(r, 0), c(r, 1), c(r, 2), c(r, 3), c(r, 4), c(r, 5), c(r, 6)))
intro = "This table lists all configurable input parameters. Users set values for these parameters to filter and control the EI's data selection and processing logic."
lines = [
    "### Parameters Reference Table",
    "",
    intro,
    "",
    "| # | Parameter | Description | Type | Length | Decimal | Data Element | Domain |",
    "|---|-----------|-------------|------|--------|---------|--------------|--------|",
]
for i, p in enumerate(params, 1):
    lines.append("| " + str(i) + " | " + p[0] + " | " + p[1] + " | " + p[2] + " | " + p[3] + " | " + p[4] + " | " + p[5] + " | " + p[6] + " |")
(RUN / "03_response.md").write_text("\n".join(lines), encoding="utf-8")
print("03_response.md:", len(params), "params")

wb2 = openpyxl.load_workbook(struct_path, read_only=True, data_only=True)
struct = list(wb2.active.iter_rows(min_row=1, max_row=150, values_only=True))
wb2.close()
lines7 = [
    "## EI Function Structure",
    "",
    "This table lists all output fields returned by the EI.",
    "",
    "| Structure Name | Field Name | Description | Data Type | Component Type |",
    "|----------------|------------|-------------|-----------|----------------|",
]
for row in struct[1:] if struct else []:
    if not row or not c(row, 0):
        continue
    lines7.append("| " + c(row, 0) + " | " + c(row, 1) + " | " + c(row, 2) + " | " + c(row, 3) + " | " + c(row, 4) + " |")
lines7.extend(["", "## ABAP Code", "", "```abap", code_path.read_text(encoding="utf-8", errors="replace"), "```"])
(RUN / "07_response.md").write_text("\n".join(lines7), encoding="utf-8")
print("07_response.md done")
