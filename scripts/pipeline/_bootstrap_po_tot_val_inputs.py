"""Bootstrap input files for CBC_PO_TOT_VAL (PO Total Group Value Control)."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "input"
OLD = INPUT / "old"
STEM = "ZSWS_CBC_10_03_PO_PER_TOT_VAL"
STRUCT_NAME = "ZSWS_CBC_10_03_PO_PER_TOT_VAL"
INDICATOR_ID = "CBC_PO_TOT_VAL"
INDICATOR_NAME = "PO - Total Group Value Control"

CODE_PARAMS = [
    "LANGU",
    "BACKDAYS",
    "DATE_REF_FLD",
    "DURATION_UNIT",
    "FRGRL",
    "EBELN",
    "BUKRS",
    "BSTYP",
    "BSART",
    "EKORG",
    "EKGRP",
    "FRGGR",
    "FRGSX",
    "FRGCO",
    "LIFNR",
    "RESWK",
    "ZTERM",
    "ERNAM",
    "AEDAT",
    "BEDAT",
    "WAERS",
    "PROCSTAT",
    "DATUM",
    "DURATION",
    "PO_GRP_AMOUNT",
]

SRC_CODE_GLOB = "Code_*PO_TOT_VAL*"
SRC_AVAIL_GLOB = "Available*PO_TOT_VAL*"


def _find(glob_pat: str) -> Path | None:
    for base in (INPUT, OLD):
        hits = sorted(base.glob(glob_pat))
        if hits:
            return hits[0]
    return None


def main() -> None:
    OLD.mkdir(exist_ok=True)

    for pat in ["*MD_CHNG_LOG*", "*CUST_CHNG*"]:
        for p in list(INPUT.glob(pat)):
            if p.is_file():
                dest = OLD / p.name
                if dest.exists():
                    dest.unlink()
                shutil.move(str(p), str(dest))

    src_code = _find(SRC_CODE_GLOB)
    src_avail = _find(SRC_AVAIL_GLOB)
    if not src_code or not src_avail:
        raise SystemExit("Missing code or available-fields source for PO_TOT_VAL")

    code_canon = INPUT / f"Code_{STEM}.txt"
    avail_canon = INPUT / f"Available fields_{STEM}.xlsx"
    struct_path = INPUT / f"Structure_{STEM}.xlsx"
    meta_path = INPUT / f"Metadata _{STEM}.xlsx"

    code_canon.write_text(src_code.read_text(encoding="utf-8"), encoding="utf-8")
    if src_code.parent == INPUT and src_code != code_canon:
        src_code.unlink()

    shutil.copy2(src_avail, avail_canon)
    if src_avail.parent == INPUT and src_avail.resolve() != avail_canon.resolve():
        src_avail.unlink()

    wb_af = openpyxl.load_workbook(avail_canon, read_only=True)
    ws_af = wb_af["Available Fields"]
    fields = []
    for r in ws_af.iter_rows(min_row=3, values_only=True):
        if r and r[0] and str(r[0]).strip() != "Field":
            fields.append(list(r)[:7])
    wb_af.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Structure"
    ws.append(["Structure Name", "Field Name", "Description", "Data Type", "Component Type"])
    for f in fields:
        fname, desc, typ, ln, dec, de, dom = (f + [None] * 7)[:7]
        dtype = f"{typ}({ln})" if typ and ln else (typ or "")
        ws.append([STRUCT_NAME, fname, desc or "", dtype, de or dom or ""])
    wb.save(struct_path)
    wb.close()

    wb_m = openpyxl.Workbook()
    ws_m = wb_m.active
    ws_m.title = "General"
    for _ in range(11):
        ws_m.append([""] * 4)
    ws_m["A8"] = "Exception indicator ID"
    ws_m["B8"] = INDICATOR_ID
    ws_m["A9"] = "Exception indicator name"
    ws_m["B9"] = INDICATOR_NAME
    wb_m.save(meta_path)
    wb_m.close()

    wb = openpyxl.load_workbook(avail_canon)
    ws = wb["Parameters"]
    oldp = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if r and r[0]:
            oldp[str(r[0]).strip().upper()] = list(r)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    ws["A1"] = f"Parameters, #of Fields = {len(CODE_PARAMS)}"
    extras = {
        "LANGU": ["Language for texts", "LANG", "1", "0", "LANGU", "SPRAS"],
        "DATE_REF_FLD": ["Date Ref Fld", "CHAR", "30", "0", "NAME_FELD", "NAME_FELD"],
        "DURATION_UNIT": ["Duration Unit", "CHAR", "1", "0", "/SKN/E_SW_DURATION_UNIT", "/SKN/D_SW_DURATION_UNIT"],
        "FRGRL": ["Release indicator", "CHAR", "1", "0", "FRGRL", "FRGRL"],
        "FRGGR": ["Release group", "CHAR", "2", "0", "FRGGR", "FRGGR"],
        "FRGSX": ["Release strategy", "CHAR", "2", "0", "FRGSX", "FRGSX"],
        "FRGCO": ["Release code", "CHAR", "2", "0", "FRGCO", "FRGCO"],
        "RESWK": ["Supplying plant", "CHAR", "4", "0", "RESWK", "WERKS"],
        "DATUM": ["Reference Date", "DATS", "8", "0", "DATUM", "DATUM"],
        "BACKDAYS": ["Days Backward", "INT4", "10", "0", "BACKDAYS", "BACKDAYS"],
        "PO_GRP_AMOUNT": ["Group Total Amount", "CURR", "13", "2", "BWERT", "WERT7"],
    }
    for i, fld in enumerate(CODE_PARAMS, start=3):
        row = oldp.get(fld.upper()) or extras.get(fld)
        ws.cell(i, 1, fld)
        if row:
            for c in range(2, 8):
                if len(row) > c - 1 and row[c - 1] not in (None, ""):
                    ws.cell(i, c, row[c - 1])
    wb.save(avail_canon)
    wb.close()

    print(f"ready {len(CODE_PARAMS)} params {len(fields)} fields")


if __name__ == "__main__":
    main()
