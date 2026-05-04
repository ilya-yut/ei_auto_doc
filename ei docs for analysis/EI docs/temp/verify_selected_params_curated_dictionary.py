from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


TEMP = Path(r"c:\vibe code dev\ei_auto_doc\ei docs for analysis\EI docs\temp")
CURATED = TEMP / "params_selected_curated.xlsx"
OUT_VERIFY = TEMP / "params_selected_curated_verification.xlsx"

REQUESTED = [
    "ERDAT",
    "CHANGENR",
    "TAB_DESC",
    "ACT_CHNGNO",
    "CHANGE_IND",
    "CHANGE_IND_DESC",
    "CHNGIND",
    "CHNGIND_DESC",
    "OBJECTCLAS",
    "OBJECT_DESC",
    "PLANCHNGNR",
    "UNIT_NEW",
    "UNIT_OLD",
    "MESSAGE",
    "OBJECT",
    "AUFNR",
]

BANNED_PATTERNS = [
    "technical selection parameter",
    "used in this monitor set",
    "align interpretation with ddic",
]


def read_curated() -> list[tuple[str, str, str, str]]:
    wb = load_workbook(CURATED, read_only=True)
    ws = wb["curated_dictionary"] if "curated_dictionary" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: list[tuple[str, str, str, str]] = []
    for r in range(9, ws.max_row + 1):
        p = ws.cell(r, 1).value
        e = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        n = ws.cell(r, 4).value
        if not p:
            continue
        out.append(
            (
                str(p).strip().upper(),
                str(e or "").strip(),
                str(c or "").strip().lower(),
                str(n or "").strip(),
            )
        )
    return out


def main() -> None:
    req_set = set(REQUESTED)
    rows = read_curated()
    row_params = [p for p, _e, _c, _n in rows]
    row_set = set(row_params)

    missing = sorted(req_set - row_set)
    unexpected = sorted(row_set - req_set)
    dupes = [p for p, n in Counter(row_params).items() if n > 1]

    banned: list[tuple[str, str, str]] = []
    length_issues: list[tuple[str, str, int]] = []

    for p, e, c, _n in rows:
        low = e.lower()
        if any(b in low for b in BANNED_PATTERNS):
            banned.append((p, c, e))
        wc = len(e.split())
        if wc < 6 or wc > 28:
            length_issues.append((p, c, wc))

    conf = Counter(c for _p, _e, c, _n in rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws["A1"] = "Verification summary"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Requested parameters"
    ws["B2"] = len(req_set)
    ws["A3"] = "Curated rows"
    ws["B3"] = len(rows)
    ws["A4"] = "Missing parameters"
    ws["B4"] = len(missing)
    ws["A5"] = "Unexpected parameters"
    ws["B5"] = len(unexpected)
    ws["A6"] = "Duplicate parameters"
    ws["B6"] = len(dupes)
    ws["A7"] = "Rows with banned phrases"
    ws["B7"] = len(banned)
    ws["A8"] = "Rows outside 6..28 words"
    ws["B8"] = len(length_issues)
    ws["A10"] = "Confidence high"
    ws["B10"] = conf.get("high", 0)
    ws["A11"] = "Confidence medium"
    ws["B11"] = conf.get("medium", 0)
    ws["A12"] = "Confidence low"
    ws["B12"] = conf.get("low", 0)

    ws2 = wb.create_sheet("issues")
    ws2.append(["issue_type", "parameter", "confidence", "details"])
    for c in range(1, 5):
        ws2.cell(1, c).font = Font(bold=True)
    for p in missing:
        ws2.append(["missing_parameter", p, "", "Requested but absent in curated output"])
    for p in unexpected:
        ws2.append(["unexpected_parameter", p, "", "Present in curated output but not requested"])
    for p in dupes:
        ws2.append(["duplicate_parameter", p, "", "More than one row for parameter"])
    for p, c, e in banned:
        ws2.append(["banned_phrase", p, c, e])
    for p, c, wc in length_issues:
        ws2.append(["length_out_of_range", p, c, f"word_count={wc}"])

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 100

    wb.save(OUT_VERIFY)

    print(f"Requested parameters: {len(req_set)}")
    print(f"Curated rows: {len(rows)}")
    print(f"Missing parameters: {len(missing)}")
    print(f"Unexpected parameters: {len(unexpected)}")
    print(f"Duplicate parameters: {len(dupes)}")
    print(f"Rows with banned phrases: {len(banned)}")
    print(f"Rows outside word-count target: {len(length_issues)}")
    print(f"Confidence high/medium/low: {conf.get('high',0)}/{conf.get('medium',0)}/{conf.get('low',0)}")
    print(f"Wrote {OUT_VERIFY}")


if __name__ == "__main__":
    main()
