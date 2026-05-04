from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from collections import Counter, defaultdict
from pathlib import Path

from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


ROOT = Path(r"c:\vibe code dev\ei_auto_doc\ei docs for analysis\EI docs")
TEMP = ROOT / "temp"

SOURCE_XLSX = TEMP / "params_used_ge4_excluding_25.xlsx"
OUT_XLSX = TEMP / "params_filtered_existing_vs_suggested.xlsx"


def _load_sap_suggestions() -> dict[str, str]:
    repo = Path(r"c:\vibe code dev\ei_auto_doc")
    sys.path.insert(0, str(repo / "tools"))
    from sap_unified_param_texts import SAP_UNIFIED_EXPLANATION  # type: ignore

    return dict(SAP_UNIFIED_EXPLANATION)


RE_PARAM_HEAD = re.compile(r"^([A-Z][A-Z0-9_]*)\s*\(.*\)\s*:?\s*$")
RE_OPTIONS = re.compile(r"^(?:\*\*)?[A-Z][A-Z0-9_]*\s*Options:\s*(?:\*\*)?$", re.I)
RE_TOK = re.compile(r"\b[A-Z][A-Z0-9_]+\b")

STOP_HEADINGS = {
    "parameter relationship",
    "parameter relationships",
    "default values",
    "practical example of parameter configuration",
    "practical configuration examples",
    "ei function structure",
    "abap code",
}


def iter_block_items(parent: DocumentType):
    body = parent.element.body
    for child in body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


@dataclass
class ParsedDoc:
    explanations: dict[str, str]
    abap_text: str
    param_meta: dict[str, dict[str, str]]


def parse_docx(path: Path) -> ParsedDoc:
    doc = Document(str(path))
    in_guidelines = False
    in_abap = False
    saw_param_ref_heading = False
    explanations: dict[str, str] = {}
    param_meta: dict[str, dict[str, str]] = {}
    current_param: str | None = None
    buf: list[str] = []
    abap_lines: list[str] = []

    def flush_param() -> None:
        nonlocal current_param, buf
        if current_param:
            txt = " ".join(x.strip() for x in buf if x.strip())
            txt = re.sub(r"\s+", " ", txt).strip()
            explanations[current_param] = txt
        current_param = None
        buf = []

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            raw = block.text or ""
            t = raw.strip()
            low = t.lower()
            if not t:
                continue

            # Start/stop ABAP section
            if low == "abap code":
                flush_param()
                in_guidelines = False
                in_abap = True
                continue
            if in_abap:
                abap_lines.append(raw)
                if "ENDFUNCTION" in t.upper():
                    in_abap = False
                continue

            # Start guidelines
            if low == "parameter configuration guidelines":
                flush_param()
                in_guidelines = True
                continue
            if low == "parameters reference table":
                saw_param_ref_heading = True
                continue

            if in_guidelines:
                if low in STOP_HEADINGS:
                    flush_param()
                    in_guidelines = False
                    continue
                mh = RE_PARAM_HEAD.match(t)
                if mh:
                    flush_param()
                    current_param = mh.group(1).upper()
                    continue
                if RE_OPTIONS.match(t):
                    flush_param()
                    continue
                if current_param:
                    buf.append(raw)

        elif isinstance(block, Table):
            # Parse Parameters Reference Table metadata when available
            if saw_param_ref_heading:
                rows = [[c.text.strip() for c in row.cells] for row in block.rows]
                if rows:
                    header = [h.strip().lower() for h in rows[0]]
                    # Expected headers: #, Field, Description, Type, Length, Decimal, Data Element, Domain
                    col = {name: idx for idx, name in enumerate(header)}
                    if "field" in col and "description" in col:
                        for rr in rows[1:]:
                            if col["field"] >= len(rr):
                                continue
                            p = rr[col["field"]].strip().upper()
                            if not p or p in {"FIELD", "---"}:
                                continue
                            meta = {
                                "description": rr[col["description"]].strip() if col["description"] < len(rr) else "",
                                "type": rr[col["type"]].strip() if "type" in col and col["type"] < len(rr) else "",
                                "data_element": rr[col["data element"]].strip()
                                if "data element" in col and col["data element"] < len(rr)
                                else "",
                                "domain": rr[col["domain"]].strip() if "domain" in col and col["domain"] < len(rr) else "",
                            }
                            param_meta[p] = meta
                saw_param_ref_heading = False

            if in_abap:
                for row in block.rows:
                    for cell in row.cells:
                        for p in cell.paragraphs:
                            tx = p.text.strip()
                            if tx:
                                abap_lines.append(tx)
                                if "ENDFUNCTION" in tx.upper():
                                    in_abap = False

    flush_param()
    return ParsedDoc(explanations=explanations, abap_text="\n".join(abap_lines), param_meta=param_meta)


def read_source_rows(path: Path) -> list[tuple[str, str, str]]:
    wb = load_workbook(path, read_only=True)
    ws = wb["params_ge4"] if "params_ge4" in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row = None
    for r in range(1, ws.max_row + 1):
        vals = [(ws.cell(r, c).value or "") for c in range(1, 6)]
        low = [str(v).strip().lower() for v in vals]
        if len(low) >= 3 and low[0] == "parameter" and low[1] == "sub-folder" and low[2] == "filename":
            header_row = r
            break
    if header_row is None:
        raise RuntimeError("Could not find detail header row (parameter/sub-folder/filename) in source xlsx.")

    out: list[tuple[str, str, str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        p = ws.cell(r, 1).value
        s = ws.cell(r, 2).value
        f = ws.cell(r, 3).value
        if p is None and s is None and f is None:
            continue
        p = str(p or "").strip()
        s = str(s or "").strip()
        f = str(f or "").strip()
        if not p or not s or not f:
            continue
        out.append((p.upper(), s, f))
    return out


def build_suggested(param: str, parsed: ParsedDoc, sap_suggest: dict[str, str]) -> str:
    base = sap_suggest.get(param, "")
    abap_upper = parsed.abap_text.upper()
    tokens = set(RE_TOK.findall(abap_upper))
    in_abap = param in tokens
    if base:
        return base

    meta = parsed.param_meta.get(param, {})
    desc = (meta.get("description") or "").strip()
    de = (meta.get("data_element") or "").strip()
    dom = (meta.get("domain") or "").strip()
    typ = (meta.get("type") or "").strip()

    if desc:
        parts: list[str] = [f"{param} controls selection by {desc.lower()}"]
        ddic_bits: list[str] = []
        if de:
            ddic_bits.append(f"DE {de}")
        if dom:
            ddic_bits.append(f"Domain {dom}")
        if typ:
            ddic_bits.append(f"Type {typ}")
        if ddic_bits:
            parts.append(f"({'; '.join(ddic_bits)}).")
        else:
            parts.append(".")
        if in_abap:
            parts.append("Used in ABAP filtering/logic for this EI.")
        return " ".join(parts)

    existing = parsed.explanations.get(param, "").strip()
    if existing:
        # Keep concise: first sentence-like chunk.
        first = re.split(r"(?<=[.!?])\s+", existing)[0].strip()
        if not first.endswith("."):
            first += "."
        return first

    if in_abap:
        return f"{param} is used in ABAP selection/processing logic for this EI."
    return f"{param} is a technical selection parameter in this EI; align description to DDIC semantics."


def _canonical_suggestion_for_param(
    param: str,
    docs_for_param: list[ParsedDoc],
    sap_suggest: dict[str, str],
) -> str:
    """Return one concise SAP-grade suggestion per parameter across all files."""
    base = sap_suggest.get(param, "").strip()
    if base:
        return base

    desc_counter: Counter[str] = Counter()
    de_counter: Counter[str] = Counter()
    dom_counter: Counter[str] = Counter()
    typ_counter: Counter[str] = Counter()
    abap_hits = 0

    for pd in docs_for_param:
        meta = pd.param_meta.get(param, {})
        d = (meta.get("description") or "").strip()
        if d:
            desc_counter[d] += 1
        de = (meta.get("data_element") or "").strip()
        if de:
            de_counter[de] += 1
        dom = (meta.get("domain") or "").strip()
        if dom:
            dom_counter[dom] += 1
        typ = (meta.get("type") or "").strip()
        if typ:
            typ_counter[typ] += 1
        if re.search(rf"\b{re.escape(param)}\b", pd.abap_text.upper()):
            abap_hits += 1

    # Prefer meaningful descriptions over placeholders like "Not used".
    meaningful_desc = Counter(
        {
            k: v
            for k, v in desc_counter.items()
            if "not used" not in k.lower() and "unused" not in k.lower()
        }
    )
    if meaningful_desc:
        desc = meaningful_desc.most_common(1)[0][0]
    else:
        desc = desc_counter.most_common(1)[0][0] if desc_counter else ""
    de = de_counter.most_common(1)[0][0] if de_counter else ""
    dom = dom_counter.most_common(1)[0][0] if dom_counter else ""
    typ = typ_counter.most_common(1)[0][0] if typ_counter else ""

    dlow = desc.lower().strip().rstrip(".")
    if dlow:
        if re.search(r"\bcount\b", dlow) or re.search(r"\bcnt\b", param.lower()):
            core = f"{param} defines the threshold/range for {dlow}."
        elif "date" in dlow:
            core = f"{param} scopes records by {dlow}."
        elif "time" in dlow:
            core = f"{param} refines records by {dlow}."
        elif "status" in dlow or "state" in dlow:
            core = f"{param} filters records by {dlow}."
        elif "user" in dlow or "uname" in param.lower():
            core = f"{param} filters records by {dlow}."
        elif "duration" in dlow:
            core = f"{param} controls elapsed-time based filtering by {dlow}."
        elif "destination" in dlow or "rfc" in dlow:
            core = f"{param} restricts processing scope by {dlow}."
        else:
            core = f"{param} scopes records by {dlow}."
        # If this parameter is effectively unused across occurrences, state that explicitly.
        if abap_hits == 0 and ("not used" in dlow or "unused" in dlow):
            core = f"{param} is a reserved/unused technical field in this monitor set."
        return core

    # Fallback when no useful description exists.
    if de or dom or typ:
        ddic = ", ".join(x for x in [de, dom, typ] if x)
        return f"{param} is used as a technical selection field ({ddic})."
    return f"{param} is used as a technical selection parameter in this monitor set."


def main() -> None:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Source xlsx not found: {SOURCE_XLSX}")

    sap_suggest = _load_sap_suggestions()
    requested = read_source_rows(SOURCE_XLSX)

    # Parse each docx once
    by_doc: dict[Path, ParsedDoc] = {}
    output_rows: list[tuple[str, str, str, str, str]] = []
    missing_docs: list[str] = []
    by_param_docs: dict[str, list[ParsedDoc]] = defaultdict(list)
    staged_rows: list[tuple[str, str, str, ParsedDoc]] = []

    for param, subfolder, filename in requested:
        docx_path = ROOT / subfolder / filename
        if not docx_path.exists():
            missing_docs.append(str(docx_path))
            continue
        if docx_path not in by_doc:
            by_doc[docx_path] = parse_docx(docx_path)
        parsed = by_doc[docx_path]
        by_param_docs[param].append(parsed)
        existing = parsed.explanations.get(param, "")
        if not existing:
            existing = "*(no matching parameter explanation found in Parameter Configuration Guidelines)*"
        staged_rows.append((param, existing, subfolder, filename, parsed))

    suggested_by_param = {
        p: _canonical_suggestion_for_param(p, docs, sap_suggest)
        for p, docs in by_param_docs.items()
    }
    for param, existing, subfolder, filename, _parsed in staged_rows:
        output_rows.append((param, existing, suggested_by_param[param], subfolder, filename))

    output_rows.sort(key=lambda x: (x[0], x[3], x[4]))

    wb = Workbook()
    ws = wb.active
    ws.title = "filtered_params"

    ws["A1"] = "Summary"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Source rows read"
    ws["B2"] = len(requested)
    ws["A3"] = "Rows written"
    ws["B3"] = len(output_rows)
    ws["A4"] = "Unique parameters"
    ws["B4"] = len({r[0] for r in output_rows})
    ws["A5"] = "Unique files"
    ws["B5"] = len({(r[3], r[4]) for r in output_rows})
    ws["A6"] = "Missing referenced files"
    ws["B6"] = len(missing_docs)

    hdr_row = 8
    headers = [
        "parameter",
        "existing explanation",
        "suggested/corrected explanation",
        "subfolder",
        "file name",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(hdr_row, c, h).font = Font(bold=True)

    r = hdr_row + 1
    for row in output_rows:
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
        r += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 68
    ws.column_dimensions["C"].width = 82
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 72

    wb.save(OUT_XLSX)

    print(f"Source rows read: {len(requested)}")
    print(f"Rows written: {len(output_rows)}")
    print(f"Unique parameters: {len({r[0] for r in output_rows})}")
    print(f"Unique files: {len({(r[3], r[4]) for r in output_rows})}")
    print(f"Missing referenced files: {len(missing_docs)}")
    if missing_docs:
        print("First missing examples:")
        for m in missing_docs[:5]:
            print(" -", m)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
