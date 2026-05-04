from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


ROOT = Path(r"c:\vibe code dev\ei_auto_doc")
PART1_DOCX_DIR = ROOT / "ei docs for analysis" / "EI docs" / "Part 1"
TEMP_DIR = ROOT / "ei docs for analysis" / "EI docs" / "temp"

SRC_MISSING_XLSX = TEMP_DIR / "params_dictionary_part1_missing.xlsx"
OUT_CURATED_XLSX = TEMP_DIR / "params_dictionary_part1_missing_curated.xlsx"

RE_NUM = re.compile(r"^\d+$")
RE_PARAM_TOKEN = re.compile(r"^[A-Z][A-Z0-9_]{1,}$")
RE_WORD = re.compile(r"[A-Za-z0-9_]+")


@dataclass
class ParamEvidence:
    descriptions: list[str]
    data_elements: list[str]
    domains: list[str]
    types: list[str]
    existing_explanations: list[str]
    files_seen: int
    abap_hits: int


def iter_block_items(parent: DocumentType):
    body = parent.element.body
    for child in body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


def _load_base_overrides() -> dict[str, str]:
    sys.path.insert(0, str(TEMP_DIR))
    from build_params_dictionary_xlsx import OVERRIDES  # type: ignore

    return {k.upper(): v.strip() for k, v in OVERRIDES.items()}


def read_missing_params(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True)
    ws = wb["missing_dictionary"] if "missing_dictionary" in wb.sheetnames else wb[wb.sheetnames[0]]
    params: list[str] = []
    # header row is 8 in current format
    for r in range(9, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        p = str(v).strip().upper()
        if p and RE_PARAM_TOKEN.match(p):
            params.append(p)
    # preserve order, unique
    seen = set()
    out: list[str] = []
    for p in params:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def _clean(s: str) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    return s


def parse_part1_docx_for_evidence(path: Path, target_params: set[str]) -> dict[str, dict[str, list[str] | int]]:
    doc = Document(str(path))
    saw_param_table = False
    in_guidelines = False
    in_abap = False

    by_param: dict[str, dict[str, list[str] | int]] = {}

    # guideline capture state
    current_param: str | None = None
    expl_buf: list[str] = []

    def flush_guideline_param() -> None:
        nonlocal current_param, expl_buf
        if current_param and current_param in target_params:
            txt = _clean(" ".join(expl_buf))
            if txt:
                by_param.setdefault(
                    current_param,
                    {
                        "descriptions": [],
                        "data_elements": [],
                        "domains": [],
                        "types": [],
                        "existing_explanations": [],
                        "abap_hits": 0,
                        "files_seen": 0,
                    },
                )
                by_param[current_param]["existing_explanations"].append(txt)  # type: ignore[index]
        current_param = None
        expl_buf = []

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            t = _clean(block.text)
            low = t.lower()
            if not t:
                continue

            if low == "parameters reference table":
                saw_param_table = True
                continue

            if low == "parameter configuration guidelines":
                flush_guideline_param()
                in_guidelines = True
                continue

            if low in {
                "parameter relationship",
                "parameter relationships",
                "default values",
                "practical example of parameter configuration",
                "practical configuration examples",
                "abap code",
            }:
                flush_guideline_param()
                if low == "abap code":
                    in_abap = True
                in_guidelines = False
                continue

            if in_guidelines:
                m = re.match(r"^([A-Z][A-Z0-9_]*)\s*\(.*\)\s*:?\s*$", t)
                if m:
                    flush_guideline_param()
                    current_param = m.group(1).upper()
                    continue
                if re.match(r"^(?:\*\*)?[A-Z][A-Z0-9_]*\s*Options:\s*(?:\*\*)?$", t, re.I):
                    flush_guideline_param()
                    continue
                if current_param:
                    expl_buf.append(t)

            if in_abap:
                words = {w.upper() for w in RE_WORD.findall(t)}
                for p in target_params:
                    if p in words:
                        by_param.setdefault(
                            p,
                            {
                                "descriptions": [],
                                "data_elements": [],
                                "domains": [],
                                "types": [],
                                "existing_explanations": [],
                                "abap_hits": 0,
                                "files_seen": 0,
                            },
                        )
                        by_param[p]["abap_hits"] = int(by_param[p]["abap_hits"]) + 1  # type: ignore[index]
                if "ENDFUNCTION" in t.upper():
                    in_abap = False

        elif isinstance(block, Table):
            if saw_param_table:
                rows = [[_clean(c.text) for c in row.cells] for row in block.rows]
                if rows:
                    header = [h.lower() for h in rows[0]]
                    if "field" in header:
                        col_field = header.index("field")
                        col_desc = header.index("description") if "description" in header else None
                        col_type = header.index("type") if "type" in header else None
                        col_de = header.index("data element") if "data element" in header else None
                        col_dom = header.index("domain") if "domain" in header else None
                        col_num = None
                        if "#" in rows[0]:
                            col_num = rows[0].index("#")
                        elif "no." in header:
                            col_num = header.index("no.")
                        elif "no" in header:
                            col_num = header.index("no")

                        for rr in rows[1:]:
                            if col_field >= len(rr):
                                continue
                            p = rr[col_field].strip().upper()
                            if not p or not RE_PARAM_TOKEN.match(p):
                                continue
                            if p not in target_params:
                                continue
                            if col_num is not None and col_num < len(rr):
                                n = rr[col_num].strip()
                                if n and not RE_NUM.match(n):
                                    break
                            by_param.setdefault(
                                p,
                                {
                                    "descriptions": [],
                                    "data_elements": [],
                                    "domains": [],
                                    "types": [],
                                    "existing_explanations": [],
                                    "abap_hits": 0,
                                    "files_seen": 0,
                                },
                            )
                            if col_desc is not None and col_desc < len(rr) and rr[col_desc]:
                                by_param[p]["descriptions"].append(rr[col_desc])  # type: ignore[index]
                            if col_de is not None and col_de < len(rr) and rr[col_de]:
                                by_param[p]["data_elements"].append(rr[col_de])  # type: ignore[index]
                            if col_dom is not None and col_dom < len(rr) and rr[col_dom]:
                                by_param[p]["domains"].append(rr[col_dom])  # type: ignore[index]
                            if col_type is not None and col_type < len(rr) and rr[col_type]:
                                by_param[p]["types"].append(rr[col_type])  # type: ignore[index]
                saw_param_table = False

    flush_guideline_param()

    for p in by_param:
        by_param[p]["files_seen"] = 1
    return by_param


def _pick_meaningful(counter: Counter[str]) -> str:
    for txt, _n in counter.most_common():
        t = txt.strip()
        if not t:
            continue
        low = t.lower()
        if "not used" in low or "unused" in low:
            continue
        return t
    return counter.most_common(1)[0][0] if counter else ""


def _make_from_desc(param: str, desc: str) -> str:
    d = desc.lower().strip().rstrip(".")
    if not d:
        return ""
    if "customer" in d:
        return f"{param} identifies the customer account and scopes records to the relevant customer population."
    if "vendor" in d or "supplier" in d:
        return f"{param} identifies the vendor account and scopes records to supplier-specific flows."
    if "material" in d:
        return f"{param} identifies the material and scopes records to product-specific transactions."
    if "company code" in d:
        return f"{param} scopes records by company code to the relevant legal entity."
    if "plant" in d:
        return f"{param} scopes records by plant to the relevant site-level operations."
    if "purchasing organization" in d:
        return f"{param} scopes records by purchasing organization for procurement control analysis."
    if "sales organization" in d:
        return f"{param} scopes records by sales organization for SD commercial analysis."
    if "currency" in d:
        return f"{param} defines the currency context used for value interpretation and filtering."
    if "status" in d or "state" in d:
        return f"{param} filters records by status/state to focus on relevant processing conditions."
    if "date" in d:
        return f"{param} scopes records by date to control the monitoring period."
    if "time" in d:
        return f"{param} refines records by time for precise intra-day event selection."
    if "count" in d:
        return f"{param} defines threshold/range control for count-based exception evaluation."
    if "amount" in d or "value" in d:
        return f"{param} is used for value-based filtering and anomaly threshold analysis."
    if "document number" in d or "number" in d:
        return f"{param} is a key identifier used for document-level traceability and drilldown."
    if "user" in d:
        return f"{param} filters records by user identity for accountability and access-focused analysis."
    if "transaction code" in d or "tcode" in d:
        return f"{param} filters records by transaction code to isolate process-specific activity."
    return f"{param} scopes records by {d}."


def _fallback_pattern(param: str) -> str:
    p = param.upper()
    if p.endswith("_DESC"):
        base = p[: -5]
        return f"{p} provides description text for {base} to keep reporting business-readable."
    if p.endswith("_OLD"):
        base = p[: -4]
        return f"{p} stores the previous value of {base} for before/after change analysis."
    if p.endswith("_NEW"):
        base = p[: -4]
        return f"{p} stores the new value of {base} for after-change impact analysis."
    if p.endswith("DAT") or p.endswith("DATE"):
        return f"{p} is a date selector used to constrain records to the relevant period."
    if p.endswith("TIM") or p.endswith("TIME"):
        return f"{p} is a time selector used to refine event windows within selected dates."
    if re.search(r"(CNT|COUNT)$", p):
        return f"{p} is a count metric used for threshold-based exception checks."
    if p.endswith("ICON"):
        return f"{p} is a visual status icon field used to present monitoring state."
    return f"{p} is an SAP technical field used to scope and contextualize records."


def build_curated(param: str, ev: ParamEvidence, overrides: dict[str, str]) -> tuple[str, str, str]:
    p = param.upper()
    if p in overrides:
        return overrides[p], "high", "curated override"

    desc = _pick_meaningful(Counter(_clean(x) for x in ev.descriptions if _clean(x)))
    explanation = _make_from_desc(p, desc) if desc else ""
    if explanation:
        conf = "high" if ev.files_seen >= 2 else "medium"
        note = f"desc='{desc[:80]}'"
        return explanation, conf, note

    # try existing guideline prose first sentence
    ex = _pick_meaningful(Counter(_clean(x) for x in ev.existing_explanations if _clean(x)))
    if ex:
        first = re.split(r"(?<=[.!?])\s+", ex)[0].strip()
        if not first.endswith("."):
            first += "."
        if len(first.split()) <= 28 and "technical selection parameter" not in first.lower():
            return first, "medium", "from existing guideline sentence"

    # pattern fallback (low)
    fb = _fallback_pattern(p)
    return fb, "low", "pattern fallback"


def main() -> None:
    target_params = read_missing_params(SRC_MISSING_XLSX)
    target_set = set(target_params)
    overrides = _load_base_overrides()

    # Aggregate evidence across Part 1 docs for each target param
    agg: dict[str, ParamEvidence] = {
        p: ParamEvidence([], [], [], [], [], 0, 0) for p in target_params
    }

    for fp in sorted(PART1_DOCX_DIR.glob("*.docx")):
        by_param = parse_part1_docx_for_evidence(fp, target_set)
        for p, raw in by_param.items():
            ev = agg[p]
            ev.files_seen += int(raw.get("files_seen", 0))  # type: ignore[arg-type]
            ev.abap_hits += int(raw.get("abap_hits", 0))  # type: ignore[arg-type]
            ev.descriptions.extend(raw.get("descriptions", []))  # type: ignore[arg-type]
            ev.data_elements.extend(raw.get("data_elements", []))  # type: ignore[arg-type]
            ev.domains.extend(raw.get("domains", []))  # type: ignore[arg-type]
            ev.types.extend(raw.get("types", []))  # type: ignore[arg-type]
            ev.existing_explanations.extend(raw.get("existing_explanations", []))  # type: ignore[arg-type]

    wb = Workbook()
    ws = wb.active
    ws.title = "curated_dictionary"

    ws["A1"] = "Summary"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Parameters requested"
    ws["B2"] = len(target_params)

    rows: list[tuple[str, str, str, str]] = []
    for p in target_params:
        expl, conf, note = build_curated(p, agg[p], overrides)
        rows.append((p, expl, conf, note))

    ws["A3"] = "Generated entries"
    ws["B3"] = len(rows)
    ws["A4"] = "High confidence"
    ws["B4"] = sum(1 for _p, _e, c, _n in rows if c == "high")
    ws["A5"] = "Medium confidence"
    ws["B5"] = sum(1 for _p, _e, c, _n in rows if c == "medium")
    ws["A6"] = "Low confidence"
    ws["B6"] = sum(1 for _p, _e, c, _n in rows if c == "low")

    hr = 8
    headers = ["parameter", "curated_explanation", "confidence", "evidence_notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(hr, i, h).font = Font(bold=True)

    r = hr + 1
    for row in rows:
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
        r += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    wb.save(OUT_CURATED_XLSX)

    print(f"Parameters requested: {len(target_params)}")
    print(f"Generated entries: {len(rows)}")
    print(f"High: {sum(1 for _p, _e, c, _n in rows if c == 'high')}")
    print(f"Medium: {sum(1 for _p, _e, c, _n in rows if c == 'medium')}")
    print(f"Low: {sum(1 for _p, _e, c, _n in rows if c == 'low')}")
    print(f"Wrote {OUT_CURATED_XLSX}")


if __name__ == "__main__":
    main()

