from __future__ import annotations

from pathlib import Path
import importlib.util
import re
from openpyxl import Workbook
from openpyxl.styles import Font


ROOT = Path(r"c:\vibe code dev\ei_auto_doc")
TEMP = ROOT / "ei docs for analysis" / "EI docs" / "temp"

SRC_CSV = Path(r"c:\Users\ilyay\Downloads\params.csv")
OUT_XLSX = Path(
    r"c:\vibe code dev\ei_auto_doc\ei docs for analysis\EI docs\temp\params_dictionary.xlsx"
)


def _load_sap_canonical_attached() -> dict[str, str]:
    path = TEMP / "sap_canonical_attached_params.py"
    spec = importlib.util.spec_from_file_location("sap_canonical_attached_params", path)
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return {str(k).upper(): str(v).strip() for k, v in mod.SAP_CANONICAL_ATTACHED.items()}


def _load_sap_unified_explanations() -> dict[str, str]:
    path = ROOT / "tools" / "sap_unified_param_texts.py"
    spec = importlib.util.spec_from_file_location("sap_unified_param_texts", path)
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return {str(k).upper(): str(v).strip() for k, v in mod.SAP_UNIFIED_EXPLANATION.items()}


OVERRIDES: dict[str, str] = {
    "ACCNT": "Account number key used to scope user/account records in the monitored dataset.",
    "ACTION": "Change-action indicator that distinguishes create/update/delete style events in change-log based monitors.",
    "AEDAT": "Changed-on date used to filter documents or master records by last maintenance activity.",
    "ANAME": "User who created the master/user record; used for creator-based accountability filtering.",
    "AREA": "Application/functional area key used to narrow logs or business process scope.",
    "AUART": "Sales document type controlling order category and therefore the SD process slice included in analysis.",
    "AUDAT": "Sales document date (order date) used for period-based SD selection.",
    "BANFN": "Purchase requisition number, the core PR document key for MM approval and lifecycle checks.",
    "BCDA1": "Date of last password change field used in user-security aging and credential hygiene checks.",
    "BEDAT": "Purchasing document date used to filter procurement documents by document creation period.",
    "BELNR": "Accounting document number, the primary FI document key for journal-level traceability.",
    "BLDAT": "Document date from the source business document, often used as legal/document reference date.",
    "BLOCK": "Blocking indicator showing whether the record is restricted for posting/processing.",
    "BNAME": "SAP user name used to restrict output to specific users or user populations.",
    "BNFPO": "Purchase requisition item number used to identify PR line-level records.",
    "BSAKZ": "Purchasing control/indicator flag used to segment procurement records by processing characteristics.",
    "BSART": "Purchasing document type that controls PO/PR business scenario and approval behavior.",
    "BSART_DESC": "Text description of purchasing document type used for readable reporting output.",
    "BSTNK": "Customer or external PO reference number used for cross-system document matching.",
    "BSTYP": "Purchasing document category (PR/PO/contract etc.) used to segment MM document classes.",
    "BSTYP_DESC": "Description of purchasing document category for business-readable output.",
    "BUDAT": "Posting date used to align analysis with accounting period recognition.",
    "BUKRS": "Company code key that scopes data to legal entity/accounting unit level.",
    "BUTXT": "Company code name/description used to present legal entity context in output.",
    "BWTAR": "Valuation type key used in split valuation scenarios (batch/material valuation layers).",
    "BZIRK": "Sales district key used for SD territory-level segmentation.",
    "CLASS": "User group/class used to filter users by administrative classification.",
    "CODV1": "Password hash/version component used in user credential state analysis.",
    "CODVN": "Password hash version indicator used to detect outdated credential hash schemes.",
    "CONVERT_KEY": "Technical conversion key used for formatting/normalization during output transformation.",
    "COSTA": "Confirmation/status indicator used to distinguish processing completion states.",
    "CPUDT": "Entry/creation date used for technical posting timestamp filtering.",
    "CUKY_NEW": "New currency key in change-log comparisons to detect currency master changes.",
    "CUKY_OLD": "Previous currency key in change-log comparisons for before/after analysis.",
    "CUST_DESC": "Customer description/name text used for readable customer-level reporting.",
    "DEPARTMENT": "Organizational department attribute used for responsibility-based slicing.",
    "DEVCLASS": "ABAP package/development class used to scope technical object ownership.",
    "DMBTR": "Amount in local currency used for FI valuation and threshold checks.",
    "DURATION_D": "Elapsed duration expressed in days for aging-style exception criteria.",
    "DURATION_H": "Elapsed duration expressed in hours for runtime or latency threshold checks.",
    "DURATION_M": "Elapsed duration expressed in minutes for near-real-time threshold checks.",
    "EBELN": "Purchasing document number (typically PO) used as the primary MM document key.",
    "EBELP": "Purchasing document item number used for line-level PO analytics.",
    "EKGRP": "Purchasing group (buyer) used for procurement ownership and control segmentation.",
    "EKGRP_DESC": "Description of purchasing group for readable buyer/team reporting.",
    "EKORG": "Purchasing organization key used to scope procurement flows by organizational unit.",
    "EKORG_DESC": "Description of purchasing organization for business-readable reporting.",
    "EKOTX": "Purchasing org/group related text field used for descriptive output enrichment.",
    "ELIKZ": "Delivery completed indicator used to identify open versus completed procurement items.",
    "ERNAM": "Created-by user ID used for maker-checker and ownership monitoring.",
    "ERZET": "Entry time used to refine timestamp windows within a selected day.",
    "ESTKZ": "Creation indicator for PR/PO source or method, used for process-origin analysis.",
    "FKART": "Billing document type used to segment SD billing scenarios.",
    "FKDAT": "Billing date used to align SD billing records with accounting/reporting periods.",
    "FNAME": "Field name key in change documents used to filter by changed attribute.",
    "FORWDAYS": "Forward-looking day horizon used for due-date/proactive exception windows.",
    "FRGC": "Release code in purchasing approvals used to analyze approval step ownership.",
    "FRGGR": "Release group key controlling the purchasing release strategy framework.",
    "FRGKE": "Release status indicator used to distinguish released vs unreleased documents.",
    "FRGRL": "Release indicator/flag used in PO/PR release strategy control.",
    "FRGZU": "Release strategy progression/status code used for approval lifecycle tracking.",
    "GLTGB": "Valid-to date used to check whether authorization/master data is still active.",
    "GLTGV": "Valid-from date used to ensure records are active in the analyzed period.",
    "GSBER": "Business area key used for FI organizational reporting segmentation.",
    "INIT_PWD_ICON": "Visual status icon indicating initial-password condition in user-security outputs.",
    "INSTANCENAME": "Application/HANA instance identifier used for server-level monitoring scope.",
    "JOBNAME": "Background job name used for scheduler/workload monitoring filters.",
    "KDAUF": "Sales order reference number at item level used for SD/MM document linkage.",
    "KDGRP": "Customer group key used for commercial segmentation in SD analysis.",
    "KDPOS": "Sales order item reference used to tie downstream records to original SD lines.",
    "KOKRS": "Controlling area key used for CO-level organizational scoping.",
    "KTOKK": "Account group (customer/vendor) used to segment master data governance rules.",
    "KUNAG": "Sold-to party/customer field used for SD partner-role based filtering.",
    "KUNNR": "KUNNR identifies the customer account and is used to scope records to specific customers across SD/FI flows.",
    "KUNRG": "Payer/customer field used to analyze SD/FI records by billing responsibility.",
    "LAND1": "Country key used for legal/geographic segmentation of business partners or plants.",
    "LANG": "Language key used for language-dependent texts and user-language filtering.",
    "LASTCHNAME": "Last changed by user field used for accountability and change ownership analysis.",
    "LFDAT": "Delivery date used for logistics due-date and fulfillment timeliness checks.",
    "LGORT": "Storage location used to segment stock/logistics movements by warehouse sub-location.",
    "LIFNR": "Vendor account number used to scope records to supplier-specific flows.",
    "LOCK_ICON": "Visual status icon indicating locked-state conditions in monitoring output.",
    "LOCNT": "Local count/occurrence metric used for threshold-based exception logic.",
    "LOEKZ": "Deletion indicator used to exclude logically deleted purchasing/material records.",
    "LTIME": "Local time field used for intra-day timestamp precision in selections.",
    "MATKL": "Material group key used for product-category segmentation in MM/SD analytics.",
    "MATNR": "Material number used as the primary product key across MM/SD records.",
    "MAT_DESC": "Material description text used to provide readable product context.",
    "MEINS": "Base unit of measure used to interpret quantity fields consistently.",
    "MEMSUM": "Aggregated memory usage metric used for technical capacity/performance thresholds.",
    "MENGE": "Quantity field used for volumetric thresholds and variance analysis.",
    "MODBE": "Modification area/module indicator used in user-change auditing.",
    "MODDA": "Modification date field used for auditing user/master changes.",
    "MODDATE": "Modification date timestamp used for recency and change-window selection.",
    "MODIFIER": "User/agent who performed the change used for accountability filtering.",
    "MODTI": "Modification time field used for precise temporal change analysis.",
    "MODTIME": "Modification time used alongside date to define exact change window.",
    "MPROK": "Material/procurement status key used to identify control-relevant status states.",
    "MPROK_DESC": "Description of material/procurement status for readable reporting.",
    "MSCDATE": "Message/status creation date used for log/event period filtering.",
    "MSCTIME": "Message/status creation time used for intra-day event analysis.",
    "MSGID": "Message class ID used to group and filter technical/application messages.",
    "NAME_FIRST": "Business partner/user first name used for readable identity output.",
    "NAME_LAST": "Business partner/user last name used for readable identity output.",
    "NETWR": "Net value amount used for commercial threshold and anomaly checks.",
    "NO_DATE_RESTRICTION": "Flag that disables default date-window filtering when set.",
    "NRPRO": "Number profile/range profile key used in numbering-control monitoring.",
    "PEINH": "Price unit denominator used to interpret per-unit purchasing prices.",
    "PERIODIC": "Periodic-job indicator used to separate recurring from one-time background jobs.",
    "POSNR": "Document item number used for line-level drilldown and joins.",
    "PRIVSUM": "Private-memory usage aggregate used for technical performance monitoring.",
    "PROGNAME": "ABAP report/program name used to scope technical execution records.",
    "PS_PSP_PNR": "WBS element key used for project-system linked cost/procurement monitoring.",
    "PWDLGNDATE": "Date of last password logon usage used for credential-age/security checks.",
    "RESWK": "Supplying/receiving plant key used in cross-plant logistics analysis.",
    "RESWK_DESC": "Plant description text used to enrich plant-level reporting.",
    "RFCDEST": "RFC destination key used to scope remote connectivity/technical checks.",
    "SDLSTRTDT": "Scheduled job start date used for scheduler planning compliance checks.",
    "SDLSTRTTM": "Scheduled job start time used for scheduler timing precision.",
    "SDLUNAME": "Scheduler user name used to analyze job ownership and change responsibility.",
    "SGTXT": "Document line text used for context and free-text pattern filters.",
    "SHKZG": "Debit/Credit indicator used to separate accounting posting direction.",
    "SLGDATE": "Application log date used for BAL log time-window filtering.",
    "SLGDATTIM": "Combined date-time in application log used for exact event sequencing.",
    "SLGLTRM": "Application log object context term used for focused BAL filtering.",
    "SLGMAND": "Client field in application log context used for tenant scoping.",
    "SLGMODE": "Application log mode indicator used to distinguish processing mode categories.",
    "SLGPROC": "Application log process identifier used to group technical process runs.",
    "SLGREPNA": "Application log report name used for report-level BAL filtering.",
    "SLGTC": "Transaction code captured in application log for process-context analysis.",
    "SLGTIME": "Application log time used for intra-day event selection.",
    "SLGTYPE": "Application log message/type category used for severity/process segmentation.",
    "SLGUSER": "Application log user used for accountability and actor-based filtering.",
    "SOBKZ": "Special stock indicator used to distinguish stock ownership categories.",
    "SOLDTO_DESC": "Sold-to party description text used for readable customer reporting.",
    "SPART": "Division key used for SD product-line segmentation.",
    "SPRAS": "Language key used for language-dependent text retrieval and filtering.",
    "STATE_DESC": "Human-readable state description used for alert/report interpretation.",
    "STATU": "Status code used for state-based filtering in process monitoring.",
    "STATU_DESC": "Status description text used for readable status analytics.",
    "SUBID": "Subsystem/sub-identifier key used to isolate technical source partitions.",
    "TAB": "Table alias/name selector used for technical table-focused filtering.",
    "TABKEY": "Composite table key value used in change-document record identification.",
    "TABNAME": "Database table name used to scope change/object monitoring to specific tables.",
    "TEXT": "General text payload field used for message/contextual filtering.",
    "TEXT_CASE": "Text case/normalization selector used for case-sensitive text filtering behavior.",
    "TRDAT": "Last logon date (or technical date marker) used for user activity recency checks.",
    "TYPE": "Type/category indicator used to segment records by business/technical class.",
    "TZONE": "Time zone key used to interpret and normalize timestamp fields.",
    "UDATE": "Update/change date used for technical recency and change-window filtering.",
    "UFLAG": "User lock/status flag used to identify locked/disabled user states.",
    "USER": "User identifier field used for actor-based filtering.",
    "USERID": "User ID key used for authentication/user master level scoping.",
    "USERNAME": "User name display field used for readable identity reporting.",
    "USNAM": "SAP changed-by/created-by user field used for accountability filtering.",
    "USTYP": "User type category used to segment dialog/system/service users.",
    "UTIME": "Update/change time used with UDATE for precise event windows.",
    "VALID_USERS_ONLY": "Boolean flag to restrict results to users validated as active/allowed.",
    "VALUE_NEW": "New value in change documents used for after-change analysis.",
    "VALUE_OLD": "Old value in change documents used for before/after comparison.",
    "VBELN": "SD document number used as primary key for sales/billing/delivery documents.",
    "VBTYP": "SD document category used to segment SD document classes.",
    "VBUND": "Trading partner/company field used for intercompany transaction analysis.",
    "VDATU": "Requested/validity date used for schedule and due-date based filtering.",
    "VENDOR_DESC": "Vendor description text used for readable supplier-level reporting.",
    "VERSN": "Version field used to separate records by versioned configuration/data state.",
    "VGABE": "Transaction/event type in purchasing history used to classify movement category.",
    "VKBUR": "Sales office key used for organizational SD segmentation.",
    "VKGRP": "Sales group key used for team-level SD analytics.",
    "VKORG": "Sales organization key used for legal/commercial SD scoping.",
    "VTWEG": "Distribution channel used for SD market/channel segmentation.",
    "WAERK": "Document currency key used for value analysis in transaction currency.",
    "WAERS": "Currency key used for monetary field interpretation and filtering.",
    "WAERS_FR": "Source/from currency key used in currency-change/translation contexts.",
    "WAS_PLANND": "Planned-state indicator used to distinguish planned versus actual execution records.",
    "WAVWR": "Statistical value amount field used for value-based exception thresholds.",
    "WERKS": "Plant key used to scope logistics/procurement records by site.",
    "WGBEZ": "Material group description used for readable category reporting.",
    "WP_TYPE": "Work process type/category used for SAP basis workload monitoring segmentation.",
    "WRBTR": "Amount in document currency used for FI/MM value-based controls.",
    "XBLNR": "Reference document number used for external document matching and traceability.",
    "XCPDK": "One-time account indicator used to identify one-time customer/vendor postings.",
}


def auto_explain(param: str) -> str:
    p = param.upper().strip()
    if p in OVERRIDES:
        return OVERRIDES[p]

    # Partner slot patterns (BP1/BP2/BP3 groups)
    m = re.match(r"BP([123])_(CODE|FUNCT|NAME)$", p)
    if m:
        slot, kind = m.groups()
        if kind == "CODE":
            return f"Business partner slot {slot} code used to identify the linked partner in multi-partner records."
        if kind == "FUNCT":
            return f"Business partner slot {slot} function/role used to classify partner responsibility."
        return f"Business partner slot {slot} name/description used for readable partner output."

    if p.endswith("_DESC"):
        base = p[: -5]
        return f"Description text for {base}, used to provide business-readable output beside technical keys."
    if p.endswith("_OLD"):
        base = p[: -4]
        return f"Previous value of {base} captured for before/after change analysis."
    if p.endswith("_NEW"):
        base = p[: -4]
        return f"New value of {base} captured for after-change impact analysis."
    if p.endswith("DATE") or p.endswith("DAT"):
        return f"{p} is a date field used to constrain the monitoring window or document recency scope."
    if p.endswith("TIME") or p.endswith("TIM"):
        return f"{p} is a time field used to refine intra-day event selection and sequencing."
    if p.endswith("ICON"):
        return f"{p} is a visual status icon field used to present state/severity in the output."
    if p.endswith("CNT") or p.startswith("CNT_"):
        return f"{p} is a count/volume metric used for threshold-based exception evaluation."
    if p.endswith("AMOUNT") or p.endswith("BTR") or p.endswith("WR"):
        return f"{p} is a monetary value field used for value-based filtering and anomaly thresholds."
    if p.startswith("Z"):
        return f"{p} is a custom field in customer namespace used for implementation-specific filtering/reporting."

    return f"{p} is an SAP application field used as a selection dimension to narrow monitoring results to the relevant record subset."


def main() -> None:
    params = [x.strip().upper() for x in SRC_CSV.read_text(encoding="utf-8").splitlines() if x.strip()]
    # keep order from file, but unique by first appearance
    seen = set()
    ordered = []
    for p in params:
        if p not in seen:
            seen.add(p)
            ordered.append(p)

    attached = _load_sap_canonical_attached()
    unified = _load_sap_unified_explanations()

    wb = Workbook()
    ws = wb.active
    ws.title = "dictionary"

    ws["A1"] = "parameter"
    ws["B1"] = "suggested/corrected explanation"
    ws["C1"] = "SAP canonical explanation"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["C1"].font = Font(bold=True)

    r = 2
    for p in ordered:
        ws.cell(r, 1, p)
        ws.cell(r, 2, auto_explain(p))
        ws.cell(r, 3, attached.get(p) or unified.get(p) or auto_explain(p))
        r += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 130
    ws.column_dimensions["C"].width = 130

    wb.save(OUT_XLSX)
    print(f"Parameters read: {len(params)}")
    print(f"Unique parameters written: {len(ordered)}")
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()

