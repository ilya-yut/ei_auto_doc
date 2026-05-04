"""
Build a new parameter dictionary workbook from params_used_twice_not_in_dictionary.csv.

- Explanations prefer tools.sap_unified_param_texts (same curated style as the reference approach).
- Remaining rows use sap_twice_csv_overlay.CANONICAL + NON_SAP_STANDARD for yellow highlighting.
- Preserves per-part usage counts from the CSV for traceability.
"""

from __future__ import annotations

import csv
import importlib.util
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(r"c:\vibe code dev\ei_auto_doc")
TEMP = ROOT / "ei docs for analysis" / "EI docs" / "temp"

SRC_CSV = TEMP / "params_used_twice_not_in_dictionary.csv"
OUT_XLSX = TEMP / "params_dictionary_from_twice_csv.xlsx"

YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")


def _load_unified() -> dict[str, str]:
    spec = importlib.util.spec_from_file_location(
        "sap_unified_param_texts", ROOT / "tools" / "sap_unified_param_texts.py"
    )
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return {str(k).upper(): str(v).strip() for k, v in mod.SAP_UNIFIED_EXPLANATION.items()}


def _load_overlay():
    spec = importlib.util.spec_from_file_location("sap_twice_csv_overlay", TEMP / "sap_twice_csv_overlay.py")
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return mod


def main() -> None:
    unified = _load_unified()
    overlay = _load_overlay()

    rows_in: list[dict[str, str]] = []
    with SRC_CSV.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows_in.append(row)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "dictionary"

    headers = [
        "parameter",
        "count_part1",
        "count_part2",
        "count_part3",
        "count_part4",
        "total_count",
        "sap_standard",
        "SAP canonical explanation",
    ]
    ws.append(headers)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)

    for r, src in enumerate(rows_in, start=2):
        p = (src.get("parameter") or "").strip().upper()
        expl = unified.get(p)
        if expl:
            sap_ok = True
        else:
            expl = overlay.CANONICAL.get(p, "")
            sap_ok = p not in overlay.NON_SAP_STANDARD

        ws.cell(r, 1, p)
        for i, col in enumerate(
            ("count_part1", "count_part2", "count_part3", "count_part4", "total_count"), start=2
        ):
            raw = src.get(col, "")
            try:
                ws.cell(r, i, int(raw) if str(raw).strip() != "" else "")
            except ValueError:
                ws.cell(r, i, raw)
        ws.cell(r, 7, "Yes" if sap_ok else "No")
        ws.cell(r, 8, expl)
        ws.cell(r, 8).alignment = Alignment(wrap_text=True, vertical="top")

        if not sap_ok:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = YELLOW

    ws.column_dimensions["A"].width = 22
    for c in range(2, 7):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 96
    ws.freeze_panes = "A2"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({len(rows_in)} parameters)")


if __name__ == "__main__":
    main()
