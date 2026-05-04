from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


ROOT = Path(r"c:\vibe code dev\ei_auto_doc\ei docs for analysis\EI docs")
TEMP = ROOT / "temp"

SOURCE_DICTIONARY = TEMP / "params_dictionary.xlsx"
OUT_XLSX = TEMP / "params_ge3_not_in_dictionary.xlsx"

DOCX_GLOBS = ["Part 1/*.docx", "Part 2/*.docx", "Part 3/*.docx", "Part 4/*.docx"]

RE_NUM = re.compile(r"^\d+$")
RE_PARAM = re.compile(r"^[A-Z][A-Z0-9_]{1,}$")
PARAM_HEADER_CANDIDATES = ("field", "parameter", "field name", "name")


def iter_block_items(parent: DocumentType):
    body = parent.element.body
    for child in body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


def read_dictionary_params(path: Path) -> set[str]:
    wb = load_workbook(path, read_only=True)
    ws = wb["dictionary"] if "dictionary" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: set[str] = set()
    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, 1).value
        if not p:
            continue
        s = str(p).strip().upper()
        if s:
            out.add(s)
    return out


def extract_params_from_docx(path: Path) -> set[str]:
    doc = Document(str(path))
    saw_param_ref_heading = False
    fallback_tables: list[list[list[str]]] = []
    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            t = block.text.strip().lower()
            if "parameters reference table" in t:
                saw_param_ref_heading = True
            continue
        if not isinstance(block, Table):
            continue

        rows = [[c.text.strip() for c in row.cells] for row in block.rows]
        if not rows:
            continue

        header = [x.lower() for x in rows[0]]
        param_col = None
        for h in PARAM_HEADER_CANDIDATES:
            if h in header:
                param_col = header.index(h)
                break
        if param_col is None:
            continue
        # skip technical mapping tables (Structure Name / Field Name ...)
        if "structure name" in header:
            continue
        if not saw_param_ref_heading:
            # keep as fallback candidate if explicit heading was not seen
            fallback_tables.append(rows)
            continue

        num_col = None
        if "#" in rows[0]:
            num_col = rows[0].index("#")
        elif "no." in header:
            num_col = header.index("no.")
        elif "no" in header:
            num_col = header.index("no")

        out: set[str] = set()
        for rr in rows[1:]:
            if param_col >= len(rr):
                continue
            p = rr[param_col].strip().upper()
            if not p or p in {"FIELD", "---"}:
                continue
            if not RE_PARAM.match(p):
                continue
            if num_col is not None and num_col < len(rr):
                n = rr[num_col].strip()
                if n and not RE_NUM.match(n):
                    break
            out.add(p)
        if out:
            return out

    # Fallback: parse first table with Field header when explicit heading not detected.
    for rows in fallback_tables:
        header = [x.lower() for x in rows[0]]
        param_col = None
        for h in PARAM_HEADER_CANDIDATES:
            if h in header:
                param_col = header.index(h)
                break
        if param_col is None:
            continue
        if "structure name" in header:
            continue
        num_col = None
        if "#" in rows[0]:
            num_col = rows[0].index("#")
        elif "no." in header:
            num_col = header.index("no.")
        elif "no" in header:
            num_col = header.index("no")

        out: set[str] = set()
        for rr in rows[1:]:
            if param_col >= len(rr):
                continue
            p = rr[param_col].strip().upper()
            if not p or p in {"FIELD", "---"}:
                continue
            if not RE_PARAM.match(p):
                continue
            if num_col is not None and num_col < len(rr):
                n = rr[num_col].strip()
                if n and not RE_NUM.match(n):
                    break
            out.add(p)
        if out:
            return out
    return set()


def main() -> None:
    if not SOURCE_DICTIONARY.exists():
        raise FileNotFoundError(f"Dictionary file not found: {SOURCE_DICTIONARY}")

    dict_params = read_dictionary_params(SOURCE_DICTIONARY)

    all_docx: list[Path] = []
    for g in DOCX_GLOBS:
        all_docx.extend(sorted(ROOT.glob(g)))

    param_files: dict[str, set[Path]] = defaultdict(set)
    skipped: list[str] = []
    for fp in all_docx:
        params = extract_params_from_docx(fp)
        if not params:
            skipped.append(str(fp.relative_to(ROOT)))
            continue
        for p in params:
            param_files[p].add(fp)

    usage = {p: len(fs) for p, fs in param_files.items()}
    result_params = sorted(
        [p for p, c in usage.items() if c >= 3 and p not in dict_params],
        key=lambda p: (-usage[p], p),
    )
    dist = Counter(usage[p] for p in result_params)

    wb = Workbook()
    ws = wb.active
    ws.title = "params_ge3_missing"

    ws["A1"] = "Summary"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "DOCX scanned"
    ws["B2"] = len(all_docx)
    ws["A3"] = "Distinct params found"
    ws["B3"] = len(param_files)
    ws["A4"] = "Params in dictionary"
    ws["B4"] = len(dict_params)
    ws["A5"] = "Result params (>=3 and not in dictionary)"
    ws["B5"] = len(result_params)
    ws["A6"] = "Skipped files (no params table parsed)"
    ws["B6"] = len(skipped)

    ws["A8"] = "usage_count"
    ws["B8"] = "parameter_count"
    ws["A8"].font = Font(bold=True)
    ws["B8"].font = Font(bold=True)
    r = 9
    for n in sorted(dist):
        ws.cell(r, 1, n)
        ws.cell(r, 2, dist[n])
        r += 1

    r += 1
    ws.cell(r, 1, "parameter").font = Font(bold=True)
    ws.cell(r, 2, "usage_count").font = Font(bold=True)
    ws.cell(r, 3, "sample_subfolder").font = Font(bold=True)
    r += 1
    for p in result_params:
        ws.cell(r, 1, p)
        ws.cell(r, 2, usage[p])
        # sample one subfolder for quick orientation
        first = sorted(param_files[p])[0]
        ws.cell(r, 3, first.relative_to(ROOT).parts[0] if len(first.relative_to(ROOT).parts) > 1 else ".")
        r += 1

    # optional detail sheet with files per parameter
    ws2 = wb.create_sheet("files")
    ws2.append(["parameter", "subfolder", "filename"])
    for c in range(1, 4):
        ws2.cell(1, c).font = Font(bold=True)
    rr = 2
    for p in result_params:
        for fp in sorted(param_files[p]):
            rel = fp.relative_to(ROOT)
            ws2.cell(rr, 1, p)
            ws2.cell(rr, 2, rel.parts[0] if len(rel.parts) > 1 else ".")
            ws2.cell(rr, 3, fp.name)
            rr += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 72

    wb.save(OUT_XLSX)

    print(f"DOCX scanned: {len(all_docx)}")
    print(f"Distinct params found: {len(param_files)}")
    print(f"Params in dictionary: {len(dict_params)}")
    print(f"Result params: {len(result_params)}")
    print(f"Skipped files: {len(skipped)}")
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()

