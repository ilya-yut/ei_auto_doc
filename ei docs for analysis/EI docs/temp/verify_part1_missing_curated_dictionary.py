from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


TEMP = Path(r"c:\vibe code dev\ei_auto_doc\ei docs for analysis\EI docs\temp")
SRC_REQ = TEMP / "params_dictionary_part1_missing.xlsx"
CURATED = TEMP / "params_dictionary_part1_missing_curated.xlsx"
OUT_VERIFY = TEMP / "params_dictionary_part1_missing_curated_verification.xlsx"

BANNED_PATTERNS = [
    "technical selection parameter",
    "used in this monitor set",
    "align interpretation with ddic",
]


def _read_requested() -> list[str]:
    wb = load_workbook(SRC_REQ, read_only=True)
    ws = wb["missing_dictionary"] if "missing_dictionary" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: list[str] = []
    for r in range(9, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not v:
            continue
        p = str(v).strip().upper()
        if p:
            out.append(p)
    # unique keep order
    seen = set()
    uniq: list[str] = []
    for p in out:
        if p not in seen:
            seen.add(p)
            uniq.append(p)
    return uniq


def _read_curated() -> list[tuple[str, str, str, str]]:
    wb = load_workbook(CURATED, read_only=True)
    ws = wb["curated_dictionary"] if "curated_dictionary" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: list[tuple[str, str, str, str]] = []
    for r in range(9, ws.max_row + 1):
        p = ws.cell(r, 1).value
        e = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        n = ws.cell(r, 4).value
        if not p:
            continue
        out.append(
            (
                str(p).strip().upper(),
                str(e or "").strip(),
                str(c or "").strip().lower(),
                str(n or "").strip(),
            )
        )
    return out


def main() -> None:
    req = _read_requested()
    rows = _read_curated()

    req_set = set(req)
    row_params = [p for p, _e, _c, _n in rows]
    row_set = set(row_params)

    missing = sorted(req_set - row_set)
    unexpected = sorted(row_set - req_set)
    dupes = [p for p, n in Counter(row_params).items() if n > 1]

    bad_rows: list[tuple[str, str, str]] = []
    weak_rows: list[tuple[str, str, str]] = []
    length_rows: list[tuple[str, str, int]] = []

    for p, e, c, _n in rows:
        low = e.lower()
        if any(bp in low for bp in BANNED_PATTERNS):
            bad_rows.append((p, c, e))
        if "not used" in low or "unused" in low:
            weak_rows.append((p, c, e))
        wc = len(e.split())
        if wc < 6 or wc > 28:
            length_rows.append((p, c, wc))

    conf = Counter(c for _p, _e, c, _n in rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    ws["A1"] = "Verification summary"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Requested parameters"
    ws["B2"] = len(req_set)
    ws["A3"] = "Curated rows"
    ws["B3"] = len(rows)
    ws["A4"] = "Unique curated parameters"
    ws["B4"] = len(row_set)
    ws["A5"] = "Missing parameters"
    ws["B5"] = len(missing)
    ws["A6"] = "Unexpected parameters"
    ws["B6"] = len(unexpected)
    ws["A7"] = "Duplicate parameter rows"
    ws["B7"] = len(dupes)
    ws["A8"] = "Rows with banned generic phrases"
    ws["B8"] = len(bad_rows)
    ws["A9"] = "Rows with not-used/unused wording"
    ws["B9"] = len(weak_rows)
    ws["A10"] = "Rows outside word-count target (6..28)"
    ws["B10"] = len(length_rows)
    ws["A12"] = "Confidence high"
    ws["B12"] = conf.get("high", 0)
    ws["A13"] = "Confidence medium"
    ws["B13"] = conf.get("medium", 0)
    ws["A14"] = "Confidence low"
    ws["B14"] = conf.get("low", 0)

    # sheet: issues
    ws2 = wb.create_sheet("issues")
    ws2.append(["issue_type", "parameter", "confidence", "details"])
    for c in range(1, 5):
        ws2.cell(1, c).font = Font(bold=True)
    for p in missing:
        ws2.append(["missing_parameter", p, "", "Absent from curated dictionary"])
    for p in unexpected:
        ws2.append(["unexpected_parameter", p, "", "Not requested in source missing list"])
    for p in dupes:
        ws2.append(["duplicate_parameter", p, "", "More than one curated row"])
    for p, c, e in bad_rows:
        ws2.append(["banned_phrase", p, c, e])
    for p, c, e in weak_rows:
        ws2.append(["weak_not_used", p, c, e])
    for p, c, w in length_rows:
        ws2.append(["length_out_of_range", p, c, f"word_count={w}"])

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 12
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 120

    wb.save(OUT_VERIFY)

    print(f"Requested parameters: {len(req_set)}")
    print(f"Curated rows: {len(rows)}")
    print(f"Missing parameters: {len(missing)}")
    print(f"Unexpected parameters: {len(unexpected)}")
    print(f"Duplicate parameter rows: {len(dupes)}")
    print(f"Rows with banned generic phrases: {len(bad_rows)}")
    print(f"Rows with not-used/unused wording: {len(weak_rows)}")
    print(f"Rows outside word-count target: {len(length_rows)}")
    print(f"Confidence high/medium/low: {conf.get('high',0)}/{conf.get('medium',0)}/{conf.get('low',0)}")
    print(f"Wrote {OUT_VERIFY}")


if __name__ == "__main__":
    main()

