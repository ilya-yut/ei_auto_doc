from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Font


ROOT = Path(r"c:\vibe code dev\ei_auto_doc\ei docs for analysis\EI docs")
TEMP = ROOT / "temp"

# Exclusion source (25-parameter list from previous step)
EXCLUDE_MD = (
    Path(r"c:\vibe code dev\ei_auto_doc")
    / "ei docs for analysis"
    / "Part 1 conv"
    / "SHARED_PARAMETER_EXPLANATIONS_GE3_NO_BACKDAYS_USER_FLD.md"
)

OUT_XLSX = TEMP / "params_used_ge4_excluding_25.xlsx"

DOCX_GLOB = ["Part 1/*.docx", "Part 2/*.docx", "Part 3/*.docx", "Part 4/*.docx"]

RE_NUM = re.compile(r"^\d+$")
RE_PARAM_TOKEN = re.compile(r"^[A-Z0-9_]{2,}$")


def iter_block_items(parent: DocumentType):
    body = parent.element.body
    for child in body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


def _extract_params_from_rows(rows: list[list[str]]) -> set[str]:
    out: set[str] = set()
    if not rows:
        return out

    header_idx = None
    field_col = None
    num_col = None
    for i, r in enumerate(rows):
        low = [x.lower() for x in r]
        has_field = any("field" == x or "field name" == x or "parameter" == x for x in low)
        has_desc = any("description" in x for x in low)
        if has_field and has_desc:
            header_idx = i
            if "field" in low:
                field_col = low.index("field")
            elif "field name" in low:
                field_col = low.index("field name")
            else:
                field_col = low.index("parameter")
            if "#" in r:
                num_col = r.index("#")
            elif "no." in low:
                num_col = low.index("no.")
            elif "no" in low:
                num_col = low.index("no")
            break
    if header_idx is None or field_col is None:
        return out

    for r in rows[header_idx + 1 :]:
        if field_col >= len(r):
            continue
        field = r[field_col].strip()
        if not field:
            continue
        if field.lower() in {"field", "---"}:
            continue
        field_norm = field.upper().strip()
        if not RE_PARAM_TOKEN.match(field_norm):
            continue
        if num_col is not None and num_col < len(r):
            n = r[num_col].strip()
            if n and not RE_NUM.match(n):
                break
        out.add(field_norm)
    return out


def read_excluded_params(path: Path) -> set[str]:
    txt = path.read_text(encoding="utf-8")
    excluded: set[str] = set()
    for line in txt.splitlines():
        s = line.strip()
        if not s.startswith("|"):
            continue
        parts = [x.strip() for x in s.strip("|").split("|")]
        if len(parts) < 4:
            continue
        if parts[0].lower() in {"parameter", "-----------"}:
            continue
        if parts[0].startswith("---"):
            continue
        excluded.add(parts[0])
    return excluded


def _extract_params_from_docx(path: Path) -> set[str]:
    doc = Document(str(path))
    saw_param_ref_heading = False
    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            t = block.text.strip().lower()
            if "parameters reference table" in t:
                saw_param_ref_heading = True
            continue
        if not isinstance(block, Table):
            continue
        if not saw_param_ref_heading:
            continue

        rows: list[list[str]] = []
        for row in block.rows:
            rows.append([c.text.strip() for c in row.cells])
        out = _extract_params_from_rows(rows)
        if out:
            return out
    return set()


def main() -> None:
    TEMP.mkdir(parents=True, exist_ok=True)
    excluded = read_excluded_params(EXCLUDE_MD)

    all_docx: list[Path] = []
    for g in DOCX_GLOB:
        all_docx.extend(sorted(ROOT.glob(g)))

    param_to_files: dict[str, set[Path]] = defaultdict(set)
    skipped: list[str] = []
    for p in all_docx:
        params = _extract_params_from_docx(p)
        if not params:
            skipped.append(str(p.relative_to(ROOT)))
            continue
        for prm in params:
            param_to_files[prm].add(p)

    usage = {k: len(v) for k, v in param_to_files.items()}
    filtered_params = sorted(
        [k for k, c in usage.items() if c >= 4 and k not in excluded],
        key=lambda k: (-usage[k], k),
    )

    dist = Counter(usage[k] for k in filtered_params)
    detail_rows: list[tuple[str, str, str]] = []
    for prm in filtered_params:
        for fp in sorted(param_to_files[prm]):
            rel = fp.relative_to(ROOT)
            sub = rel.parts[0] if len(rel.parts) > 1 else "."
            detail_rows.append((prm, sub, fp.name))

    wb = Workbook()
    ws = wb.active
    ws.title = "params_ge4"

    r = 1
    ws.cell(r, 1, "Summary")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1
    ws.cell(r, 1, "DOCX scanned")
    ws.cell(r, 2, len(all_docx))
    r += 1
    ws.cell(r, 1, "Distinct parameters found")
    ws.cell(r, 2, len(param_to_files))
    r += 1
    ws.cell(r, 1, "Excluded list size")
    ws.cell(r, 2, len(excluded))
    r += 1
    ws.cell(r, 1, "Parameters in result (>=4, excluding list)")
    ws.cell(r, 2, len(filtered_params))
    r += 2

    ws.cell(r, 1, "usage_count")
    ws.cell(r, 2, "parameter_count")
    ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2).font = Font(bold=True)
    r += 1
    for n in sorted(dist):
        ws.cell(r, 1, n)
        ws.cell(r, 2, dist[n])
        r += 1
    r += 2

    ws.cell(r, 1, "parameter")
    ws.cell(r, 2, "sub-folder")
    ws.cell(r, 3, "filename")
    for c in range(1, 4):
        ws.cell(r, c).font = Font(bold=True)
    r += 1

    for prm, sub, fname in detail_rows:
        ws.cell(r, 1, prm)
        ws.cell(r, 2, sub)
        ws.cell(r, 3, fname)
        r += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 72

    wb.save(OUT_XLSX)

    print(f"Scanned DOCX: {len(all_docx)}")
    print(f"Distinct params: {len(param_to_files)}")
    print(f"Excluded params: {len(excluded)}")
    print(f"Result params (>=4, excluded list removed): {len(filtered_params)}")
    print(f"Result rows: {len(detail_rows)}")
    if skipped:
        print(f"Skipped (no params table parsed): {len(skipped)}")
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
