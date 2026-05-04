from __future__ import annotations

import importlib.util
import re
from collections import Counter
from pathlib import Path

from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Font


ROOT = Path(r"c:\vibe code dev\ei_auto_doc")
DOCS_ROOT = ROOT / "ei docs for analysis" / "EI docs"
TEMP = DOCS_ROOT / "temp"

OUT_XLSX = TEMP / "params_attached_curated.xlsx"

PARAM_TEXT = """
ERDAT
CHANGENR
TAB_DESC
ACT_CHNGNO
CHANGE_IND
CHANGE_IND_DESC
CHNGIND
CHNGIND_DESC
OBJECTCLAS
OBJECT_DESC
PLANCHNGNR
UNIT_NEW
UNIT_OLD
MESSAGE
OBJECT
AUFNR
ACTION_DESC
AGR_FDATE
AGR_TDATE
ARKTX
AS4DATE
AS4TIME
ATTRBT
BADAT
BANKN
BLART
BPMNG
BPRME
BUZEI
CHARG
COMP_CODE_DESC
COMP_OPERATOR
COND
DMBTR_FR
ENDDATE
ENDTIME
EREKZ
ERFME
ERFMG
ERRNO
ETENR
FABKL
FACDATE
FKSTO
FKTYP
FRGKZ
GRUND
HWAER
INCLUDENAME
KDATB
KDATE
KDEIN
KLMENG
KONZS
KOSTL
KSCHL
KTOPL
KWMENG
LBKUM
MBLNR
MJAHR
MODBE_NAME_FIRST
MODBE_NAME_LAST
MSGNO
MSGTY
MTART
NEW_VAL
OBJNR
OLD_VAL
PARVW
PAYER_DESC
PLPLA
PRCTR
PROCSTAT
PSTYV
QNAME
QRFCFNAM
QRFCUSER
REF_FIELD_NAME1
REF_FIELD_NAME2
REQ_CNT
RESULT_COMP
RFBSK
SAKTO
SALK3
SPERR
STAT
STKZN
SUBOBJECT
SUBSYSTEM
SYHOST
SYUSER
TABE
TID
TIME_REF_FLD
TRKORR
TRN_BY
TRN_EX
TRSTATUS
TXT
UNAME
USTYP_DESC
VARIANT
VGART
VPRSV
VRKME
WEPOS
WERKS_DESC
WORKING_DAYS
W_VARIANT
ZEILE
ZEIT
ZTERM
"""

RE_NUM = re.compile(r"^\d+$")
RE_PARAM = re.compile(r"^[A-Z][A-Z0-9_]{1,}$")
RE_WORD = re.compile(r"[A-Za-z0-9_]+")
PARAM_HEADER_CANDIDATES = ("field", "parameter", "field name", "name")


def get_params() -> list[str]:
    seen = set()
    ordered: list[str] = []
    for line in PARAM_TEXT.strip().splitlines():
        p = line.strip().upper()
        if p and p not in seen:
            seen.add(p)
            ordered.append(p)
    return ordered


def load_sap_canonical_attached() -> dict[str, str]:
    path = TEMP / "sap_canonical_attached_params.py"
    spec = importlib.util.spec_from_file_location("sap_canonical_attached_params", path)
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return {k.upper(): v.strip() for k, v in mod.SAP_CANONICAL_ATTACHED.items()}


def iter_block_items(parent: DocumentType):
    body = parent.element.body
    for child in body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


def extract_evidence(path: Path, targets: set[str]) -> dict[str, dict[str, Counter[str] | int]]:
    doc = Document(str(path))
    out: dict[str, dict[str, Counter[str] | int]] = {}
    in_abap = False
    saw_param_ref_heading = False

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            t = " ".join((block.text or "").split())
            low = t.lower()
            if low == "parameters reference table":
                saw_param_ref_heading = True
                continue
            if low == "abap code":
                in_abap = True
                continue
            if in_abap:
                words = {w.upper() for w in RE_WORD.findall(t)}
                for p in targets:
                    if p in words:
                        out.setdefault(
                            p,
                            {"desc": Counter(), "de": Counter(), "dom": Counter(), "typ": Counter(), "abap_hits": 0},
                        )
                        out[p]["abap_hits"] = int(out[p]["abap_hits"]) + 1  # type: ignore[index]
                if "ENDFUNCTION" in t.upper():
                    in_abap = False
            continue

        if not isinstance(block, Table):
            continue
        rows = [[" ".join((c.text or "").split()) for c in row.cells] for row in block.rows]
        if not rows:
            continue
        header = [h.lower() for h in rows[0]]

        param_col = None
        for h in PARAM_HEADER_CANDIDATES:
            if h in header:
                param_col = header.index(h)
                break
        if param_col is None:
            continue
        if "structure name" in header:
            continue
        if not saw_param_ref_heading:
            continue

        col_desc = header.index("description") if "description" in header else None
        col_type = header.index("type") if "type" in header else None
        col_de = header.index("data element") if "data element" in header else None
        col_dom = header.index("domain") if "domain" in header else None
        col_num = None
        if "#" in rows[0]:
            col_num = rows[0].index("#")
        elif "no." in header:
            col_num = header.index("no.")
        elif "no" in header:
            col_num = header.index("no")

        for rr in rows[1:]:
            if param_col >= len(rr):
                continue
            p = rr[param_col].strip().upper()
            if p not in targets or not RE_PARAM.match(p):
                continue
            if col_num is not None and col_num < len(rr):
                n = rr[col_num].strip()
                if n and not RE_NUM.match(n):
                    break
            out.setdefault(
                p,
                {"desc": Counter(), "de": Counter(), "dom": Counter(), "typ": Counter(), "abap_hits": 0},
            )
            if col_desc is not None and col_desc < len(rr) and rr[col_desc]:
                out[p]["desc"][rr[col_desc].strip()] += 1  # type: ignore[index]
            if col_de is not None and col_de < len(rr) and rr[col_de]:
                out[p]["de"][rr[col_de].strip()] += 1  # type: ignore[index]
            if col_dom is not None and col_dom < len(rr) and rr[col_dom]:
                out[p]["dom"][rr[col_dom].strip()] += 1  # type: ignore[index]
            if col_type is not None and col_type < len(rr) and rr[col_type]:
                out[p]["typ"][rr[col_type].strip()] += 1  # type: ignore[index]

    return out


def main() -> None:
    params = get_params()
    param_set = set(params)
    canonical = load_sap_canonical_attached()
    missing = [p for p in params if p not in canonical]
    extra = sorted(set(canonical.keys()) - param_set)
    if missing or extra:
        raise SystemExit(f"SAP canonical dictionary out of sync. Missing={missing!r} Extra={extra!r}")

    files = []
    for part in ("Part 1", "Part 2", "Part 3", "Part 4"):
        files.extend(sorted((DOCS_ROOT / part).glob("*.docx")))

    files_seen: dict[str, set[str]] = {p: set() for p in params}
    desc = {p: Counter() for p in params}
    de = {p: Counter() for p in params}
    dom = {p: Counter() for p in params}
    typ = {p: Counter() for p in params}
    abap_hits = {p: 0 for p in params}

    for fp in files:
        ev = extract_evidence(fp, param_set)
        for p in ev:
            files_seen[p].add(fp.name)
            desc[p].update(ev[p]["desc"])  # type: ignore[arg-type]
            de[p].update(ev[p]["de"])  # type: ignore[arg-type]
            dom[p].update(ev[p]["dom"])  # type: ignore[arg-type]
            typ[p].update(ev[p]["typ"])  # type: ignore[arg-type]
            abap_hits[p] += int(ev[p]["abap_hits"])  # type: ignore[arg-type]

    rows: list[tuple[str, str, str, str]] = []
    for p in params:
        top_desc = desc[p].most_common(1)[0][0] if desc[p] else ""
        top_de = de[p].most_common(1)[0][0] if de[p] else ""
        top_dom = dom[p].most_common(1)[0][0] if dom[p] else ""

        expl = canonical[p]
        origin = "sap_canonical"
        confidence = "high"

        note = (
            f"origin={origin}; files={len(files_seen[p])}; abap_hits={abap_hits[p]}; "
            f"desc='{top_desc}'; de='{top_de}'; dom='{top_dom}'"
        )
        rows.append((p, expl, confidence, note))

    wb = Workbook()
    ws = wb.active
    ws.title = "curated_dictionary"

    ws["A1"] = "Summary"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Parameters requested"
    ws["B2"] = len(params)
    ws["A3"] = "DOCX scanned"
    ws["B3"] = len(files)
    ws["A4"] = "Generated entries"
    ws["B4"] = len(rows)
    ws["A5"] = "High confidence"
    ws["B5"] = sum(1 for _, _, c, _ in rows if c == "high")
    ws["A6"] = "Medium confidence"
    ws["B6"] = sum(1 for _, _, c, _ in rows if c == "medium")
    ws["A7"] = "Low confidence"
    ws["B7"] = sum(1 for _, _, c, _ in rows if c == "low")

    header_row = 9
    headers = ["parameter", "curated_explanation", "confidence", "evidence_notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(header_row, i, h).font = Font(bold=True)

    r = header_row + 1
    for row in rows:
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
        r += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 120

    wb.save(OUT_XLSX)
    print(f"Parameters requested: {len(params)}")
    print(f"Generated entries: {len(rows)}")
    print(f"High/Medium/Low: {sum(1 for _, _, c, _ in rows if c == 'high')}/{sum(1 for _, _, c, _ in rows if c == 'medium')}/{sum(1 for _, _, c, _ in rows if c == 'low')}")
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
