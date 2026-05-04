from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Font


ROOT = Path(r"c:\vibe code dev\ei_auto_doc")
DOCS_ROOT = ROOT / "ei docs for analysis" / "EI docs"
TEMP = DOCS_ROOT / "temp"
OUT_XLSX = TEMP / "params_selected_curated.xlsx"

PARAMS = [
    "ERDAT",
    "CHANGENR",
    "TAB_DESC",
    "ACT_CHNGNO",
    "CHANGE_IND",
    "CHANGE_IND_DESC",
    "CHNGIND",
    "CHNGIND_DESC",
    "OBJECTCLAS",
    "OBJECT_DESC",
    "PLANCHNGNR",
    "UNIT_NEW",
    "UNIT_OLD",
    "MESSAGE",
    "OBJECT",
    "AUFNR",
]

CURATED = {
    "ERDAT": "ERDAT is the record creation date and scopes monitoring to items created in the selected period.",
    "CHANGENR": "CHANGENR is the change document number used to trace one business change event end-to-end.",
    "TAB_DESC": "TAB_DESC provides short table text so technical table names remain business-readable in monitoring output.",
    "ACT_CHNGNO": "ACT_CHNGNO stores the active change document number used to correlate current change-header processing records.",
    "CHANGE_IND": "CHANGE_IND marks whether the application object was inserted, changed, or deleted for header-level change analysis.",
    "CHANGE_IND_DESC": "CHANGE_IND_DESC provides business-readable text for CHANGE_IND codes to simplify interpretation of change states.",
    "CHNGIND": "CHNGIND marks row-level operation type (insert, update, delete) within change document item details.",
    "CHNGIND_DESC": "CHNGIND_DESC provides business-readable text for CHNGIND values in change item reporting.",
    "OBJECTCLAS": "OBJECTCLAS identifies the change-document object class, scoping records to a specific SAP business object.",
    "OBJECT_DESC": "OBJECT_DESC provides the descriptive name of the referenced object so technical keys are understandable.",
    "PLANCHNGNR": "PLANCHNGNR stores the planning change number used to track plan-version adjustments across updates.",
    "UNIT_NEW": "UNIT_NEW stores the post-change unit of measure for quantitative fields in change comparison.",
    "UNIT_OLD": "UNIT_OLD stores the pre-change unit of measure for before/after comparison of quantity values.",
    "MESSAGE": "MESSAGE contains returned message text that explains processing outcomes, warnings, and error details.",
    "OBJECT": "OBJECT identifies the relevant SAP object key or type used to scope object-level monitoring records.",
    "AUFNR": "AUFNR identifies the internal order number and scopes records to specific Controlling order activity.",
}

RE_NUM = re.compile(r"^\d+$")
RE_PARAM = re.compile(r"^[A-Z][A-Z0-9_]{1,}$")
RE_WORD = re.compile(r"[A-Za-z0-9_]+")
PARAM_HEADER_CANDIDATES = ("field", "parameter", "field name", "name")


def iter_block_items(parent: DocumentType):
    body = parent.element.body
    for child in body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


def extract_evidence(path: Path, targets: set[str]) -> dict[str, dict[str, Counter[str] | int]]:
    doc = Document(str(path))
    out: dict[str, dict[str, Counter[str] | int]] = {}
    in_abap = False
    saw_param_ref_heading = False

    for block in iter_block_items(doc):
        if isinstance(block, Paragraph):
            t = " ".join((block.text or "").split())
            low = t.lower()
            if low == "parameters reference table":
                saw_param_ref_heading = True
                continue
            if low == "abap code":
                in_abap = True
                continue
            if in_abap:
                words = {w.upper() for w in RE_WORD.findall(t)}
                for p in targets:
                    if p in words:
                        out.setdefault(p, {"desc": Counter(), "de": Counter(), "dom": Counter(), "typ": Counter(), "abap_hits": 0})
                        out[p]["abap_hits"] = int(out[p]["abap_hits"]) + 1  # type: ignore[index]
                if "ENDFUNCTION" in t.upper():
                    in_abap = False
            continue

        if not isinstance(block, Table):
            continue

        rows = [[" ".join((c.text or "").split()) for c in row.cells] for row in block.rows]
        if not rows:
            continue
        header = [h.lower() for h in rows[0]]

        param_col = None
        for h in PARAM_HEADER_CANDIDATES:
            if h in header:
                param_col = header.index(h)
                break
        if param_col is None:
            continue
        if "structure name" in header:
            continue
        if not saw_param_ref_heading:
            continue

        col_desc = header.index("description") if "description" in header else None
        col_type = header.index("type") if "type" in header else None
        col_de = header.index("data element") if "data element" in header else None
        col_dom = header.index("domain") if "domain" in header else None
        col_num = None
        if "#" in rows[0]:
            col_num = rows[0].index("#")
        elif "no." in header:
            col_num = header.index("no.")
        elif "no" in header:
            col_num = header.index("no")

        for rr in rows[1:]:
            if param_col >= len(rr):
                continue
            p = rr[param_col].strip().upper()
            if p not in targets or not RE_PARAM.match(p):
                continue
            if col_num is not None and col_num < len(rr):
                n = rr[col_num].strip()
                if n and not RE_NUM.match(n):
                    break

            out.setdefault(p, {"desc": Counter(), "de": Counter(), "dom": Counter(), "typ": Counter(), "abap_hits": 0})
            if col_desc is not None and col_desc < len(rr) and rr[col_desc]:
                out[p]["desc"][rr[col_desc].strip()] += 1  # type: ignore[index]
            if col_de is not None and col_de < len(rr) and rr[col_de]:
                out[p]["de"][rr[col_de].strip()] += 1  # type: ignore[index]
            if col_dom is not None and col_dom < len(rr) and rr[col_dom]:
                out[p]["dom"][rr[col_dom].strip()] += 1  # type: ignore[index]
            if col_type is not None and col_type < len(rr) and rr[col_type]:
                out[p]["typ"][rr[col_type].strip()] += 1  # type: ignore[index]

    return out


def main() -> None:
    targets = set(PARAMS)
    files = []
    for part in ("Part 1", "Part 2", "Part 3", "Part 4"):
        files.extend(sorted((DOCS_ROOT / part).glob("*.docx")))

    files_seen: dict[str, set[str]] = {p: set() for p in PARAMS}
    desc = {p: Counter() for p in PARAMS}
    de = {p: Counter() for p in PARAMS}
    dom = {p: Counter() for p in PARAMS}
    typ = {p: Counter() for p in PARAMS}
    abap_hits = {p: 0 for p in PARAMS}

    for fp in files:
        ev = extract_evidence(fp, targets)
        for p in ev:
            files_seen[p].add(fp.name)
            desc[p].update(ev[p]["desc"])  # type: ignore[arg-type]
            de[p].update(ev[p]["de"])  # type: ignore[arg-type]
            dom[p].update(ev[p]["dom"])  # type: ignore[arg-type]
            typ[p].update(ev[p]["typ"])  # type: ignore[arg-type]
            abap_hits[p] += int(ev[p]["abap_hits"])  # type: ignore[arg-type]

    wb = Workbook()
    ws = wb.active
    ws.title = "curated_dictionary"

    ws["A1"] = "Summary"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "Parameters requested"
    ws["B2"] = len(PARAMS)
    ws["A3"] = "DOCX scanned"
    ws["B3"] = len(files)
    ws["A4"] = "Generated entries"
    ws["B4"] = len(PARAMS)

    header_row = 8
    headers = ["parameter", "curated_explanation", "confidence", "evidence_notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(header_row, i, h).font = Font(bold=True)

    r = header_row + 1
    for p in PARAMS:
        conf = "high" if len(files_seen[p]) >= 4 else "medium"
        top_desc = desc[p].most_common(1)[0][0] if desc[p] else ""
        top_de = de[p].most_common(1)[0][0] if de[p] else ""
        top_dom = dom[p].most_common(1)[0][0] if dom[p] else ""
        note = f"files={len(files_seen[p])}; abap_hits={abap_hits[p]}; desc='{top_desc}'; de='{top_de}'; dom='{top_dom}'"
        ws.cell(r, 1, p)
        ws.cell(r, 2, CURATED[p])
        ws.cell(r, 3, conf)
        ws.cell(r, 4, note)
        r += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 110
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 120

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
