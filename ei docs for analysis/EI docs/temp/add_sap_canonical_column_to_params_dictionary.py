"""
Fill column C 'SAP canonical explanation' on params_dictionary.xlsx for every data row.

Population order:
1. sap_canonical_attached_params.SAP_CANONICAL_ATTACHED (curated attached batch)
2. tools/sap_unified_param_texts.SAP_UNIFIED_EXPLANATION (unified DDIC-oriented strings)
3. build_params_dictionary_xlsx.auto_explain (includes OVERRIDES + pattern fallbacks) — never leaves blank when column A has a parameter.
"""

from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


ROOT = Path(r"c:\vibe code dev\ei_auto_doc")
TEMP = ROOT / "ei docs for analysis" / "EI docs" / "temp"
DICT_XLSX = TEMP / "params_dictionary.xlsx"


def _load_sap_canonical_attached() -> dict[str, str]:
    path = TEMP / "sap_canonical_attached_params.py"
    spec = importlib.util.spec_from_file_location("sap_canonical_attached_params", path)
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return {str(k).upper(): str(v).strip() for k, v in mod.SAP_CANONICAL_ATTACHED.items()}


def _load_sap_unified_explanations() -> dict[str, str]:
    path = ROOT / "tools" / "sap_unified_param_texts.py"
    spec = importlib.util.spec_from_file_location("sap_unified_param_texts", path)
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return {str(k).upper(): str(v).strip() for k, v in mod.SAP_UNIFIED_EXPLANATION.items()}


def _load_dictionary_builder():
    path = TEMP / "build_params_dictionary_xlsx.py"
    spec = importlib.util.spec_from_file_location("build_params_dictionary_xlsx", path)
    mod = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(mod)
    return mod


def main() -> None:
    if not DICT_XLSX.exists():
        raise FileNotFoundError(str(DICT_XLSX))

    attached = _load_sap_canonical_attached()
    unified = _load_sap_unified_explanations()
    bd = _load_dictionary_builder()
    auto_explain = bd.auto_explain

    wb = load_workbook(DICT_XLSX)
    ws = wb["dictionary"] if "dictionary" in wb.sheetnames else wb[wb.sheetnames[0]]

    ws["C1"] = "SAP canonical explanation"
    ws["C1"].font = Font(bold=True)

    filled_attached = 0
    filled_unified = 0
    filled_generated = 0
    blank_param_rows = 0

    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if not raw:
            ws.cell(r, 3, "")
            blank_param_rows += 1
            continue
        p = str(raw).strip().upper()
        canon_attached = attached.get(p, "")
        canon_unified = unified.get(p, "")
        if canon_attached:
            val = canon_attached
            filled_attached += 1
        elif canon_unified:
            val = canon_unified
            filled_unified += 1
        else:
            val = auto_explain(p)
            filled_generated += 1
        ws.cell(r, 3, val)

    ws.column_dimensions["C"].width = 130
    wb.save(DICT_XLSX)

    nonempty = 0
    wb2 = load_workbook(DICT_XLSX, read_only=True)
    ws2 = wb2["dictionary"] if "dictionary" in wb2.sheetnames else wb2[wb2.sheetnames[0]]
    for r in range(2, ws2.max_row + 1):
        if ws2.cell(r, 1).value and not (str(ws2.cell(r, 3).value or "").strip()):
            raise RuntimeError(f"Canonical column still empty for row {r} parameter {ws2.cell(r,1).value!r}")
        if ws2.cell(r, 1).value:
            nonempty += 1

    print(f"Wrote column C on {DICT_XLSX}")
    print(f"Data rows (non-blank parameter): {nonempty}")
    print(f"Filled from attached canonical: {filled_attached}")
    print(f"Filled from unified canonical: {filled_unified}")
    print(f"Filled via auto_explain (OVERRIDES + patterns + default): {filled_generated}")
    print(f"Blank parameter rows (C left empty): {blank_param_rows}")
    print("Verification: no empty canonical text when column A is present.")


if __name__ == "__main__":
    main()
